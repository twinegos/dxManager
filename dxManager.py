# encoding:utf-8

# 표준 라이브러리 임포트
import os
import sys
import time
import json
import subprocess
from datetime import datetime, timedelta
from multiprocessing import Process
from types import MethodType

# 서드파티 라이브러리 임포트
import pandas as pd
import openpyxl
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pymongo import MongoClient

# PySide2 imports - wildcard 방식으로 변경 (버전 호환성 문제 해결)
from PySide2.QtCore import *
from PySide2.QtWidgets import *  
from PySide2.QtGui import *

from PySide2 import QtWidgets

# 설정 파일 임포트
import config

# 클래스 모듈 임포트

# 외부 라이브러리 임포트
import dxConfig
import requests
# 데이터베이스 연결 설정
DBIP = dxConfig.getConf("DB_IP")
client = MongoClient(DBIP)
# KEEP: 데이터베이스 테스트 코드 - 삭제 금지
#db = client['SCHEDULE_test']
#collection = db['test']
#collection.insert_one({'a':1})
#client.drop_database('SCHEDULE_test')



# 상수 정의
#TACTIC
TATIC_API_KEY = config.TACTIC_API_KEY

currentPath = config.CURRENT_PATH

# 로컬 모듈 경로 설정
# UI 모듈 임포트
UI_path = config.UI_PATH
sys.path.append(UI_path)
# UI 모듈 imports
import leadMainWindow_UI
import dayScheduleDyn_UI
import dashboard_UI
import teamMemberData_UI
import memberDataInfo_UI
import darkTheme 



# 클래스 모듈 임포트
classPath = currentPath + "/class"
uiPath = currentPath + "/UI"
sys.path.append(classPath)
sys.path.append(uiPath)
# 클래스 모듈 imports
import dragDropListView as ddlv
from sort_manager import SortManager 
import editMemberDialog as memberDialog
import taskInfoDialog
import status_projManday
import proj_Manday
import status_teamManday
import loadingProgress
import leadMainWindow
import Edit_mandays
import scheduleListview as lv
import multiColorProgressBar as mpb
import progressBar_Thread as pb
import updateUserInfo_Thread as updateUser
import readOnlyDelegate as readOnly
import loading_workData as loadData
from tactic_api_client import TacticAPIClient
from data_manager import DataManager
from ui_controller import UIController
from team_manager import TeamManager
from event_manager import EventManager
from ui_layout_manager import UILayoutManager
from validation_manager import ValidationManager
from file_manager import FileManager

userID = config.get_user_id()

# TACTIC API 클라이언트 초기화
tactic_client = TacticAPIClient()

# 데이터 매니저 초기화
data_manager = DataManager(userID)

# JSON 스케줄 파일 임포트
jsonData = []
jsonFilePath = config.SCHEDULE_DATA_PATH
jsonFile = config.get_schedule_data_file(userID)


# 전역 변수
dropData = {} # 드래그 드랍으로 받은 데이타와 받은 날짜
teamTreeHierarchy = {} # 팀원과 팀원의 하위멤버의 하이라키 구조를 저장
teamInfo_ = [] # 현재유저 하위의 모든 팀원의 정보를 저장
gradeScore = {"상":100, "중상":80, "중":60, "중하":40, "하":20}
connect_shot_proj = {} # 프로세스내 사용되는 샷리스트와 프로젝트 매칭을 위한 전역변수
connect_shot_schedule = {} # 프로세스내 사용되는 샷리스트와 스케쥴 매칭을 위한 전역변수
sortOrder_date = {} # 스케쥴 리스트뷰의 소팅오더를 저장
sortColumn_date = {} # 스케쥴 리스트뷰의 소팅한 최종 컬럼이 어떤것인지 저장
DOUBLE_CLICK_DELAY = 110 # 멤버 트리뷰 더블클릭 간격






# UILayoutManager 클래스



# 메인 클래스
class DxManager(QMainWindow):
    def __init__(self,parent=None):
        QMainWindow.__init__(self,parent)
       
        self.manager = lv.ActiveListviewManager()

        self.memberCache_ = {} # 태스크정보를 읽어온 유저들의 한글이름과 영어이름을 짝지어 저장
        self.deadLine_flag = 0
        self.memberEN  = []
        self.taskInfoJson = {}
        self.sort_order = Qt.AscendingOrder
        self.sort_column = 1
        self.comboBoxDate = []
        self.rangeEndDate = []
        self.Q_leader_app_task = 0
        self.select_ctxMenu = 0 # 멤버 트리뷰의 컨텍스트메뉴 실행시 사용
        self.reload_tatic = 0
        self.listview_minimum = 310
        self.currentProjs = [] # 현재 선택된 프로젝트들

        self.MT_reloadShotList = 0
        self.MT_selAll_thread = 0

        # 팀 관리자 초기화 (checkUserInfo 호출 전에 초기화 필요)
        self.team_manager = TeamManager(self)
        

        # UI 레이아웃 관리자 초기화
        self.ui_layout_manager = UILayoutManager(self)

        # FileManager 클래스 초기화
        self.file_manager = FileManager(self)

        self.currName_kr, self.role, self.team, self.job, self.department = self.checkUserInfo(userID) # 현재유저의 정보 가져오기


        self.contextFix = {} # 잘못 저장된 컨텍스트를 수정하기 위한 임시 변수


        # treeView 의 selectAll의 쓰레드풀 초기화 ########################################
        self.threadpool = QThreadPool()
        self.threadpool.setMaxThreadCount(QThread.idealThreadCount())
        self.signals = loadData.LoadWorkDataSignals()        
        #################################################################################


        self.ui = leadMainWindow_UI.Ui_MainWindow()

        self.dayUi = dayScheduleDyn_UI.Ui_daySchedule_frame()

        self.dashboardUi = dashboard_UI.Ui_dashboard_frm()
        self.teamMemberUi = teamMemberData_UI.Ui_memberData_frm()
        self.memberDataInfoUi = memberDataInfo_UI.Ui_workData_frm()



        self.ui.setupUi(self)


        icon = QIcon()
        icon.addFile(config.REFRESH_ICON, QSize(), QIcon.Normal, QIcon.Off)
        self.ui.reload_Button.setIcon(icon)

        self.ui.splitter_2.setSizes([200,300])
        self.dataInfoBtn = self.teamMemberUi.workData_btn
        self.dayScheduleFrame = QFrame()

        self.dayUi.setupUi(self.dayScheduleFrame)
        self.dayUi.splitter.setChildrenCollapsible(False)
        self.dayUi.horizontalLayout_3.setAlignment(Qt.AlignRight)
        self.dayUi.horizontalLayout_2.setAlignment(Qt.AlignLeft)
        self.dayUi.horizontalLayout_6.setAlignment(Qt.AlignRight)

        self.dashboardFrame = self.dashboardUi.dashboard_frm
        self.ui.mainInfo_layout.addWidget(self.dayScheduleFrame)
        self.ui.mainInfo_layout.addWidget(self.dashboardFrame)
        self.dashboardFrame.setVisible(False)       

        self.memberDataFrame = self.teamMemberUi.memberData_frm
        self.dashboardUi.scrollContents_layout.addWidget(self.memberDataFrame, 0, Qt.AlignHCenter)

        self.ui.line.setStyleSheet("QFrame { background-color: #303030; height: 1px; border: none}")

        self.ui.menuPreferences.setTitle("Edit")
        self.ui.pushButton_3.setText("Export")
        self.ui.pushButton_3.setEnabled(False)


        self.resize(1700, 900)

        # UI 변수 선언
        ##############################################################
        self.menu_Edit = self.ui.actionEdit_member


        self.teamTree = self.ui.team_treeView
        self.preferenceMenu = self.ui.menuPreferences

        self.leftSideInfoFrm = self.ui.leftSide_frame

        self.artistLabel = self.ui.artistNameLabel
        self.artistLabel_EN = self.ui.artistNameLabel_EN
        self.teamLabel = self.ui.teamLabel

        self.projListview = self.ui.projectListView
        self.shotListview = self.ui.shotListView

        self.scheduleViewBtn = self.ui.scheduleView_btn
        self.artistFrame = self.ui.frame_2
        
        # UI 컨트롤러 초기화
        self.ui_controller = UIController(self)
        
        # EventManager 초기화
        self.event_manager = EventManager(self)

        # SortManager 초기화
        self.sort_manager = SortManager(self)

        # ValidationManager 초기화
        self.validation_manager = ValidationManager(self)


        # KEEP: 내용정리 될때까지 하이드시키기 #################
        self.dashboardViewBtn = self.ui.dashboard_btn
        self.dashboardViewBtn.hide()
        ################################################

        self.forwardDayBtn = self.dayUi.forwardDayButton
        self.backwardDayBtn = self.dayUi.backwardDayButton
        self.day_listViewLayout = self.dayUi.scrollAreaLayout
        self.todayBtn = self.dayUi.Button_today

        self.listViewNumLineEdit = self.dayUi.listViewNum_lineEdit
        self.listViewSlider = self.dayUi.listViewNumSlider

        self.listViewNumLineEdit.setReadOnly(True)

        self.listViewScroll = self.dayUi.scrollArea

        self.startDay = self.dayUi.start_day
        self.startMonth = self.dayUi.start_month

        self.saveButton = self.ui.saveButton



        ### 사용자에 따라 ui 조정 ############################################

        self.scheduleViewBtn.hide()
        self.dashboardViewBtn.hide()

        #팀원의 경우 프리퍼런스메뉴,스케쥴버튼,대시보드,팀트리뷰창을 숨김
        if self.job == "팀원" :
            self.preferenceMenu.menuAction().setVisible(False)

        ######################################################################


        # 아티스트 라벨 설정
        self.teamLabel.setText(self.team)
        self.artistLabel.setText(self.currName_kr)
        self.artistLabel_EN.setText(userID)

        # 프로젝트 리스트뷰 설정 
        self.projects = getProjectList()
        self.proj_model = QStandardItemModel()

        for x in self.projects:
            self.proj_model.appendRow(QStandardItem(self.projects[x][1]))#x.upper()))
        self.projListview.setModel(self.proj_model)
        self.projSelection_model = self.projListview.selectionModel()
        self.projListview.setEditTriggers(QAbstractItemView.NoEditTriggers) # 리스트뷰내 아이템 편집되지 않게 함

        # 팀 트리뷰 모델 설정 
        self.teamTreeModel = QStandardItemModel()
        self.teamTreeModel.setColumnCount(1)
        self.teamTree.setModel(self.teamTreeModel)


        # 트리뷰의 멤버들을 선택하거나 포커싱되지 않도록 함
        self.teamTree.setSelectionMode(QAbstractItemView.SingleSelection)


        self.teamLeader = QStandardItem(self.currName_kr) 
        self.teamLeader.setCheckable(True)

        self.invisibleRootItem = self.teamTreeModel.invisibleRootItem()
        self.invisibleRootItem.appendRow(self.teamLeader)


        # 트리뷰 모델의 헤더 이름 설정
        self.teamTreeModel.setHorizontalHeaderLabels([self.team])

        # 날짜 콤보박스 설정
        self.comboDay = self.dayUi.comboBox_day
        self.comboMonth = self.dayUi.comboBox_month
        self.comboYear = self.dayUi.comboBox_year
        self.comboYear.setStyleSheet ( "QComboBox { font-size: 15px; text-align: right;}")
        self.comboYear.setFixedWidth(100)

        currentYear, currentMonth, currentDay, currentDow, years, allMonths, allDays = self.setDate()
        self.comboDay.addItems(allDays)
        self.comboDay.setEditable(True)
        self.comboDay.lineEdit().setAlignment(Qt.AlignCenter)        
        self.comboDay.lineEdit().setReadOnly(True)                
        self.comboDay.setCurrentText(str(currentDay))

        self.comboMonth.addItems(allMonths)
        self.comboMonth.setEditable(True)
        self.comboMonth.lineEdit().setAlignment(Qt.AlignCenter)        
        self.comboMonth.lineEdit().setReadOnly(True)                
        self.comboMonth.setCurrentText(str(currentMonth))        

        self.comboYear.addItems(years)
        self.comboYear.setCurrentText(str(currentYear))

        self.comboBoxDate.append(self.comboYear.currentText())
        self.comboBoxDate.append(self.comboMonth.currentText())
        self.comboBoxDate.append(self.comboDay.currentText())

        # EventManager 연결 설정 (모든 UI 요소가 정의된 후)
        self.event_manager.setup_connections()

        # 리스트뷰 UI 설정
        self.listViewNumLineEdit.setText(str(self.listViewSlider.value()))

        # 이전에 저장된 스케쥴의 제이슨파일을 읽어와서  전역변수 jsonData 에 저장
        # 팀원 제이슨 정보를 읽어와서 ui에 반영시킬수 있도록 함
        self.getJsonData() 

        # KEEP: MongoDB 관련 코드는 database_manager.py로 이동됨
        # from database_manager import DatabaseManager
        # self.db_manager = DatabaseManager()
        # self.db_manager.connect()


        # 컨트롤러 속성 연결
 
        self.saveButton.clicked.connect(self.saveFile)        
        self.teamTreeModel.dataChanged.connect(self.handle_item_state_change)


        # 멤버트리뷰의 클릭/더블클릭 겹침방지를 위한 타이머 객체 설정 ###################
        self.click_timer = QTimer()
        self.click_timer.setSingleShot(True)
        self.pending_click = None
        self.checkArea_click = 0 # 클릭한 위치가 체크박스영역인지 텍스트 영역인지 확인
        ##############################################################################


        self.memberStatusDialog = None # 다이얼로그 객체 초기화
        self.teamTree.clicked.connect(self.on_item_clicked) # 체크박스를 정확히 클릭하지 않고 이름을 클릭해도 태스크 리스트를 읽어오도록 함.
        self.teamTree.doubleClicked.connect(self.show_status_manday)        

        self.ui.deadlineButton.clicked.connect(self.deadLine_filter) # 이번주 완료예정인 태스크만 보이도록 함
        self.ui.reload_Button.clicked.connect(self.referesh_manager) # 전체 새로고침

        self.click_timer.timeout.connect(self.singleClick_select_member)


        self.taskInfomationDialog = None # 다이얼로그 객체 초기화
        self.editMandays = None # 맨데이 편집창 객체 초기화

        self.shotListview.doubleClicked.connect(lambda index: self.openMedia(self.shotListview, index))

        artist = self.artistLabel.text()



        self.sliderTimer = QTimer()
        self.sliderTimer.setInterval(0)
        self.sliderTimer.setSingleShot(True)
        self.last_value = self.listViewSlider.value()


        self.listViewSlider.valueChanged.connect(self.setListviewLineEdit)

        self.listViewSlider.sliderReleased.connect(self.changeListViewUI_notTracking)

        # 초기 기본 5개의 리스트뷰 생성
        self.setUp_listViewUI()

        # 각 날짜의 데이타가 아직 생성되지 않은 초기상태에서 초기 json 파일 생성
        if not jsonData:
            self.makeInit_data(currentPath, userID)

        # 샷 리스트뷰에 컨텍스트 메뉴 연결
        self.shotListview.setContextMenuPolicy(Qt.CustomContextMenu)
        self.shotListview.customContextMenuRequested.connect(
            lambda pos, lv=self.shotListview: self.show_context_menu(lv, pos))

        # 팀 트리뷰에 컨텍스트 메뉴 연결
        self.teamTree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.teamTree.customContextMenuRequested.connect(self.member_context_menu)

        # 오늘날짜로 스케쥴 리스트뷰 갱신
        self.todayBtn.clicked.connect(self.move_TodayListview)

        # 리스트뷰의 뷰필터 
        self.dayUi.checkBox_approved_v.stateChanged.connect(self.listView_filter)
        self.dayUi.checkBox_inprogress_v.stateChanged.connect(self.listView_filter)
        self.dayUi.checkBox_ready_v.stateChanged.connect(self.listView_filter)
        self.dayUi.checkBox_review_v.stateChanged.connect(self.listView_filter)
        self.dayUi.checkBox_hold_v.stateChanged.connect(self.listView_filter)
        self.dayUi.checkBox_ok_v.stateChanged.connect(self.listView_filter)
        self.dayUi.checkBox_wait_v.stateChanged.connect(self.listView_filter)
        self.dayUi.checkBox_omit_v.stateChanged.connect(self.listView_filter)
        self.dayUi.checkBox_retake_v.stateChanged.connect(self.listView_filter)


        # 진급, 및 부서이동으로 인한 개인정보 변경시 이를 적용하기 위한 쓰레드 실행
        members = list(self.memberCache_.keys())
        self.updateUserInfoThread = updateUser.updateUserInfoThread(members, self)
        self.updateUserInfoThread.start()


        # 프로젝트를 선택하지 않아도 각 날짜 리스트뷰에 태스크들이 그대로 남아있게 하는옵션 - 차후 업데이트
        # 차후에 난이도와 스타트/엔드데이트 컬럼 추가한후 진행예정
        self.ui.showAllTask_checkBox.setEnabled(False)

        # ui가 렌더링된 이후에 스크롤바 조정이 가능하기 때문에, 렌더링후 스크롤바 조정 메서드 실행
        QTimer.singleShot(100, self.moveScrollBar)


        def teamTree_mousePress(self, event):
            super(QTreeView, self).mousePressEvent(event)

            # self.teamTree 로 전달되어야 할 마우스 클릭 이벤트가 계속 미리소비되어져서, 
            # 마우스클릭 이벤트를 eventFilter 로 강제전달함
            QApplication.sendEvent(self, event) 


        # MethodType 메서드로 teamTree_mousePress 메서드를 self.teamTree 객체의 메서드로 동적 할당
        # QTreeView의 기본 mousePressEvent를 교체하여 MouseButtonPress 처리 강제
        self.teamTree.mousePressEvent = MethodType(teamTree_mousePress, self.teamTree)    

        # self.teamTree 이벤트필터 설치
        self.teamTree.installEventFilter(self)

        self.teamLeader.setCheckState(Qt.Checked)






        #self.makeMirrorData()

        # DEPRECATED: DB 마이그레이션 시 제거 예정 - 미러 데이터 스레드 셋업
        # MongoDB에서는 트랜잭션과 관계형 구조로 대체됨
        # 스케쥴 미러데이터를 만들기 위한 쓰레드 셋업 - 팀장의 미러데이타가 저장됳때 팀원의 미러데이타도 함께 업데이트 되게 하여야함
        #self.makeMirrorThread = WorkerThread_makeMirror(jsonData, self)
        #self.makeMirrorThread.finished.connect(self.makeMirrorFinished)
        #self.makeMirrorThread.error.connect(self.makeMirrorError)

        #self.makeMirrorThread.start()




    ###############################################################################################################################################################
    ###############################################################################################################################################################
    ###############################################################################################################################################################
    ###############################################################################################################################################################



    # 태스크리스트뷰 혹은 스케쥴 리스트뷰의 태스크를 더블클릭했을때, 현재 작업중인 최신프리뷰 실행
    def openMedia(self, listview, index):

        task_name = ""
        context_name = ""
        part = ""
        works_path = ""

        item = listview.model().data(index, Qt.DisplayRole)

        if "/" in item[1]:
            task_name = item[1].split('/')[0]
            context_name = item[1].split('/')[1]
    
        else:
            task_name = item[1]
            context_name = None

        if item[4] == "A": part = "animation"
        elif item[4] == "M": part = "matchmove"
        elif item[4] == "R" : part = "creature"

        proj_upp = item[0]

        proj_code = self.projects[proj_upp.lower()][0]
     

        default_path = config.get_tactic_asset_path(proj_code)

        asset_worksPath = default_path +"/asset/"+task_name+"/"+part


        if os.path.exists(asset_worksPath) == False:

            shot_worksPath = default_path +"/shot/"+task_name+"/"+part
            if os.path.exists(shot_worksPath) == True:

                if context_name == None:
                    if part=="animation": context_name = "animation"
                    elif part =="matchmove": context_name = "matchmove"
                    elif part =="creature": context_name = "creature"
                shot_worksPath = shot_worksPath+"/"+context_name

                if os.path.exists(shot_worksPath):
                    self.file_manager.get_latest_file(shot_worksPath)

                elif os.path.exists(shot_worksPath) == False:
                    QMessageBox.information(None, "알 림", "프리뷰가 존재하지 않습니다.")



            elif os.path.exists(shot_worksPath) == False:

                edit_path = default_path + "/shot/"+task_name+"/publish/edit/"

                if os.path.exists(edit_path) == True:
                    self.file_manager.get_latest_file(edit_path)

                else:
                    QMessageBox.information(None, "알 림", "프리뷰가 존재하지 않습니다.")



        elif os.path.exists(asset_worksPath) == True:

            if context_name == None:
                if part=="animation": context_name = "animation"
                elif part =="matchmove": context_name = "matchmove"
                elif part =="creature": context_name = "creature"

            asset_worksPath = asset_worksPath+"/"+context_name

            if os.path.exists(asset_worksPath):
                self.file_manager.get_latest_file(asset_worksPath)




















    # ShotListView 및 Schedule ListView의 정보를 Tactic의 최신 정보로 업데이트
    def referesh_manager(self):
        
        self.reload_tatic = 1

        self.reloadShotList()
        self.updateJson(0)

        self.reload_tatic = 0



    # UI가 렌더링된 이후에 스크롤바 조정이 가능하기 때문에, 렌더링 후 스크롤바 조정하는 메서드
    def moveScrollBar(self):

        numberListView = int(self.listViewNumLineEdit.text())

        today = QDate.currentDate()
        monday = today.addDays(-today.dayOfWeek()+1)
        friday = monday.addDays(4) 


        dow = today.dayOfWeek()
        scrollWidth = self.listViewScroll.horizontalScrollBar().maximum()

        if dow == 1 or dow == 2:
            self.listViewScroll.horizontalScrollBar().setValue(0)

        elif dow == 3:
            self.listViewScroll.horizontalScrollBar().setValue(scrollWidth/2)

        elif dow == 4 or dow == 5:
            self.listViewScroll.horizontalScrollBar().setValue(scrollWidth)









    # josnData   확인해야함.
    def makeMirrorSchedule(self, jsonData_):

        members = list(self.memberCache_.keys())

        # { 프로젝트: {태스크1:파트, 태스크2:파트} } 형태의 딕셔너리
        tasks = []
        proj_tasks = {}
        for data in jsonData_:
            if data['tasks'] != [{}] or data['tasks'][0] != []:
                for proj in list(data['tasks'][0].keys()):
                    for task in data['tasks'][0][proj]:
                        if proj in proj_tasks:

                            proj_tasks[proj][task] =  data['department'] 

                        elif proj not in proj_tasks:
                            proj_tasks[proj] = {}
                            proj_tasks[proj][task] =  data['department'] 



        #task_resultData, mirror_Dic = self.convert_to_resultData_forMirror(tasks, members)
        task_resultData, mirror_Dic = self.convert_to_resultData_forMirror(proj_tasks, members)


        if os.path.exists(config.get_schedule_mirror_file(userID)):

            savedMirrorData = self.file_manager.import_Json_Thread(config.SCHEDULE_MIRROR_PATH, userID+"_scheduleMirror.json")

            del_sav_task = {} # 미러데이터를 만드는 경우 이미 미러데이터 제이슨 파일에 같은 이름의 데이타가 있으면 삭제 <-- 맨데이나 스테이터스등 수정된 최신데이타로 바꾸기 위해
            add_mir_task = {} # 삭제된 이전 미러 데이타를 대체하여 들어갈 최신 미러데이터 딕셔너리

            for m_proj in mirror_Dic:
                if m_proj not in savedMirrorData:
                    savedMirrorData[m_proj] = mirror_Dic[m_proj]

                elif m_proj in savedMirrorData:

                    del_tasks = [] 
                    add_tasks = []

                    for m_task in mirror_Dic[m_proj]:
                        for sav_task in savedMirrorData[m_proj]:
                            if sav_task[1]==m_task[1] and sav_task[4]==m_task[4]:
                                del_tasks.append(sav_task)
                                add_tasks.append(m_task)

                    del_sav_task[m_proj] = del_tasks
                    add_mir_task[m_proj] = add_tasks


            if del_sav_task:
                for proj in del_sav_task:
                    for i in range(len(del_sav_task[proj])):
                        savedMirrorData[proj].remove(del_sav_task[proj][i])
                        if add_mir_task[proj][i] not in savedMirrorData[m_proj]:
                            savedMirrorData[proj].append(add_mir_task[proj][i])


            self.file_manager.export_Json_Thread(config.SCHEDULE_MIRROR_PATH, userID+"_scheduleMirror.json", savedMirrorData)

            # 미러데이타 백업
            backup_dir = "mirrorData"
            backup_name = userID +"_scheduleMirror"


        else:
            self.file_manager.export_Json_Thread(config.SCHEDULE_MIRROR_PATH, userID+"_scheduleMirror.json", mirror_Dic)            

            # 미러데이타 백업
            backup_dir = "mirrorData"
            backup_name = userID +"_scheduleMirror"
            self.file_manager.backup_schedule(mirror_Dic, userID, backup_dir, backup_name)



                



    def updateUserInfo(self, members):

        name_kr = ""
        role = ""
        team = ""
        job = ""
        department = ""        

        for member in members:
            name_kr, role, team, job, department = getUserInfo(member)

            if os.path.exists(config.get_user_info_file(member)):
                memberUserInfo = self.file_manager.import_Json(config.USER_INFO_PATH, member+"_userInfo.json")

            if (memberUserInfo[0]["role"] != role) or (memberUserInfo[0]["team"] != team) or (memberUserInfo[0]["job"] != job) or (memberUserInfo[0]["department"] != department):

                memberUserInfo[0]["role"] = role
                memberUserInfo[0]["team"] = team
                memberUserInfo[0]["job"] = job
                memberUserInfo[0]["department"] = department

                self.file_manager.export_Json(config.USER_INFO_PATH, member+"_userInfo.json", memberUserInfo)                

        print (" - User information updated successfully - ")






    def set_GradeScore(self, checked_list):

        top_lv = 100
        uppMid_lv = 80
        mid_lv = 60
        lowMid_lv = 40
        low_lv = 20

        #gradeScore = ["상":100, "중상":80, "중":60, "중하":40, "하":20]

        if (checked_list):
            lv_set = []
            for item in checked_list:
                grade = checked_list[item]
                lv_set.append(gradeScore[grade])            

            for grade in gradeScore:
                if gradeScore[grade] == max(lv_set):
                    return grade




    def get_taskScore(self,project, process, task, context, artist):

        rig_taskContexts = ["asset", "shot"]
        ani_taskContexts = ["aniClip", "cache", "shot"]
        mmv_taskContexts = ["lens"]

        contextFolders = ""
        if process == "animation":
            part = "ANI"
            contextFolders = ani_taskContexts

        elif process == "creature":
            part = "RIG"
            contextFolders = rig_taskContexts

        elif process == "matchmove":
            part = "MMV"
            contextFolders = mmv_taskContexts


        if context == None:
            context = process



        # rig실의 asset/shot 폴더를 검색하여 해당 태스크 찾기 ########
        processPath = config.get_show_process_path(project, part)

        checkList_path = ""
        for taskContext in contextFolders:
            if os.path.isdir(os.path.join(processPath, taskContext)):
                listDir = os.listdir(processPath + "/" + taskContext)
                for dir in listDir:
                    if os.path.isdir(processPath + "/" + taskContext + "/" + dir):
                        listFolders = os.listdir(processPath + "/" + taskContext + "/" + dir)
                        for folder in listFolders:
                            if folder == task:
                                checkList_path = processPath + "/" + taskContext + "/" + dir + "/" + task + "/" + context + "/.data/checkList" 





        checkList_file_name = artist + "_" + context + "_data.json"

        # 태스크의 난이도 정보
        checkList_data = self.file_manager.import_Json(checkList_path, checkList_file_name) 

        if os.path.exists(os.path.join(checkList_path, checkList_file_name)):
            checked_list = {}
            parentItem = ""

            self.get_checked_items(checkList_data[0]["children"], checked_list, parentItem)

            grade = self.set_GradeScore(checked_list)

            return grade


        else:
            grade = "X"
            return grade





    def get_checked_items(self,items, checked_list, parentItem):

        for item in items:
            if "children" in item:
                if item["checked"] == True:
                    parentItem = parentItem + item["name"]+"--"
                self.get_checked_items(item["children"], checked_list, parentItem)

            if "children" not in item:
                if item["checked"] == True:
                    category = parentItem+item["name"]
                    checked_list[category] = item["score"]                    






    def move_TodayListview(self):

        today = datetime.today()
        year = today.strftime("%Y")#-%m-%d")
        month = today.strftime("%m")
        day = today.strftime("%d")

        listview_num = self.listViewNumLineEdit.text()
        addDay = int(int(listview_num)/2)

        endDate = QDate(int(year), int(month), int(day)).addDays(addDay)
        
        endDate_year = QDate(endDate).year()
        endDate_month = QDate(endDate).month()
        endDate_day = QDate(endDate).day()

        self.comboDay.setCurrentText(str(endDate_day))
        self.comboMonth.setCurrentText(str(endDate_month))            
        self.comboYear.setCurrentText(str(endDate_year))

        # ui의 날짜 콤보박스에 반영된 날짜대로 리스트뷰 이동시키기
        self.changeComboDate()

        mm = self.listViewScroll.horizontalScrollBar().maximum()

        self.listViewScroll.horizontalScrollBar().setValue(mm/2)



    # 리더나 수퍼바이져 하위의 팀원을 모두 읽어와 리스트 반환
    def get_tree_member(self, member, teamMember):

        member_info = self.file_manager.import_Json(currentPath + "/.user_Info", member+"_userInfo.json")

        if member_info[0]["role"] == "Lead" or member_info[0]["role"] == "Supervisor" or member_info[0]["role"] == "Manager":

            if (os.path.exists(currentPath + "/.team_Info/"+member+"_teamInfo.json")):
                team_info = self.file_manager.import_Json(currentPath + "/.team_Info", member+"_teamInfo.json")  

                for mem in team_info[0]:
                    self.get_tree_member(mem, teamMember)

        teamMember.append(member)




    # 트리뷰에서 텍스트로 아이템 찾기
    def find_item_by_text(self, text):

        # 트리뷰에서 해당 아이템을 찾기위한 재귀함수
        def search_recursive(parent, target_text):
            for row in range(parent.rowCount()):
                child = parent.child(row)
                if child.text() == target_text:
                    return child
                if child.hasChildren():
                    result = search_recursive(child, target_text)                    
                    if result:
                        return result
            return None

        root = self.teamTreeModel.invisibleRootItem()
        return search_recursive(root, text)



    # 프로젝트 리스트 뷰 의 선택이 변경되면 샷리스트뷰를 새로고침
    def process_sel_project(self):

        # 프로젝트 리스트뷰의 선택변경이 팀트리뷰 컨텍스트메뉴에 의한것인지 확인
        if self.select_ctxMenu == 0:
            self.reloadShotList()
            self.refereshListViews(0, jsonData)
            #self.contextFix = {}  




    # 팀 트리 뷰 우클릭 메뉴
    def member_context_menu(self, position):

        index = self.teamTree.indexAt(position)

        if index.isValid():

            item = index.data()
            memberEN = [member for member in self.memberCache_ if self.memberCache_[member] == item][0]

            teamMember = []
            self.get_tree_member(memberEN, teamMember)

            memberEN = [member for member in self.memberCache_ if self.memberCache_[member] == item][0]
            member_info = self.file_manager.import_Json(currentPath + "/.user_Info", memberEN+"_userInfo.json")

            if member_info[0]["role"] == "Lead" or member_info[0]["role"] == "Supervisor" or member_info[0]["role"] == "Manager":
                if (os.path.exists(currentPath + "/.team_Info/"+memberEN+"_teamInfo.json")):

                    context_menu = QMenu(self)
                    selectAll = context_menu.addAction("Select All")
                    deselectAll = context_menu.addAction("Deselect All")

                    action = context_menu.exec_(self.teamTree.mapToGlobal(position))

                    QApplication.setOverrideCursor(Qt.WaitCursor)


                    if action == selectAll:

                        i=0 # 프로그레스바 진행률
                        self.select_ctxMenu = 1  # 프로젝트 리스트뷰에서 프로젝트가 선택되더라도 select_ctxMenu 가 1로 되어있으면 reloadShotList가 실행되지 않도록 함
                        num_Mem = len(teamMember)

                        for memEN in teamMember:
                            memKR = self.memberCache_[memEN]
                            selItem = self.find_item_by_text(memKR)
                            selItem.setCheckState(Qt.Checked)

                        if self.ui.sel_allProj_checkBox.checkState() == Qt.Checked:
                            self.reloadShotList() 

                        else:
                            self.reloadShotList()

                        self.select_ctxMenu = 0 # 다시 0으로 원위치 시켜서 프로젝트 리스트뷰의 선택이 변경되면 reloadShotList가 실행되도록함



                    elif action == deselectAll:

                        self.select_ctxMenu = 1    # 프로젝트 리스트뷰에서 프로젝트가 선택되더라도 select_ctxMenu 가 1로 되어있으면 reloadShotList가 실행되지 않도록 함                     
                        for memEN in teamMember:
                            memKR = self.memberCache_[memEN]
                            selItem = self.find_item_by_text(memKR)
                            selItem.setCheckState(Qt.Unchecked)

                        if self.ui.sel_allProj_checkBox.checkState() == Qt.Checked:
                            self.reloadShotList()
                            self.updateJson(0)

                        else:
                            self.reSelect_proj()
                            self.reloadShotList()                       

                        self.select_ctxMenu = 0   # 다시 0으로 원위치 시켜서 프로젝트 리스트뷰의 선택이 변경되면 reloadShotList가 실행되도록함

        QApplication.restoreOverrideCursor()






    # 샷 리스트 우클릭 메뉴
    #def show_context_menu(self, position):
    def show_context_menu(self, listview, position):

        index = listview.indexAt(position)

        if index.isValid():

            context_menu = QMenu(self)

            task_info = context_menu.addAction("Task Info")
            add_action = context_menu.addAction("Assign Schedule")
            del_action = context_menu.addAction("Delete Schedule")
            find_action = context_menu.addAction("Find Schedule")
            edit_mandays = context_menu.addAction("Edit Man-days")

            action = context_menu.exec_(listview.mapToGlobal(position))

            if action == task_info:
                # 태스크 정보 표시하기
                self.openTaskInfoWin(listview, index)


            if action == add_action:
                selected_indexes = listview.selectedIndexes()
                # 현재 샷 리스트에서 선택한 태스크들을 스케쥴에 맞게 리스트뷰에 배치시킴
                self.add_schedule(selected_indexes)                


            elif action == find_action:
                selected_indexes = listview.selectedIndexes()
                self.find_schedule(selected_indexes)


            elif action == del_action:
                selectedTasks = []
                selected_indexes = listview.selectedIndexes()
                self.delete_schedule(selected_indexes, listview)


            elif action == edit_mandays:
                QApplication.setOverrideCursor(Qt.WaitCursor)
                selectedTasks = []                
                selected_indexes = listview.selectedIndexes()
                self.edit_mandays(selected_indexes, listview)
                QApplication.restoreOverrideCursor()




    def edit_mandays(self, selected_indexes, currListview):

        selectedItems = []
        for index in selected_indexes:
            item = currListview.model().data(index, Qt.DisplayRole)            
            selectedItems.append(item)

        column = ["Grade", "Task name", "BD manday", "Status", "Part"]

        # 판다스 데이타 프레임 설정
        df_taskList = pd.DataFrame(selectedItems)

        # 데이타프레임 컬럼 이름 변경
        df_taskList = df_taskList.set_axis(column, axis="columns")

        # grade 열 삭제
        del df_taskList["Grade"]

        # 데이터프레임에 컬럼추가
        df_taskList["Project"] = "Unknown"
        df_taskList["Act manday"] = "Unknown"
        df_taskList["Note"] = ""


        # 컬럼 순서 조정
        new_order = ["Project", "Part", "Task name", "Act manday", "BD manday", "Status", "Note"]

        # 조정된 컬럼순서 데이타프레임에 적용
        df_taskList = df_taskList[new_order]

        for index in df_taskList.index:
            mandayRate = df_taskList["BD manday"][index]
            bd_manday = mandayRate.split("/")[1]
            act_manday = mandayRate.split("/")[0]
            df_taskList["BD manday"][index] = bd_manday
            df_taskList["Act manday"][index] = act_manday

            for task in connect_shot_proj:
                if task[1] == df_taskList["Task name"][index] and task[4] == df_taskList["Part"][index]:
                    project = connect_shot_proj[task]
                    df_taskList["Project"][index] = self.projects[project][1]

            part = df_taskList["Part"][index]
            if part == "A": part = "ANI"
            elif part == "R": part = "RIG"
            elif part == "M": part = "MMV"
            df_taskList["Part"][index] = part

        self.editMandays = Edit_mandays.EditMandays(self)

        # 판다스 데이터프레임의 컬럼 이름 가져오기
        tableColumn = list(df_taskList.columns)

        # 테이블 위젯의 컬럼수 설정
        self.editMandays.ui.table_editMandays.setColumnCount(len(tableColumn))

        # 테이블 위젯의 컬럼 이름 설정
        self.editMandays.ui.table_editMandays.setHorizontalHeaderLabels(tableColumn)

        # 테이블 위젯의 행수 설정
        self.editMandays.ui.table_editMandays.setRowCount(len(df_taskList))                

        # 테이블 위젯의 크기정책 설정
        self.editMandays.ui.table_editMandays.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.editMandays.ui.table_editMandays.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # 테이블 위젯의 컬럼중 수동으로 조정해야할수 있는 컬럼의 수동 조정설정
        self.editMandays.ui.table_editMandays.horizontalHeader().setSectionResizeMode(0, QHeaderView.Interactive)
        self.editMandays.ui.table_editMandays.horizontalHeader().setSectionResizeMode(1, QHeaderView.Interactive)
        self.editMandays.ui.table_editMandays.horizontalHeader().setSectionResizeMode(2, QHeaderView.Interactive)
        self.editMandays.ui.table_editMandays.horizontalHeader().setSectionResizeMode(3, QHeaderView.Interactive)
        self.editMandays.ui.table_editMandays.horizontalHeader().setSectionResizeMode(4, QHeaderView.Interactive)
        self.editMandays.ui.table_editMandays.horizontalHeader().setSectionResizeMode(5, QHeaderView.Interactive)                        


        # 테이블 위젯에 데이타 저장
        for row_index, row_data in df_taskList.iterrows():
            for col_index, cell_value in enumerate(row_data):
                item = QTableWidgetItem(str(cell_value))
                self.editMandays.ui.table_editMandays.setItem(row_index, col_index, item)
                item.setTextAlignment(Qt.AlignCenter)

        # 각 셀의 아이템의 넓이에 맞춰 컬럼의 넓이 조정
        for col in range(len(df_taskList.columns)):
            self.delegate = readOnly.ReadOnlyDelegate(self.editMandays.ui.table_editMandays) # 편집불가능하게 하는 객체선언
            self.editMandays.ui.table_editMandays.resizeColumnToContents(col)
            if col !=4 and col !=6:
                self.editMandays.ui.table_editMandays.setItemDelegateForColumn(col, self.delegate)

        self.editMandays.ui.Button_save.clicked.connect(lambda: self.save_bidManday(selectedItems))
        self.editMandays.ui.Button_export.clicked.connect(self.export_excel)

        self.editMandays.show()



    def export_excel(self):

        exportPath = "/home/" + userID

        try:
            file_path, _ = QFileDialog.getSaveFileName(self, "Excel 파일저장", exportPath, "Excel Files (*.xlsx);;All Files (*)")

            if file_path:
                if not file_path.endswith(".xlsx"):
                    file_path += ".xlsx"

                df = self.table_to_excel(file_path)

        except Exception as e:
            print (f"{str(e)}")



    def table_to_excel(self, output_file):
        
        # 테이블 위젯을 데이타 프레임으로 변환
        headers = []
        column_widths = [] # 각 컬럼의 너비 저장


        table = self.editMandays.ui.table_editMandays

        for col in range(table.columnCount()):
            header_item = table.horizontalHeaderItem(col)
            
            if header_item is not None:
                headers.append(header_item.text())

            # 픽셀단위의 컬럼 너비를 엑셀단위로 변환(근사값)
            # 엑셀의 1단위는 약 7픽셀에 해당
            excel_width = table.columnWidth(col) / 7
            column_widths.append(excel_width)



        data = []
        for row in range(table.rowCount()):
            row_data = []
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append("")                    
            data.append(row_data)

        table_df = pd.DataFrame(data, columns=headers)
        table_df.to_excel(output_file, index=False)


        # openpyx 을 사용하여 열 너비 조정
        wb = load_workbook(output_file) # 저장된 엑셀파일을 여는 명령
        ws = wb.active # 현재 활성화된 워크시트를 가져옴

        # 각 컬럼의 너비 설정
        for i, width in enumerate(column_widths, 1):
            #최소너비 설정
            adjusted_width = max(width,8) #최소너비를 8로 설정

            # 최대너비 제한
            adjusted_width = min(adjusted_width, 50) # 최대 너비를 50으로 제한
            ws.column_dimensions[get_column_letter(i)].width = adjusted_width

        wb.save(output_file)
        return table_df




    # edittd_mandayData.json파이을 확인하여 현재 수정한 태스크의 bd맨데이가 택틱의 bd맨데이와 다르면, 수정된 
    # 맨데이를 edited_mandayData.json 에 저장
    # 저장하는 과정에서 active가 아닌 프로젝트의 데이타들은 모두 삭제
    def save_bidManday(self, tasks):
        row = self.editMandays.ui.table_editMandays.rowCount()
        column = self.editMandays.ui.table_editMandays.columnCount()

        saveItem = []
        for row in range(row):

            editMandayData = []
            bd_manday = self.editMandays.ui.table_editMandays.item(row, 4).text()
            taskName = self.editMandays.ui.table_editMandays.item(row, 2).text()
            part = self.editMandays.ui.table_editMandays.item(row, 1).text()
            project = self.editMandays.ui.table_editMandays.item(row, 0).text()

            if part == "ANI": part = "A"
            elif part == "RIG": part = "R"
            elif part == "MMV": part = "M"            

            editMandayData = [taskName, bd_manday, part, project]

            for task in tasks:
                if taskName == task[1] and part == task[4]:
                    if bd_manday != task[2].split("/")[1]:
                        saveItem.append(editMandayData)

        edited_mandayData_path = os.path.join(config.CURRENT_PATH, ".temp_BDmanday", "edited_mandayData.json")

        edited_Mandays_json = self.file_manager.import_Json(os.path.join(config.CURRENT_PATH, ".temp_BDmanday"), "edited_mandayData.json")


        projectList_model = self.projListview.model()

        activeProjects = []
        for row in range(projectList_model.rowCount()):
            item = projectList_model.item(row)
            if item is not None:
                activeProjects.append(item.text())

        del_tasks = [] # active 아닌 프로젝트에 해당되는 태스크는 삭제리스트에 저장
        for editedTask in edited_Mandays_json:
            if editedTask[3] not in activeProjects:
                del_tasks.append(editedTask)

        # edited_mandays_json을 순회하면서 del_tasks에 없는 아이템만 남기기
        edited_Mandays_json = [x for x in edited_Mandays_json if x not in del_tasks]


        add_tasks = []
        if edited_Mandays_json == []:
            self.file_manager.export_Json(os.path.join(config.CURRENT_PATH, ".temp_BDmanday"), "edited_mandayData.json", saveItem)

        elif edited_Mandays_json != []:
            for item in saveItem:                
                for editedItem in edited_Mandays_json:
                    if editedItem [0] == item[0] and editedItem [2] == item[2]:
                        editedItem[1] = item[1]
                        
                    elif item not in edited_Mandays_json:
                        add_tasks.append(item)

            for add_task in add_tasks:
                if add_task not in edited_Mandays_json:
                    edited_Mandays_json.append(add_task)

            self.file_manager.export_Json(currentPath+"/.temp_BDmanday", "edited_mandayData.json", edited_Mandays_json)

        self.reloadShotList() 
        self.refereshListViews(0, jsonData)           







    # 현재 shotlist 에서 선택한 태스크의 전체 스케쥴을 삭제
    def delete_schedule(self, indexs, currListview):

        global jsonData

        unscheduled_shots = [] # start/end date가 없는 샷들
        scheduleRange = []

        task_date = []
        for index in indexs:
            item = index.data()

            checkedMem, uncheckedMem = self.get_teamTree_member() # 현재 선택된 멤버 리스트 가져오기

            task = ""
            context = ""
            proj = connect_shot_proj[item]


            # 컨텍스트가 있는 태스크인지 확인
            if "/" not in (item[1]):
                task = item[1]

            else:
                taskContext = item[1].split("/")
                task = taskContext[0]
                context = taskContext[1]

            
            for mem in checkedMem:
                memEN = ""
                for member in self.memberCache_:
                    if self.memberCache_[member] == mem:
                        memEN = member


                # jsonData 내에 해당 태스크를 찾아 지움
                for data in jsonData:
                    if data["artist"] == memEN:
                        if proj in data["tasks"][0]:
                            if item[1] in data["tasks"][0][proj]:
                                data["tasks"][0][proj].remove(item[1])


        
        # 현재 스케쥴 리스트뷰의 갯수
        num_listview = self.dayUi.splitter.count()
        items = []
        for index in indexs:    
            items.append(currListview.model().data(index, Qt.DisplayRole))

        for item in items:
            # 현재 스케쥴 리스트뷰의 태스크들을 listView_task 리스트에 저장
            listView_task = []
            for i in range(num_listview):
                widget = self.dayUi.splitter.widget(i)
                listview_layout = widget.layout()

                if listview_layout:
                    frame = listview_layout.itemAt(2).widget()
                    frameLayout = frame.layout()
                    schListview = frameLayout.itemAt(0).widget()   

                    self.delete_items(schListview, currListview, item)


        self.updateJson(0)
        self.reloadShotList()







    # 전체스케쥴중 현재 선택한 태스크 찾기 ---> 현재 샷 리스트에서 선택한 태스크가 존재하는 리스트뷰를 찾아서 해당 태스크를 하이라이트함
    def find_schedule(self, indexs):

        item = indexs[0].data()
        unscheduled_shots = [] # start/end date가 없는 샷들
        scheduleRange = []

        checkedMem, uncheckedMem = self.get_teamTree_member() # 현재 선택된 멤버 리스트 가져오기

        task = ""
        context = ""
        proj = connect_shot_proj[item]

        # 컨텍스트가 있는 태스크인지 확인
        if "/" not in (item[1]):
            task = item[1]

        else:
            taskContext = item[1].split("/")
            task = taskContext[0]
            context = taskContext[1]

        task_date = []
        for mem in checkedMem:
            memEN = ""
            for member in self.memberCache_:
                if self.memberCache_[member] == mem:
                    memEN = member


            #task_date = []
            for data in jsonData:
                if data["artist"] == memEN:
                    if proj in data["tasks"][0]:
                        if item[1] in data["tasks"][0][proj]:
                            year = data["year"]
                            month = data["month"]
                            day = data["day"]

                            date = QDate(year, month, day)
                            task_date.append(date)

        if task_date != []:
            start = min(task_date).toPython()
            end = max(task_date).toPython()
            days_difference = (end - start).days

            # 스케쥴 리스트뷰들이 홀수로만 생성되고, 가운데 리스트뷰를 기준으로 양쪽으로 하나씩 두개가 붙기때문에 
            # 최소날짜와 최대날짜의 차이가 짝수이면 홀수로 바꾸어 주어야 함
            remainder = days_difference % 2

            if remainder == 1:
                self.changeListViewUI_(days_difference+4)
                schedule_EndDate = start + timedelta(days=days_difference+2)

            else:                
                self.changeListViewUI_(days_difference+2)
                schedule_EndDate = start + timedelta(days=days_difference+1)

            # 리스트뷰의 엔드날짜 
            numberDays = ((self.rangeEndDate.toPython()) - schedule_EndDate).days


            self.comboDay.setCurrentText(str(schedule_EndDate.day))
            self.comboMonth.setCurrentText(str(schedule_EndDate.month))            
            self.comboYear.setCurrentText(str(schedule_EndDate.year))

            # ui의 날짜 콤보박스에 반영된 날짜대로 리스트뷰 이동시키기
            self.changeComboDate()

            # 현재 스케쥴 리스트뷰의 갯수
            #num_listview = self.day_listViewLayout.count()
            num_listview = self.dayUi.splitter.count()

            for i in range(num_listview):
                #listview_layout = self.day_listViewLayout.layout().itemAt(i)
                widget = self.dayUi.splitter.widget(i)
                listview_layout = widget.layout()


                if listview_layout:
                    frame = listview_layout.itemAt(2).widget()
                    frameLayout = frame.layout()
                    listview = frameLayout.itemAt(0).widget()
                    
                    tasks = self.getListViewItem(listview) # 현재 ui상에 디스플레이되어있는 아이템들 가져오기

                    # 현재 찾으려는 태스크와 동일한 태스크가 리스트뷰에 존재하면 하이라이트 시킴
                    if item in tasks:

                        model = listview.model()

                        for row in range(model.rowCount(QModelIndex())):
                            index = model.index(row)
                            taskName = model.data(index, Qt.DisplayRole)
                            
                            if taskName == item:
                                listview.setCurrentIndex(index)



    # 추가한 아이템, 혹은 찾은 아이템을 확인하기 쉽도록 하일라이트 시킴
    def selectItems(self, indexes):
        num_listview = self.dayUi.splitter.count()

        for i in range(num_listview):
            #listview_layout = self.day_listViewLayout.layout().itemAt(i)
            widget = self.dayUi.splitter.widget(i)
            listview_layout = widget.layout()

            if listview_layout:
                frame = listview_layout.itemAt(2).widget()
                frameLayout = frame.layout()
                listview = frameLayout.itemAt(0).widget()
                
                tasks = self.getListViewItem(listview) # 현재 ui상에 디스플레이되어있는 아이템들 가져오기

                # 현재 찾으려는 태스크와 동일한 태스크가 리스트뷰에 존재하면 하이라이트 시킴

                model = listview.model()
                selection_model = listview.selectionModel()
                for row in range(model.rowCount(QModelIndex())):
                    index = model.index(row, 0)
                    taskName = model.data(index, Qt.DisplayRole)
                    if taskName in indexes:
                        selection_model.select(index, QItemSelectionModel.Select)



    # 현재 샷 리스트에서 선택한 태스크들을 스케쥴에 맞게 리스트뷰에 배치시키는 메서드
    def add_schedule(self, indexs):

        unscheduled_shots = [] # start/end date가 없는 샷들
        scheduleRange = []
        taskInfoJson = self.file_manager.import_Json(currentPath+"/.task_info", userID+"_task_info.json") # 현재 로드되어있는 태스크 정보 가져오기        

        for index in indexs:
            item = index.data()

            checkedMem, uncheckedMem = self.get_teamTree_member() # 현재 선택된 멤버 리스트 가져오기


            task = ""
            context = ""

            # 컨텍스트가 있는 태스크인지 확인
            if "/" not in (item[1]):
                task = item[1]

            else:
                taskContext = item[1].split("/")
                task = taskContext[0]
                context = taskContext[1]


            for mem in checkedMem:

                taskList = []
                memEN = []
                
                for nameEN in self.memberCache_:
                    if self.memberCache_[nameEN] == mem:
                        memEN.append(nameEN)


                for name in memEN:
                    for taskInfo in taskInfoJson[name]:
                        process = taskInfo["process"]

                        if context != "": processContext = process+"/"+ context
                        else : processContext = process

                        dep = "" 
                        proj = connect_shot_proj[item]
                        proj_code = self.projects[proj][0]


                        if (taskInfo["extra_code"] == task and taskInfo["context"] == processContext) and taskInfo["project_code"] == proj_code:

                            if taskInfo["start_date"] != None and taskInfo["end_date"] != None:
                                startDate_str =  taskInfo["start_date"]
                                endDate_str = taskInfo["end_date"]
                                process = taskInfo["process"]
                                #proj = connect_shot_proj[item]

                                start_dt = datetime.strptime(startDate_str, "%Y-%m-%d %H:%M:%S")
                                end_dt = datetime.strptime(endDate_str, "%Y-%m-%d %H:%M:%S")

                                startDate = QDate(start_dt.year, start_dt.month, start_dt.day)
                                endDate = QDate(end_dt.year, end_dt.month, end_dt.day)
                                
                                between_date = startDate
                                betweenDate_list = []

                                while between_date <= endDate:
                                    date_key = between_date.toPython() # Qt의 QDate를 python의 datetime으로 변환
                                    betweenDate_list.append(date_key)
                                    between_date = between_date.addDays(1)

                                if process == "matchmove": dep = "MMV"
                                elif process == "animation": dep = "ANI"
                                elif process == "creature": dep = "RIG"

                                self.set_schedule_StartEnd(betweenDate_list, name, proj, item[1], taskList, dep)

                                for date in betweenDate_list:
                                    scheduleRange.append(date)

                            else:
                                unscheduled_shots.append(item[1])



        if unscheduled_shots != []:

            str_shots = ""
            for shot in unscheduled_shots:
                str_shots = str_shots + "\n" + shot

            QMessageBox.information(self, "    start/end date가 존재하지 않습니다.    ", f" : {str_shots}")


        # 반영된 스케쥴이 있는 날짜로 ui 이동
        if scheduleRange != []:
            minDate = min(scheduleRange)
            maxDate = max(scheduleRange)
            days_difference = (maxDate - minDate).days + 1



            # 스케쥴 리스트뷰들이 홀수로만 생성되고, 가운데 리스트뷰를 기준으로 양쪽으로 하나씩 두개가 붙기때문에 
            # 최소날짜와 최대날짜의 차이가 짝수이면 홀수로 바꾸어 주어야 함
            remainder = days_difference % 2

            if days_difference > 20: # 스케쥴 리스트뷰의 최대개수 제한
                days_difference = 20

            total_addDays = 0

            if remainder == 1:
                total_addDays = days_difference
                #self.changeListViewUI_(days_difference+4)
                schedule_EndDate = minDate + timedelta(days=days_difference-1)


            else:      
                total_addDays = days_difference+1
                #self.changeListViewUI_(days_difference+2)
                schedule_EndDate = minDate + timedelta(days=days_difference)

            # 스케쥴만큼 리스트뷰 갯수 추가 혹은 삭제
            self.changeListViewUI_(total_addDays)

            self.comboYear.setCurrentText(str(schedule_EndDate.year))
            self.comboMonth.setCurrentText(str(schedule_EndDate.month))            
            self.comboDay.setCurrentText(str(schedule_EndDate.day))

            # ui의 날짜 콤보박스에 반영된 날짜대로 리스트뷰 이동시키기
            self.changeComboDate()

            addDatas = [index.data() for index in indexs if index] 


            self.reloadShotList()
            self.selectItems(addDatas)





    def set_schedule_StartEnd(self, dateList, member, project, task, taskList, department):

        global jsonData

        removeData = []
        for data in jsonData:
            if data['tasks'] == [{}]:
                removeData.append(data)

        for rm_data in removeData:
            jsonData.remove(rm_data)

        add_ScheduleData = []


        for date in dateList:

            day_num = date.isoweekday()
            if day_num != 6 and day_num != 7:

                scheduleData = {}

                projectTasks = {project : [task]}

                scheduleData["artist"] = member
                scheduleData["year"] = date.year
                scheduleData["month"] = date.month
                scheduleData["day"] = date.day
                scheduleData["tasks"] = [projectTasks]
                scheduleData["department"] = department


                if taskList == []:
                    taskList.append(scheduleData)


                elif  (self.validation_manager.check_date_exists(taskList, scheduleData) == False):
                        taskList.append(scheduleData)


                elif taskList != []:
                    for data in taskList:
                        if data["artist"] == member and data["year"] == date.year and data["month"] == date.month and data["day"] == date.day and data["tasks"] != scheduleData["tasks"]:
                            #taskList 내에 동일한 날짜, 동일한 아티스트의 태스크리스트에 해당 프로젝트가 있을경우 태스크를 이 프로젝트에 포함시킴
                            if project in data["tasks"][0]:
                                if task not in data["tasks"][0][project]:
                                    data["tasks"][0][project].append(task)


                            else: 
                                data["tasks"][0][project] = [task]




                for addTask in taskList:
                    if jsonData != []:
                        # 날짜가 존재하는지 확인후 존재하지 않으면 스케쥴에 추가
                        if  (self.validation_manager.check_date_exists(jsonData, addTask) == False):
                            jsonData.append(addTask)

        
                        # 날짜가 존재하고 아티스트도 같은 태스크가 존재할 경우 
                        elif (self.validation_manager.check_dateMember_exists(jsonData, addTask) == True):# and self.check_artist_exists(jsonData, addTask) == True):
                            for data in jsonData:
                                if data["artist"] == addTask["artist"] and data["year"] == addTask["year"] and data["month"] == addTask["month"] and data["day"] == addTask["day"]:# and data["tasks"] != scheduleData["tasks"]:
                                    if project in data["tasks"][0]:
                                        for _task in addTask["tasks"][0][project]:
                                            if _task not in data["tasks"][0][project]:
                                                data["tasks"][0][project].append(_task)

                                    elif project not in data["tasks"][0]:
                                        data["tasks"][0][project] = [task]

                        else:

                            jsonData.append(addTask)                        


                    elif jsonData == []:
                        jsonData.append(addTask)


                    

    def get_teamTree_member(self):

        # 현재유저의 팀원정보 가져오기
        teamInfoData = self.team_manager.get_team_info(userID)

        if userID not in self.memberCache_:
            self.memberCache_[userID] = self.currName_kr

        if teamInfoData != []:
            existMember_inJson = list(teamInfoData[0].keys())
        else:
            existMember_inJson = []            

        invisibleRoot = self.teamTreeModel.invisibleRootItem()
        rootItem = invisibleRoot.child(0)

        checked_items_list = []
        unchecked_items_list = []

        self.team_manager.get_tree_all_items(rootItem, checked_items_list, unchecked_items_list, existMember_inJson, teamTreeHierarchy, 0)

        return checked_items_list, unchecked_items_list

    # 주어진 month가 어느분기에 속하는지 찾아서 반환
    def getQuarter(self, month):
        
        #month = date.month()

        if 1 <= month <=3:
            return  1

        elif 4 <= month <=6:
            return 2

        elif 7 <= month <=9:
            return 3

        elif 10 <= month <=12:
            return 4

        else:
            return -1




    # 분기별 작업량 취합하여 가져오기
    def get_workload_JsonData(self, itemName):

        global jsonData

        currentDate = QDate.currentDate()
        currentYear = currentDate.year()
        currentMonth = currentDate.month()
        currentQuarter = self.getQuarter(currentMonth)

        periodWorkload = {}
        for data in jsonData:

            if data["tasks"] != [{}] and data["artist"] == itemName:

                year = data["year"]
                month = data["month"]
                quarter_data = self.getQuarter(month)

                if year not in list(periodWorkload.keys()):
                    periodWorkload[year] = {}


                if quarter_data != -1:

                    # 1분기
                    if quarter_data == 1:
                        if "Q1" in list(periodWorkload[year].keys()):
                            periodWorkload[year]["Q1"].append(data)

                        elif "Q1" not in list(periodWorkload[year].keys()):
                            periodWorkload[year]["Q1"] = [data]

                    # 2분기
                    elif quarter_data == 2:
                        if "Q2" in list(periodWorkload[year].keys()):
                            periodWorkload[year]["Q2"].append(data)

                        elif "Q2" not in list(periodWorkload[year].keys()):
                            periodWorkload[year]["Q2"] = [data]

                    # 3분기
                    elif quarter_data == 3:
                        if "Q3" in list(periodWorkload[year].keys()):
                            periodWorkload[year]["Q3"].append(data)

                        elif "Q3" not in list(periodWorkload[year].keys()):
                            periodWorkload[year]["Q3"] = [data]

                    # 4분기
                    elif quarter_data == 4:
                        if "Q4" in list(periodWorkload[year].keys()):
                            periodWorkload[year]["Q4"].append(data)

                        elif "Q4" not in list(periodWorkload[year].keys()):
                            periodWorkload[year]["Q4"] = [data]


        thisQuarter = str(currentYear) + "/" + "Q" + str(currentQuarter)

        return thisQuarter, periodWorkload



    def export_QuarterlyWorkload(self):
        
        for member in self.memberCache_:

            thisQuarter, periodWorkload = self.get_workload_JsonData(member)
            self.file_manager.export_Json(currentPath+"/.quarterlyWorkload", member+"_workload.json", periodWorkload)






    def get_selected_projects(self):
        """프로젝트 리스트뷰에서 선택된 프로젝트 코드들을 반환"""
        selected_indexes = self.projListview.selectedIndexes()
        selectedItems = [index.data() for index in selected_indexes]

        selectedProj = []
        for item in selectedItems:
            for proj in self.projects:
                if self.projects[proj][1] == item:
                    selectedProj.append(proj)

        return selectedProj

    def get_cached_task_info(self, members, taskInfoJson=None):
        """멤버별 태스크 정보를 캐싱하여 반환"""
        taskInfo = {}
        
        if taskInfoJson is None:
            taskInfoJson = self.file_manager.import_Json(currentPath+"/.task_info", userID+"_task_info.json")
        
        for member in members:
            if member in list(taskInfoJson.keys()):
                taskInfo[member] = taskInfoJson[member]
            elif member not in list(taskInfoJson.keys()):
                taskInfo[member] = self.getTaskInfo(member)
                taskInfoJson[member] = taskInfo[member]
                self.file_manager.export_Json(currentPath+"/.task_info", userID+"_task_info.json", taskInfoJson)
        
        return taskInfo

    def hide_completed_project_ui(self, ui_object, background_color="#252525"):
        """완료된 프로젝트 UI를 숨기고 어두운 색상으로 설정"""
        ui_object.hide()
        ui_object.ui.frame_mandayStatus.setStyleSheet(f"QFrame {{ background-color : {background_color};}}")
        ui_object.ui.label_project.setStyleSheet("color : #555555")
        ui_object.ui.label_shots.setStyleSheet("color : #555555")
        ui_object.ui.label_manday.setStyleSheet("color : #555555")
        ui_object.ui.label_amount_shots.setStyleSheet("color : #555555")
        ui_object.ui.label_amount_md.setStyleSheet("color : #555555")

    # 멤버의 이름을 더블클릭했을때 나오는 현재 프로젝트별 진행현형 다잉얼로그 창 실행
    def show_status_manday(self, index):

        self.click_timer.stop()
        self.pending_click = None

        if self.memberStatusDialog is not None: # 다이얼로그가 이미 열려있는지 확인
            self.memberStatusDialog.close() # 기존 다이얼로그를 닫음
            self.memberStatusDialog.deleteLater() # 기존 다이얼로그 메모리 해제
            self.memberStatusDialog = None # 다이얼로그 객체 삭제

        # 현재 새로 드래그드롭 되어진 스케쥴까지 포함시켜야 해서 업데이트를 한번 실행해야함
        self.updateJson(0)

        self.memberStatusDialog = status_projManday.ProjMandayDialog(self)
        self.memberStatusDialog.setWindowTitle("Project Status Dashboard")   

        itemName = self.teamTreeModel.data(index, Qt.DisplayRole)

        # 현재유저가 팀원이면 팀정보를 보는 버튼을 감춤
        if self.job == "팀원" :
                    self.memberStatusDialog.ui.button_teamWorkload.hide()            

        # 현재 더블클릭한 멤버가 팀원이면 팀정보를 볼수 있는 버튼을 감춤
        if teamInfo_ != []:
            for teamMember in teamInfo_[0]:
                if teamInfo_[0][teamMember]["nameKr"] == itemName  and  teamInfo_[0][teamMember]["job"] == "팀원":
                    self.memberStatusDialog.ui.button_teamWorkload.hide()



        itemNameEN = []
        for member in self.memberCache_:
            if self.memberCache_[member] == itemName:
                itemNameEN.append(member)

        active_projects = list(self.projects.keys())
        taskInfoJson = self.file_manager.import_Json(currentPath+"/.task_info", userID+"_task_info.json") 
        userInfoJson = self.file_manager.import_Json(currentPath+"/.user_Info", itemNameEN[0]+"_userInfo.json") 

        thisQuarter, periodWorkload = self.get_workload_JsonData(itemNameEN[0])


        # 현재사용자의 태스크 정보는 이미 읽어온 상태이므로 클릭한 멤버를 확인하여 현재사용자가 아닌경우 그 사용자의 태스크 정보를 읽어옴
        if itemNameEN[0] != userID and (itemNameEN[0] not in list(taskInfoJson.keys())):
            taskInfoJson[itemNameEN[0]] = self.getTaskInfo(itemNameEN[0])

        # 현재사용자의 체크박스가 off 되면 taskInfo에서 현재사용자의 데이타가 삭제되는데, 그때 현재사용자를 더블클릭했을경우 현재사용자의 태스크 정보를 읽어오기 위해
        elif itemNameEN[0] == userID:
            taskInfoJson[itemNameEN[0]] = self.getTaskInfo(itemNameEN[0])

        # 각 프로젝트별 생성될 UI를 저장할 리스트
        self.memberStatusDialog.projUI = []

        # 분기별 진행한 태스크의 갯수와 완료현황, 맨데이 현황을 가져옴
        taskWorkload_json, mandayStatus_json  = self.get_shotWorkload_byPeriod (taskInfoJson, itemNameEN[0], periodWorkload)

        # 클릭한 멤버의 각 프로젝트의 맨데이 현황을 읽어옴
        for proj in active_projects:
            projManday, progressValue, mandayProgressBar_set, bidManday = self.get_status_projectManday(taskInfoJson, proj, itemNameEN[0] ) # 프로젝트별 맨데이 현황 가져오기
            shotStatus, shotProgressValue, progressBar_set, wipValue, wip_shots = self.get_status_projectShot(taskInfoJson, proj, itemNameEN[0]) # 프로젝트별 샷진행 현황 가져오기

            # 현재 선택한 아티스트의 현재분기 작업량의 전체태스크가 아닌 jsonData(스케쥴UI에 기록된)의 샷 진행현황만 분기별로 정리해서 가져옴
            appTasks = "0"
            allTasks = "0"
            app_allTask = "0"
            usedManday = "0.0"
            allManday = "0.0"
            used_allManday = "0.0"
            progressValue_totalTasks = 0
            progressValue_totalMandays = 0
            progressBar_setup_totalTasks = ""
            progressBar_setup_totalMandays = ""


            if taskWorkload_json != {}:

                for period in taskWorkload_json:
                    if period == thisQuarter:
                        thisQuarter_shotWorkload = taskWorkload_json[period]
                        thisQuarter_mandayStatus = mandayStatus_json[period]

                        task_workload = thisQuarter_shotWorkload.split("/")

                        appTasks = task_workload[0]
                        allTasks = task_workload[1]

                        if allTasks != "0":
                            app_allTask = str(appTasks) + " / " + str(allTasks)
                            progressValue_totalTasks = (int(appTasks) / int(allTasks))*100
                            progressBar_setup_totalTasks = (self.setup_shot_progressBar(allTasks, appTasks)).replace("height: 15px", "height: 25px")

                            manday_status = thisQuarter_mandayStatus.split("/")
                            usedManday = manday_status[0]
                            allManday = manday_status[1]
                            used_allManday = str(usedManday) + " / " + str(allManday)
                            if allManday != "0.0":
                                progressValue_totalMandays = (float(usedManday) / float(allManday))*100
                                if progressValue_totalMandays > 100:
                                    progressValue_totalMandays = 100

                            progressBar_setup_totalMandays = (self.setup_manday_progressBar(float(usedManday), float(allManday), progressValue_totalMandays)).replace("height: 15px", "height: 25px")

                           
            if projManday != "":

                app = int(shotStatus.split("/")[0])
                all = int(shotStatus.split("/")[1])

                projMandayUI = proj_Manday.ProjManday()
                self.memberStatusDialog.ui.v_Layout_allManday.addWidget(projMandayUI)
                projMandayUI.ui.label_project.setText(self.projects[proj][1])
                projMandayUI.ui.label_progressBar_md.setValue(progressValue)
                projMandayUI.ui.label_amount_md.setText(projManday)

                # 커스텀 멀티색깔 프로그래스바를 사용하기 위해 기존 UI로 만들었던 프로그래스바를 삭제
                del_label_progressBar_shots = projMandayUI.ui.label_progressBar_shots
                del_label_progressBar_shots.deleteLater()
                #parent = del_label_progressBar_shots.parent()

                # 멀티색깔 프로그래스바 객체 생성
                multi_label_progressBar_shots = mpb.MultiColorProgressBar()

                # 기존 삭제한 프로그래스바의 위치에 새로 생성한 멀티색깔 프로그래스바 배치
                projMandayUI.ui.h_Layout_shots.insertWidget(1, multi_label_progressBar_shots)

                # 각 색깔별 프로그래스바의 영역 설정
                if shotProgressValue == wipValue:
                    wipValue = wipValue-1

                multi_label_progressBar_shots.addSegment(shotProgressValue,'#484970')#'#968ae1')
                multi_label_progressBar_shots.addSegment(wipValue,'#5b8961')#'#c1cec1')

                projMandayUI.ui.label_amount_shots.setText(shotStatus)
                multi_label_progressBar_shots.setStyleSheet( progressBar_set )
                projMandayUI.ui.label_progressBar_md.setStyleSheet( mandayProgressBar_set )

                if bidManday == 0:
                    projMandayUI.ui.label_progressBar_md.setTextVisible(False)
                    projMandayUI.ui.label_amount_md.setStyleSheet("color : #D0D0D0")                

                self.memberStatusDialog.projUI.append(projMandayUI)

                if app == all: # 작업이 완료된 프로젝트는 숨기고, unhide시킬경우 어두운색깔로 표시되게 함
                    self.hide_completed_project_ui(projMandayUI, "#252525")                                        


        self.memberStatusDialog.ui.button_teamWorkload.clicked.connect(lambda: self.show_team_status(index)) 
        self.verticalSpacer = QSpacerItem(20,40,QSizePolicy.Minimum, QSizePolicy.Expanding)
        self.memberStatusDialog.ui.v_Layout_allManday.addItem(self.verticalSpacer)
        self.memberStatusDialog.ui.label_name.setText(itemName)
        self.memberStatusDialog.ui.label_team.setText(userInfoJson[0]["team"])
        self.memberStatusDialog.ui.label_allTasks.setText(appTasks)
        self.memberStatusDialog.ui.label_allMandays.setText(usedManday)
        self.memberStatusDialog.ui.label_wip_allTasks.setText(app_allTask)
        self.memberStatusDialog.ui.label_used_allMandays.setText(used_allManday)
        self.memberStatusDialog.ui.progressBar_quarterTasks.setValue(progressValue_totalTasks)
        self.memberStatusDialog.ui.progressBar_quarterMandays.setValue(progressValue_totalMandays)
        self.memberStatusDialog.ui.progressBar_quarterTasks.setStyleSheet( progressBar_setup_totalTasks )
        self.memberStatusDialog.ui.progressBar_quarterMandays.setStyleSheet( progressBar_setup_totalMandays )        
        self.memberStatusDialog.ui.checkBox_visivilityCompletedProj.stateChanged.connect(self.visibilityProjStatusUI)
        self.memberStatusDialog.ui.line.setStyleSheet("QFrame { background-color: #323232; height: 2px; border: none;}")

        self.memberStatusDialog.show()




    # 프로젝트별 현황UI 중 완료된 프로젝트의 비져빌리티 옵션        
    def visibilityProjStatusUI(self, state):

        for proj_ui in self.memberStatusDialog.projUI:
            #mmm = proj_ui.ui.label_amount_md.text()
            app = int((proj_ui.ui.label_amount_shots.text()).split("/")[0])
            all = int((proj_ui.ui.label_amount_shots.text()).split("/")[1])

            if (app == all) and (state == 0):
                proj_ui.hide()

            elif  (state == 2):
                proj_ui.show()



    # 팀의 전체 정보를 표시할때, 팀 멤버들의 프로젝트별 데이타를 합치기 위해서 모든 하위 팀 멤버들을 하나의 리스트로 만듦
    def make_orderList(self, index, orderList):

        childNum = self.teamTreeModel.rowCount(index)
        selectedName = self.teamTreeModel.data(index, Qt.DisplayRole)
        
        selectedNameEN = ""
        for member in self.memberCache_:
            if self.memberCache_[member] == selectedName:
                selectedNameEN = member


        if childNum != 0:
            for i in range(childNum):
                childItem = index.child(i,0)
                nameKr = self.teamTreeModel.data(childItem, Qt.DisplayRole)

                for member in teamInfo_[0]:
                    if (teamInfo_[0][member]["nameKr"] == nameKr):
                        orderList.append(member)

                self.make_orderList(childItem, orderList)

        if selectedNameEN not in orderList:
            orderList.append(selectedNameEN)





    def get_child_items(self, index, total_app_tasks, total_tasks, total_act_mandays, total_mandays, teamProjManday, teamTaskStatus, total_steps, teamWipStatus, orderList):

        i = 0
        numberStep = len(orderList)
        for member in orderList:

            # 현재사용자의 태스크정보 읽어오기
            taskInfoJson = self.file_manager.import_Json(currentPath+"/.task_info", userID+"_task_info.json")

            # 현재사용자의 태스크 정보가 없는경우 빈 리스트 항목만 만들어준다.
            if userID not in taskInfoJson:
                taskInfoJson[userID] = []

            # 현재사용자의 태스크 정보는 이미 읽어온 상태이므로 클릭한 멤버를 확인하여 현재사용자가 아닌경우 그 사용자의 태스크 정보를 읽어옴
            if member != userID and (member not in list(taskInfoJson.keys())):
                taskInfoJson[member] = self.getTaskInfo(member)


            # 각 멤버별 프로젝트별 맨데이 합산
            active_projects = list(self.projects.keys())

            for proj in active_projects:
                if proj != "testshot":
                    projManday, progressValue, mandayProgressBar_set, bidManday = self.get_status_projectManday(taskInfoJson, proj, member ) # 프로젝트별 맨데이 현황 가져오기
                    taskStatus, shotProgressValue, progressBar_set, wipValue, wip_shots = self.get_status_projectShot(taskInfoJson, proj, member) # 프로젝트별 샷진행 현황 가져오기

                    if "/" in taskStatus and taskStatus != "0 / 0":
                        appTask = int(taskStatus.split("/")[0])
                        allTask = int(taskStatus.split("/")[1])

                        if proj not in teamTaskStatus:
                            teamTaskStatus[proj] = taskStatus
                            teamWipStatus[proj] = str(wip_shots) + " / " + str(allTask)

                        elif proj in teamTaskStatus:
                            existAppTask = int(teamTaskStatus[proj].split("/")[0])
                            existAllTask = int(teamTaskStatus[proj].split("/")[1])                                    

                            existWipTask = int(teamWipStatus[proj].split("/")[0]) 

                            addedAppTask = appTask + existAppTask
                            addedAllTask = allTask + existAllTask
                            addedWipTask = wip_shots + existWipTask

                            addedTaskStatus = str(addedAppTask) + " / " + str(addedAllTask)
                            addedWipStatus = str(addedWipTask) + " / " + str(addedAllTask)

                            teamTaskStatus[proj] = addedTaskStatus
                            teamWipStatus[proj] = addedWipStatus


                    if projManday != "":
                        actManday = round(float(projManday.split("/")[0]),2)
                        bidManday = round(float(projManday.split("/")[1]),2)

                        if proj not in teamProjManday:
                            teamProjManday[proj] = projManday

                        elif proj in teamProjManday:
                            existActManday = round(float(teamProjManday[proj].split("/")[0]),2)
                            existBidManday = round(float(teamProjManday[proj].split("/")[1]),2)

                            addedActManday = actManday + existActManday
                            addedBidManday = bidManday + existBidManday

                            addedProjManday = str(addedActManday) + " / " + str(addedBidManday)

                            teamProjManday[proj] = addedProjManday



            #for proj in active_projects:
            thisQuarter, periodWorkload = self.get_workload_JsonData(member)
            taskWorkload_json, mandayStatus_json  = self.get_shotWorkload_byPeriod (taskInfoJson, member, periodWorkload)

            if thisQuarter in taskWorkload_json:
                tasks = taskWorkload_json[thisQuarter].split("/")
                appTasks = tasks[0]
                allTasks = tasks[1]

                total_app_tasks = total_app_tasks + int(appTasks)
                total_tasks = total_tasks + int(allTasks)

                mandays = mandayStatus_json[thisQuarter].split("/")
                actMandays = mandays[0]
                allMandays = mandays[1]

                total_act_mandays = total_act_mandays + round(float(actMandays), 2)
                total_mandays = total_mandays + round(float(allMandays), 2)

            i = i+1
            progress = int((i+1) / numberStep * 100)

            if self.update_progress:
                self.update_progress(progress)

        return total_app_tasks, total_tasks, total_act_mandays, total_mandays, teamProjManday, teamTaskStatus, teamWipStatus



    # show_team_status메서드의 진행이 끝나면 창을 닫기위한 시그널 함수
    def progressBar_task_finished(self):
        self.loadingProgressDialog.close()



    # 하위 멤버들의 어싸인된 프로젝트를 로딩이 끝나면 창을 닫기위한 시그널 함수
    def loadingProj_progress_task_finished(self):
        self.projSelectionProgressDialog.close()



    # 프로그레스바의 설정을 위한 쓰레드내의 작업결과 변수들을 가져오기 위한 시그널에 연결된 메서드
    # 그리고 그 변수들을 사용하기 위해 show_team_status메서드를 둘로 나눠서 나머지를 이 메서드내에서 실행되게 함. ㅠㅠ...
    def handle_result(self, result):

        projManday = result[4]
        projShotStatus = result[5]
        projWipStatus = result[6]        

        # 프로젝트별로 생성되는 ui를 저장해두는 리스트(체크박스 on/off 위해)
        self.teamStatusDialog.projUI_list = []
        if projManday != "":
            # 프로젝트별 UI 생성
            for proj in projManday:

                actManday = float(projManday[proj].split("/")[0])
                bidManday = float(projManday[proj].split("/")[1])

                mandayProgressValue = 0.0
                if bidManday == 0: mandayProgressValue = 0
                else:  
                    mandayProgressValue = (actManday/bidManday)*100
                    if mandayProgressValue > 100:
                        mandayProgressValue = 100

                appTasks = float(projShotStatus[proj].split("/")[0])
                allTasks = float(projShotStatus[proj].split("/")[1])
                wipTasks = float(projWipStatus[proj].split("/")[0])

                tasksProgressValue = 0.0
                if allTasks == 0 : tasksProgressValue = 0
                else: 
                    tasksProgressValue = (appTasks/allTasks)*100
                    tasksWipValue = (wipTasks/allTasks)*100

                projMandayProgressSet = self.setup_manday_progressBar(actManday, bidManday, mandayProgressValue)    
                projTasksProgressSet = self.setup_shot_progressBar(appTasks, allTasks)


                # UI 설정
                projTeamMandayUI = proj_Manday.ProjManday()

                self.teamStatusDialog.projUI_list.append(projTeamMandayUI)

                self.teamStatusDialog.ui.v_Layout_scrollArea.addWidget(projTeamMandayUI)
                projTeamMandayUI.ui.label_project.setText(self.projects[proj][1])
                projTeamMandayUI.ui.label_progressBar_md.setValue(mandayProgressValue)
                projTeamMandayUI.ui.label_amount_md.setText(projManday[proj])

                #projTeamMandayUI.ui.label_progressBar_shots.setValue(tasksProgressValue)
                # 커스텀 멀티색깔 프로그래스바를 사용하기 위해 기존 UI로 만들었던 프로그래스바를 삭제
                del_label_progressBar_teamShots = projTeamMandayUI.ui.label_progressBar_shots
                del_label_progressBar_teamShots.deleteLater()

                # 멀티색깔 프로그래스바 객체 생성
                multi_label_progressBar_teamShots = mpb.MultiColorProgressBar()

                # 기존 삭제한 프로그래스바의 위치에 새로 생성한 멀티색깔 프로그래스바 배치
                projTeamMandayUI.ui.h_Layout_shots.insertWidget(1, multi_label_progressBar_teamShots)


                # 각 색깔별 프로그래스바의 영역 설정
                if tasksProgressValue == tasksWipValue:
                    tasksWipValue = tasksWipValue-1


                # 각 색깔별 프로그래스바의 영역 설정
                multi_label_progressBar_teamShots.addSegment(tasksProgressValue,'#484970')
                multi_label_progressBar_teamShots.addSegment(tasksWipValue,'#5b8961')

                projTeamMandayUI.ui.label_amount_shots.setText(projShotStatus[proj])

                multi_label_progressBar_teamShots.setStyleSheet( projTasksProgressSet )
                projTeamMandayUI.ui.label_progressBar_md.setStyleSheet( projMandayProgressSet )


                if appTasks == allTasks:
                    self.hide_completed_project_ui(projTeamMandayUI, "#303030")                         


        all_app = result[0]
        all_tasks = result[1]
        all_actManday = round(result[2],1)
        all_manday = round(result[3],1)

        taskProgressSet = (self.setup_shot_progressBar(all_tasks, all_app)).replace("height: 15px", "height: 25px")
        Q_app_all_tasks = str(all_app)+"/"+str(all_tasks)
        Q_act_bid_mandays = str(all_actManday)+"/"+str(all_manday)


        # manday
        Q_mandayProgressValue = 0.0
        if float(all_manday) == 0:
            Q_mandayProgressValue = 0

        else:
            Q_mandayProgressValue = (float(all_actManday)/float(all_manday))*100
            if Q_mandayProgressValue > 100:
                Q_mandayProgressValue = 100

        mandayProgressSet = (self.setup_manday_progressBar(all_actManday, all_manday, Q_mandayProgressValue)).replace("height: 15px", "height: 25px")

        # task
        Q_taskProgressValue = 0.0
        if all_tasks == 0:
            Q_taskProgressValue = 0

        else:
            Q_taskProgressValue = (float(all_app) / float(all_tasks))*100

        self.teamStatusDialog.ui.label_wip_allTeamTasks.setText(Q_app_all_tasks)
        self.teamStatusDialog.ui.label_used_allTeamMandays.setText(Q_act_bid_mandays)        
        self.teamStatusDialog.ui.label_allTasks.setText(str(all_app))        
        self.teamStatusDialog.ui.label_allTeamMandays.setText(str(all_actManday))        

        self.teamStatusDialog.ui.progressBar_QteamTasks.setValue(Q_taskProgressValue)
        self.teamStatusDialog.ui.progressBar_QteamTasks.setStyleSheet( taskProgressSet )
        self.teamStatusDialog.ui.progressBar_QteamMandays.setValue(Q_mandayProgressValue)
        self.teamStatusDialog.ui.progressBar_QteamMandays.setStyleSheet( mandayProgressSet )

        self.teamStatusDialog.ui.v_Layout_scrollArea.addStretch()
        self.teamStatusDialog.ui.v_Layout_scrollArea.setAlignment(Qt.AlignTop)

        self.teamStatusDialog.ui.checkBox_visivilityCompletedProj.stateChanged.connect(self.visibilityTeamStatusUI)
        self.teamStatusDialog.ui.line.setStyleSheet("QFrame { background-color: #323232; height: 2px; border: none;}")

        self.teamStatusDialog.show()





    def show_team_status(self, index):
        
        self.teamStatusDialog = status_teamManday.TeamMandayDialog(self)
        self.teamStatusDialog.setWindowTitle("Team Status Dashboard")

        itemName = self.teamTreeModel.data(index, Qt.DisplayRole)

        itemNameEN = ""
        for member in self.memberCache_:
            if self.memberCache_[member] == itemName:
                itemNameEN = member

        userInfoJson = self.file_manager.import_Json(currentPath+"/.user_Info", itemNameEN+"_userInfo.json") 

        job = userInfoJson[0]["job"] + " : " + itemName
        team = userInfoJson[0]["team"]


        invisibleRoot = self.teamTreeModel.invisibleRootItem()
        rootItem = invisibleRoot.child(0)

        total_app_tasks = 0
        total_tasks = 0

        total_act_mandays = 0.0
        total_mandays = 0.0

        team_projManday = {}
        team_shotStatus = {}
        team_wipStatus = {}
        orderList = []


        # 현재 선택한 멤버의 하위 모든 멤버의 이름 리스트 만들기
        self.make_orderList(index, orderList)

        self.loadingProgressDialog = loadingProgress.LoadingProgressDialog(self)
        self.loadingProgressDialog.setModal(True)
        self.loadingProgressDialog.ui.progressBar_loading.setValue(0)
        self.loadingProgressDialog.setWindowTitle("Loading data")
        self.loadingProgressDialog.show()

        self.progressBar_thread = pb.ProgressBar_Thread(index, total_app_tasks, total_tasks, total_act_mandays, total_mandays, team_projManday, team_shotStatus, team_wipStatus, orderList, self)
        self.progressBar_thread.finished_signal.connect(self.progressBar_task_finished)
        self.progressBar_thread.progress_signal.connect(self.loadingProgressDialog.ui.progressBar_loading.setValue) # 프로그레스 바의 진행률을 표시하기위해 쓰레드내 작업의 진행률을 가져와서 프로그레스 바의 setValue와 연결
        self.progressBar_thread.result_ready_signal.connect(self.handle_result) # 프로그레스바의 설정을 위한 쓰레드내의 작업결과 변수들을 가져오기 위한 시그널, 이 변수들을 사용하기 위해 
                                                                           # show_team_status를 둘로 나눈뒤 절반을 self.handle_result 로 옮겨서 나머지 실행함.
        self.progressBar_thread.start()


        # 팀 스테이터스 창 팀과 팀장의 job 라벨 ( 프로그래스바 설정을 위한 쓰레드 셋업때문에 show_team_status 를 둘로 나누면서 남은 코드 - team,job변수가 이쪽이 있어서)
        self.teamStatusDialog.ui.label_teamName.setText(team)
        self.teamStatusDialog.ui.label_teamLeader.setText(job)




    # 프로젝트별 현황UI 중 완료된 프로젝트의 비져빌리티 옵션        
    def visibilityTeamStatusUI(self, state):

        for proj_ui in self.teamStatusDialog.projUI_list:
            app = int((proj_ui.ui.label_amount_shots.text()).split("/")[0])
            all = int((proj_ui.ui.label_amount_shots.text()).split("/")[1])

            if (app == all) and (state == 0):
                proj_ui.hide()

            elif  (state == 2):
                proj_ui.show()




    def on_toggle(self, checked):
        self.content.setVisible(checked)



    def getStatusTask(self, task, proj, department, taskInfo, member):

        # 택틱 bd 맨데이와 스케쥴bd맨데이가 다른 태스크의 데이타프레임 만들기
        tempTask_sr = None
        tempBid_path = currentPath + "/.temp_BDmanday/edited_mandayData.json" 
        if(os.path.exists(tempBid_path)):
            tempBid_df = pd.read_json(tempBid_path)
            tempTask_sr = tempBid_df[0]  # 


        status = ""
        contextFullName = ""
        bd_manday = 0.0
        act_manday = 0.0
        part = "" 

        if department == "ANI" : process = "animation"
        elif department == "MMV" : process = "matchmove"
        elif department == "RIG" : process = "creature"

        if "/" not in task:
            contextFullName = process

        elif "/" in task:
            taskContext = task.split("/")
            task = taskContext[0]
            context = taskContext[1]
            contextFullName = process + "/" + context 


        # 현재 열려있는 프로젝트
        if proj in self.projects:
            for taskData in taskInfo[member]:

                if process == "animation" : part = "A"
                elif process == "creature" : part = "R"
                elif process == "matchmove" : part = "M"

                if taskData["project_code"] == self.projects[proj][0] and taskData["extra_code"] == task and taskData["context"] == contextFullName:
                    # 편집된 BD맨데이가 있는지 확인했을때 없는경우
                    if (tempTask_sr is not None and task not in tempTask_sr.values) or (tempTask_sr is None):
                        status = taskData["status"]
                        bd_manday = taskData["bid_manday"]
                        act_manday = taskData["active_manday"]
                        """
                        if project != "testshot": ## 관리개발은 전체 맨데이 산정에서 제외시키기
                            if taskData["status"] != "Omit": # 오밋된 샷은 샷갯수 산정에서 제외시키기
                                all_shot += 1

                            if taskData["status"] == "Approved": 
                                app_shot += 1
                        """

                elif tempTask_sr is not None and task in tempTask_sr.values:
                    indices = tempBid_df[(tempBid_df[0] == task) & (tempBid_df[2] == part)].index.tolist() # 해당 태스크의 인덱스 가져오기                        
                    if indices:
                        bd_manday = float(tempBid_df.loc[indices[0]][1])


            if bd_manday == None:
                bd_manday = 0
            if act_manday == None:
                act_manday = 0

            return status, bd_manday, act_manday



        # 프로젝트가 종료되어 현재 프로젝트가 열려있지 않은경우
        elif proj not in self.projects:
            saved_mirrData = self.file_manager.import_Json(currentPath+"/.scheduleData_mirror", userID+"_scheduleMirror.json")
            for mirrTask in saved_mirrData[proj]:
                if mirrTask[1] == task:
                    mandays = mirrTask[2]

                    if "/" in str(mandays):
                        act = mandays.split("/")[0]
                        bd = mandays.split("/")[1]

                        if bd == "None":
                            bd_manday = 0
                        elif  bd != "None":
                            bd_manday = float(bd)

                        if act == "None":
                            act_manday = 0
                        elif act != "None":
                            act_manday = float(act)

                    elif "/" not in str(mandays):
                        bd_manday = 0
                        act_manday = 0

            return status, bd_manday, act_manday                    



    def get_shotWorkload_json(self, taskInfo, member, workload_json, department):

        workload_result_json = {}
        manday_result_json = {}

        if workload_json != {}:
            for year in workload_json:
                if workload_json[year] != {}:
                    for quarter in workload_json[year]:
                        year_quarter = str(year) + "/" + str(quarter)
                        all_shot = 0
                        app_shot = 0
                        all_manday = 0
                        used_manday = 0

                        if workload_json[year][quarter] != {}:
                            for proj in workload_json[year][quarter]:
                                if workload_json[year][quarter][proj] != []:
                                    for task in workload_json[year][quarter][proj]:

                                        status, bd_manday, act_manday = self.getStatusTask(task, proj, department, taskInfo, member)
                                        
                                        if status != "Omit":
                                            all_shot = all_shot+1
                                            if status == "Approved":
                                                app_shot = app_shot+1

                                            all_manday = float(all_manday) + float(bd_manday)
                                            used_manday = round((float(used_manday) + float(act_manday)),2)

                        workload_result_json[year_quarter] = str(app_shot) + "/" + str(all_shot)
                        manday_result_json[year_quarter] = str(used_manday) + "/" + str(all_manday)

        return workload_result_json, manday_result_json





    # 분기별 진행한 태스크의 갯수와 완료현황, 맨데이 현황을 가져옴
    def get_shotWorkload_byPeriod(self, taskInfo, member, periodWorkload):

        global jsonData

        all_shot = 0
        app_shot = 0

        workload_json = {}
        department = ""

        if periodWorkload != {}:
            for year in periodWorkload:
                workload_json[year] = {}

                if periodWorkload[year] != {}:
                    for quarter in periodWorkload[year]:
                        workload_json[year][quarter] = {}

                        if periodWorkload[year][quarter] != []:
                            tasks = {}                            
                            for data in periodWorkload[year][quarter]:

                                if data["tasks"][0] != {}:
                                    for proj in data["tasks"][0]:
                                        if proj != "testshot": # 관리개발은 포함시키지 않음
                                            department = data["department"]

                                            for task in data["tasks"][0][proj]:
                                                if proj not in workload_json[year][quarter]:
                                                    workload_json[year][quarter][proj] = [task]

                                                elif proj in workload_json[year][quarter]:
                                                    if task not in workload_json[year][quarter][proj]:
                                                        workload_json[year][quarter][proj].append(task)

        shotWorkload_json, mandayStatus_json = self.get_shotWorkload_json(taskInfo, member, workload_json, department)

        return shotWorkload_json, mandayStatus_json






    # 프로젝트의 진행샷 현황 통계데이타 가져오기
    def get_status_projectShot(self, taskInfo, project, member):

        global jsonData

        all_shot = 0
        app_shot = 0
        wip_shot = 0
        shotStatus = ""
        progressValue = 0.0
        wipValue = 0
        progressBar_setup = ""

        for taskData in taskInfo[member]:
            if taskData["project_code"] == self.projects[project][0]:
                if project != "testshot": ## 관리개발은 전체 맨데이 산정에서 제외시키기
                    if taskData["status"] != "Omit": # 오밋된 샷은 샷갯수 산정에서 제외시키기
                        all_shot += 1

                    if taskData["status"] == "Approved": 
                        app_shot += 1

                    if taskData["status"] == "In-Progress" or taskData["status"] == "Review" or taskData["status"] == "OK":
                        wip_shot += 1

            if all_shot == 0:
                progressValue = 0

            else:
                progressValue = (float(app_shot) / float(all_shot))*100
                wipValue = (float(wip_shot) / float(all_shot))*100

            shotStatus = str(int(app_shot)) + " / " + str(int(all_shot))

            progressBar_setup = self.setup_shot_progressBar(all_shot, app_shot)


        return shotStatus, progressValue, progressBar_setup, wipValue, wip_shot





    # 프로젝트의 맨데이 통계데이타 가져오기
    def get_status_projectManday(self, taskInfo, project, member):

        # 택틱 bd 맨데이와 스케쥴bd맨데이가 다른 태스크의 데이타프레임 만들기
        tempBid_path = currentPath + "/.temp_BDmanday/edited_mandayData.json" 
        if(os.path.exists(tempBid_path)):
            tempBid_df = pd.read_json(tempBid_path)
            tempTask_sr = tempBid_df[0]  # 

        else:
            tempBid_df = None
            tempTask_sr = None


        bid_proj = 0.0
        act_proj = 0.0
        progressValue_proj = 0
        projManday = ""

        for taskData in taskInfo[member]:
            
            task_code = taskData["extra_code"]
            process = taskData["process"]
            context = taskData["context"]

            part = ""
            if process == "animation" : part = "A"
            elif process == "creature" : part = "R"
            elif process == "matchmove" : part = "M"

            taskName = ""
            if "/" in context:
                contextName = context.split("/")[1]
                taskName = task_code +"/"+contextName

            elif "/" not in context:
                taskName =task_code

            if taskData["project_code"] == self.projects[project][0]:
                if project != "testshot": ## 관리개발은 전체 맨데이 산정에서 제외시키기
                    if taskData["status"] != "Omit": # 오밋된 샷은 샷갯수 산정에서 제외시키기

                        # 편집된 BD맨데이가 있는지 확인했을때 없는경우
                        if (tempTask_sr is not None and taskName not in tempTask_sr.values) or (tempTask_sr is None):
                            if taskData["bid_manday"] != None:
                                bid_proj += float(taskData["bid_manday"])

                        # 편집된 BD맨데이가 있는경우 편집된 맨데이를 가져와서 더해준다.
                        elif tempTask_sr is not None and taskName in tempTask_sr.values:
                            indices = tempBid_df[(tempBid_df[0] == taskName) & (tempBid_df[2] == part)].index.tolist() # 해당 태스크의 인덱스 가져오기                        
                            if indices:
                                bidManday = float(tempBid_df.loc[indices[0]][1])
                                bid_proj += bidManday


                        # act 맨데이 더해주기
                        if taskData["active_manday"] != None:
                            act_proj += float(taskData["active_manday"])

                        else:
                            act_proj += 0

                        if bid_proj == 0:
                            progressValue_proj = 0

                        else:
                            progressValue_proj = (float(act_proj) / float(bid_proj))*100
                            if progressValue_proj > 100:
                                progressValue_proj = 100

                        projManday = str(int(act_proj)) + " / " + str(int(bid_proj))

        mandayProgressBar_setup = self.setup_manday_progressBar(act_proj, bid_proj, progressValue_proj)

        return projManday, progressValue_proj, mandayProgressBar_setup, bid_proj 









    # 맨데이 프로그레스바 셋업
    def setup_manday_progressBar(self, act_manday, bd_manday, value_manday):

        if bd_manday == 0: 

            progressBar_setup =     ("""
                    QProgressBar {
                        border: 1px solid #333333;
                        border-radius: 0px;
                        text-align: center;
                        height: 15px;
                        font-size: 10px;
                        background-color: #535353;
                    }
                    
                """)

        #6b2c2f;
        elif value_manday > 90:

            progressBar_setup =     ("""
                    QProgressBar {
                        border: 1px solid #353535;
                        border-radius: 0px;
                        text-align: center;
                        height: 15px;
                        font-size: 10px;

                    }
                    
                    QProgressBar::chunk {
                        background-color: #6b2c2f;
                    }
                """)


        else:
            progressBar_setup =     ("""
                    QProgressBar {
                        border: 1px solid #353535;
                        border-radius: 0px;
                        text-align: center;
                        height: 15px;
                        font-size: 10px;
                        background-color: #303030;
                    }

                    QProgressBar::chunk {
                        background-color: #557055
                    }

                """)
        return progressBar_setup





    # 샷 진행률 프로그레스바  셋업
    def setup_shot_progressBar(self, app, all):

        if app != all: # approved의 수가 전체샷개수 보다 적을때(완료되지 않았을때)
            
            progressBar_setup =     ("""
                    QProgressBar {
                        border: 1px solid #353535;
                        border-radius: 0px;
                        text-align: center;
                        height: 15px;
                        font-size: 10px;
                        background-color: #303030

                    }
                    
                    QProgressBar::chunk {
                        background-color: #484970;
                    }
                """)

        else:
            progressBar_setup =     ("""
                    QProgressBar {
                        border: 1px solid #353535;
                        border-radius: 0px;
                        text-align: center;
                        height: 15px;
                        font-size: 10px;

                    }

                    QProgressBar::chunk {
                        background-color: #484970;
                    }

                """)

        return progressBar_setup


    # 트리뷰에서 팀원을 선택하거나 해제할때 샷리스뷰와 리스트뷰의 내용들을 갱신
    def handle_item_state_change(self, index):
        self.clickedAction_memberCheck(index)



    # 싱글클릭과 더블클릭을 구분짓기 위한 클릭딜레이 체크
    def on_item_clicked(self, index):
        self.pending_click = index

        #  ===  최상위 멤버만 클릭이 두번 되어지는 현상을 막기위해 
        itemData = index.data()

        if itemData == self.memberCache_[userID]:
            self.pending_click = None
        # ======================================================

        self.click_timer.start(DOUBLE_CLICK_DELAY)




    # 체크박스를 정확히 클릭하지 않고 이름을 클릭해도 태스크 리스트를 읽어오도록 함.
    def singleClick_select_member(self):#, index):

        if self.pending_click:

            item = self.teamTreeModel.itemFromIndex(self.pending_click)

            if item and self.checkArea_click == 0 : # 클릭한 인덱스에 아이템이 존재하고, 클릭한 위치가 텍스트영역일때
                current_state = item.checkState()
                item.setCheckState(Qt.Checked if current_state == Qt.Unchecked else Qt.Unchecked)

            elif item and self.checkArea_click == 1 : # 클릭한 인덱스에 아이템이 존재하고, 클릭한 위치가 체크박스 영역일때
                """
                체크박스 영역은 클릭하였을때 자동으로 on/off 가 되므로 토글방식을 적용하면 다시 돌아가게 된다. 
                그래서 토글방식을 적용하지 않고,다른 프로세스 없이 체크박스 영역이 선택되었음을 나타내는 변수만 다시 0 으로 리셋함.
                """
                self.checkArea_click = 0



    @Slot(int)
    def setListviewLineEdit(self, value): # 슬라이더의 value가 홀수만 출력되도록하며, lineEdit의 텍스트가 그값을 받아 표시되도록 함

        if value % 2 == 0:
            self.listViewNumLineEdit.setText(str(value+1)) 

        else:
            self.listViewNumLineEdit.setText(str(value))




    @Slot()
    def changeListViewUI_notTracking(self):

        scaleUI = 300/(int(self.listViewNumLineEdit.text())/3) 

        current_Listview_Num = self.dayUi.splitter.count()

        changed_ListView_Num = int(self.listViewNumLineEdit.text())
        number_add = changed_ListView_Num - current_Listview_Num

        if 300 < scaleUI < 800:
            fontScale = 12

        elif scaleUI <= 300:
            scaleUI = 300 # 리스트뷰의 최소크기
            fontScale = 12
            
        elif scaleUI >= 800:
            fontScale = 25


        if number_add > 0:
            self.addListView(number_add, current_Listview_Num, scaleUI)

        elif number_add < 0:
            self.removeListView(number_add, current_Listview_Num)         

        mm = self.listViewScroll.horizontalScrollBar().maximum()
        self.listViewScroll.horizontalScrollBar().setValue(mm)




    @Slot()
    def changeListViewUI_(self, text):

        self.listViewSlider.setValue(int(text))  

        scaleUI = 300/(int(self.listViewNumLineEdit.text())/3) 

        current_Listview_Num = self.dayUi.splitter.count()

        changed_ListView_Num = int(self.listViewNumLineEdit.text())
        number_add = changed_ListView_Num - current_Listview_Num

        if 300 < scaleUI < 800:
            fontScale = 12

        elif scaleUI <= 300:
            scaleUI = 300 # 리스트뷰의 최소크기
            fontScale = 12
            
        elif scaleUI >= 800:
            fontScale = 25


        if number_add > 0:
            self.addListView(number_add, current_Listview_Num, scaleUI)

        elif number_add < 0:
            self.removeListView(number_add, current_Listview_Num)         

        max_scroll = self.listViewScroll.horizontalScrollBar().maximum()
        self.listViewScroll.horizontalScrollBar().setValue(max_scroll)#self.listViewScroll.horizontalScrollBar().maximum())

          






    def addListView(self, num, currentNum, scaleUI):

        global jsonData

        half_val = int(num/2) # 추가되는 리스트뷰중 앞쪽 혹은 뒤쪽에 붙일 리스트뷰의 개수(전체추가되는 개수의 절반)
        half_existList = int(currentNum / 2) # 이미 존재하는 리스트뷰 개수기준으로 현자날짜기준 앞쪽 혹은 뒤쪽에 붙일 리스트뷰개수(전체개수의 절반)

        halfDay = QDate(self.rangeEndDate.year(), self.rangeEndDate.month(), self.rangeEndDate.day()).addDays(-1*half_existList)
        
        currentDay = halfDay.day()
        currentMonth = halfDay.month()
        currentYear = halfDay.year()

        num_all_listview = currentNum + num
        value_addDay = int((num_all_listview) / 2) 

        for i in range(num):

            listViewName = "listView_" + str(i+currentNum)

            num_addDay = i- half_val    
            dayNumber_add = 0

            if i < half_val:
                dayNumber_add = num_addDay-half_existList

            elif i >= half_val:
                dayNumber_add = num_addDay+half_existList+1

            weekDay = {1:'월', 2:'화', 3:'수', 4:'목', 5:'금', 6:'토', 7:'일'}
            label_year = QDate(int(currentYear), int(currentMonth), int(currentDay)).addDays(dayNumber_add).year()
            label_month = QDate(int(currentYear), int(currentMonth), int(currentDay)).addDays(dayNumber_add).month()
            label_day = QDate(int(currentYear), int(currentMonth), int(currentDay)).addDays(dayNumber_add).day()
            label_dow = QDate(int(currentYear), int(currentMonth), int(currentDay)).addDays(dayNumber_add).dayOfWeek()

            listview_labelName = str(label_month) + "/" + str(label_day) + " ( "+weekDay[label_dow]+" )"

            layout_parent = self.dayUi.splitter
            layout_child = listViewName + "_layout"

            newListview = lv.ScheduleListView(listview_labelName, listViewName, layout_parent, layout_child, scaleUI, self, self.manager, self.dayScheduleFrame, self)

            # 새로 생성한 리스트뷰에 드래그 드롭이 발생한경우, self.get_dropEvent 메서드를 실행시킴
            newListview.listView_day.itemsDropped.connect(self.get_dropEvent)

            # 리스트뷰내 아이템을 더블클릭하면 테스크 정보창이 생성
            newListview.listView_day.doubleClicked.connect(lambda index: self.openMedia(self.current_focused_list_view, index))

            # 리스트뷰에 컨텍스트 메뉴 연결
            newListview.listView_day.setContextMenuPolicy(Qt.CustomContextMenu)
            newListview.listView_day.customContextMenuRequested.connect(
                lambda pos, lv=newListview.listView_day: self.show_context_menu(lv, pos))


            # label_dow 가 토요일이나 일요일일경우 색깔변경
            if label_dow == 6 or label_dow == 7:
                newListview.listView_day.setStyleSheet("QListView { background-color: #232323; }")
            else:                
                newListview.listView_day.setStyleSheet("QListView { background-color: #2a2a2a; }")            

            # 스케쥴뷰의 리스트뷰의 갯수가 늘어날 경우 앞날짜의 리스트뷰와 뒷날짜의 리스트뷰가 늘어나게 되는데
            # 앞날짜의 리스트뷰와 뒷날짜의 리스트뷰가 모두 뒤로 붙지 않고 앞날짜는 앞에 뒷날짜는 뒤로 붙게 생성.
            if i < half_val:
                for j in range(currentNum):
                    widget = self.dayUi.splitter.widget(i)
                    exist_layout = widget.layout()

                    parentWidget = exist_layout.parentWidget()

                    self.dayUi.splitter.addWidget(parentWidget)

        newRangeEndDate = QDate(self.rangeEndDate.year(), self.rangeEndDate.month(), self.rangeEndDate.day()).addDays(half_val)
        newRangeStartDate = QDate(newRangeEndDate.year(), newRangeEndDate.month(), newRangeEndDate.day()).addDays(-1*(num+currentNum)+1)
       
        self.update_range_Label(newRangeStartDate, newRangeEndDate)
        self.rangeEndDate = newRangeEndDate

        listViewNum = self.dayUi.splitter.count()


        # 현재 ui에 표시되어있는 스케쥴 리스트뷰 가져오기(옮기기전)
        listViews=[]
        layouts = []
        for i in range(listViewNum):
            #layout = self.day_listViewLayout.itemAt(i)
            widget = self.dayUi.splitter.widget(i)
            layout = widget.layout()

            layouts.append(layout)

            if layout:
                frame = layout.itemAt(2).widget()
                frameLayout = frame.layout()
                listView = frameLayout.itemAt(0).widget()
                self.manager.Move_sideway_setContainer(frame, layout)

        self.refereshListViews(0, jsonData)
        self.ui_layout_manager.update_splitter_size() # 생성된 리스트뷰들의 갯수의 합보다 scroll area 의 크기를 더크게 셋팅하여 splitter 작동되도록 함
        self.ui_layout_manager.reset_splitter_size() # 스케쥴 리스트뷰의 스플리터 사이즈 리셋




    def _remove_layout(self, remItem):

        if isinstance(remItem, QWidget):
            layout = remItem.layout()
            if layout:
                self._remove_layout(layout)

            remItem.setParent(None) # 부모에서 제거
            remItem.deleteLater()

        elif isinstance(remItem, QHBoxLayout):
            while remItem.count():
                item = remItem.takeAt(0)

                if isinstance(item, QWidgetItem):
                    widget = item.widget()

                    if widget:
                        sub_layout = widget.layout() # 하위에 레이아웃이 있는지 확인
                        if sub_layout:
                            self._remove_layout(sub_layout) # 재귀적으로 하위 레이아웃 삭제

                        widget.setParent(None) # 부모에서 제거
                        widget.deleteLater()

            remItem.setParent(None) # 부모에서 제거            
            remItem.deleteLater()





    def removeListView(self, removeNum, currentNum):

        global jsonData

        half_val = abs(int(removeNum/2)) # 추가되는 리스트뷰중 앞쪽 혹은 뒤쪽에 삭제할 리스트뷰의 개수(전체삭제되는 개수의 절반)
        half_existList = int(currentNum / 2) # 이미 존재하는 리스트뷰 개수기준으로 현재날짜기준 앞쪽 혹은 뒤쪽의 리스트뷰개수(전체개수의 절반)

        widgets_to_remove = []
        for i in range(abs(removeNum)):
            if  i < half_val:
                widgets_to_remove.append(self.dayUi.splitter.widget(i))

            if i >= half_val:
                numLayout = self.dayUi.splitter.count()  
                widgets_to_remove.append(self.dayUi.splitter.widget((numLayout-1)-(i-half_val)))

        if widgets_to_remove:

            widgets_to_check = widgets_to_remove[:]
            for widget in widgets_to_remove:
                self._remove_layout(widget)

            QApplication.processEvents() # 삭제 이벤트 처리
            widgets_to_remove.clear() # 참조 제거

        newRangeEndDate = QDate(self.rangeEndDate.year(), self.rangeEndDate.month(), self.rangeEndDate.day()).addDays(-1*half_val)
        newRangeStartDate = QDate(newRangeEndDate.year(), newRangeEndDate.month(), newRangeEndDate.day()).addDays(-1*(removeNum+currentNum)+1)
       
        self.update_range_Label(newRangeStartDate, newRangeEndDate)
        self.rangeEndDate = newRangeEndDate

        self.ui_layout_manager.reset_splitter_size() # 스케쥴 리스트뷰의 스플리터 사이즈 리셋



    def get_TaskInfo_wip(self, item, taskInfoJson, project):

        # 택틱 bd 맨데이와 스케쥴bd맨데이가 다른 태스크의 데이타프레임 만들기
        tempBid_path = currentPath + "/.temp_BDmanday/edited_mandayData.json" 

        if(os.path.exists(tempBid_path)):
            tempBid_df = pd.read_json(tempBid_path)
            tempTask_sr = tempBid_df[0]

        else:
            tempBid_df = None
            tempTask_sr = None



        if "/" in item[1]:
            task_name = item[1].split('/')[0]
            context_name = item[1].split('/')[1]

        else:
            task_name = item[1]
            context_name = None

        #트리뷰를 새로고침하고 전체 팀원정보를 읽어옴
        teamInfo = teamInfo_

        artist_task = None
        artist_kr = ""
        projTitle = ""
        process = ""

        artistList = list(taskInfoJson.keys())
        project_code = self.projects[project][0]

        for artist in artistList:
            for task in  taskInfoJson[artist]:
                if context_name == None:
                    if (task["extra_code"] == task_name) and project_code == task["project_code"] and task["context"] == task["process"]:
                        process = task["process"]
                        artist_task = artist
                    
                        if artist == userID:
                            artist_kr = self.currName_kr
                        else:
                            artist_kr = teamInfo[0][artist]["nameKr"]

                else:
                    if (task["extra_code"] == task_name and task["context"] == task["process"]+"/"+context_name) and project_code == task["project_code"]:
                        artist_task = artist
                        #project = connect_shot_proj[item]
                        process = task["process"]

                        if artist == userID:
                            artist_kr = self.currName_kr
                        else:
                            artist_kr = teamInfo[0][artist]["nameKr"]                    

        start=""
        end=""
        shotManday = ""
        projManday = ""
        status = ""
        identifier = "noContext"

        progressValue_shot = 0.0
        progressValue_proj = 0.0

        if artist_task != None :

            bid_proj = 0.0
            act_proj = 0.0
            taskName = item[1]

            if "/" in taskName: # 군중,페이셜등의 context 식별자 설정
                identifier = "haveContext"

            for taskData in taskInfoJson[artist_task]:
                # facial, crowd 등의 컨텍스트 식별
                if identifier == "haveContext":
                    task = taskName.split('/')[0]
                    context = taskData["process"]+"/" + taskName.split('/')[1]

                else: 
                    task = taskName
                    context = taskData["process"]

                if project.lower() in list(self.projects.keys()):
                    projTitle = self.projects[project.lower()][1]


                # 샷 맨데이 설정
                bid = 0.0
                act = 0.0

                if (taskData["extra_code"] == task and taskData["context"] == context) and project_code == taskData["project_code"]:
                    indices = tempBid_df[(tempBid_df[0] == item[1]) & (tempBid_df[2] == item[4])].index.tolist() # 해당 태스크의 인덱스 가져오기

                    if indices == []:
                        if taskData["start_date"] != None:
                            start = taskData["start_date"].split(" ")[0]
                        else: start = None                        

                        if taskData["end_date"] != None:
                            end = taskData["end_date"].split(" ")[0]
                        else: end = None

                        bid = taskData["bid_manday"]
                        act = taskData["active_manday"]
                        shotManday = str(act) + " / " + str(bid)

                    elif indices != []:
                        if taskData["start_date"] != None:
                            start = taskData["start_date"].split(" ")[0]
                        else: start = None                        

                        if taskData["end_date"] != None:
                            end = taskData["end_date"].split(" ")[0]
                        else: end = None

                        taticMandays = item[2]
                        taticAct = str(taticMandays.split("/")[0])
                        editBD = str(tempBid_df.loc[indices[0]][1]) # 편집된 BD 맨데이 가져오기               

                        newMandays = taticAct + " / " + editBD
                        shotManday = newMandays


                    status = taskData["status"]
                    shot_act = shotManday.split("/")[0]
                    shot_bid = shotManday.split("/")[1]

                    if shot_bid == " None": shot_bid = None

                    if shot_bid != None and shot_act != None:
                        if float(shot_bid) != 0:
                            progressValue_shot = (float(shot_act) / float(shot_bid))*100
                            if progressValue_shot > 100:
                                progressValue_shot = 100

                        elif bid == 0:
                            progressValue_shot = 0                            

                    else:
                        progressValue_shot = 0



                # 프로젝트 맨데이 설정
                if taskData["project_code"] == self.projects[project.lower()][0]:

                    taskContext = ""

                    if taskData["process"] == "animation":
                        if "/" in taskData["context"]:
                            context = taskData["context"].split("/")[1]
                            taskContext = taskData["extra_code"] + "/" + context

                        elif taskData["context"] == "animation":
                            taskContext = taskData["extra_code"]


                    elif taskData["process"] == "matchmove":
                        if "/" in taskData["context"]:
                            context = taskData["context"].split("/")[1]
                            taskContext = taskData["extra_code"] + "/" + context

                        elif taskData["context"] == "matchmove":
                            taskContext = taskData["extra_code"]


                    elif taskData["process"] == "creature":
                        if "/" in taskData["context"]:
                            context = taskData["context"].split("/")[1]
                            taskContext = taskData["extra_code"] + "/" + context

                        elif taskData["context"] == "creature":
                            taskContext = taskData["extra_code"]


                    if taskData["process"] == "animation": part = "A"
                    elif taskData["process"] == "creature": part = "G"
                    elif taskData["process"] == "matchmove": part = "M"

                    # bd맨데이를 수정한 태스크인 경우 edited_mandayData.json에서 bd맨데이를 가져옴
                    if tempTask_sr is not None and taskContext in tempTask_sr.values:
                        indices = tempBid_df[(tempBid_df[0] == taskContext) & (tempBid_df[2] == part)].index.tolist() # 해당 태스크의 인덱스 가져오기                        
                        if indices:
                            bidManday = float(tempBid_df.loc[indices[0]][1])
                            bid_proj += bidManday

                    #elif taskContext not in tempTask_sr.values:
                    elif (tempTask_sr is not None and taskContext not in tempTask_sr.values) or (tempTask_sr is None):
                        if taskData["bid_manday"] != None:
                            bid_proj += float(taskData["bid_manday"])

                    if taskData["active_manday"] != None:
                        act_proj += float(taskData["active_manday"])

                    else:
                        act_proj += 0

                    if bid_proj == 0:
                        progressValue_proj = 0

                    else:
                        progressValue_proj = (float(act_proj) / float(bid_proj))*100
                        if progressValue_proj > 100:
                            progressValue_proj = 100

                    projManday = str(round(float(act_proj),2)) + " / " + str(round(float(bid_proj),2))

        return start, end, shotManday, projManday, progressValue_shot, progressValue_proj, status, artist_task, artist_kr, projTitle, process




    def get_TaskInfo(self, item, taskInfoJson):

        if "/" in item[1]:
            task_name = item[1].split('/')[0]
            context_name = item[1].split('/')[1]

        else:
            task_name = item[1]
            context_name = None

        #트리뷰를 새로고침하고 전체 팀원정보를 읽어옴
        teamInfo = teamInfo_

        artist_task = None
        artist_kr = ""
        project = ""
        projTitle = ""
        process = ""

        artistList = list(taskInfoJson.keys())

        for artist in artistList:
            for task in  taskInfoJson[artist]:

                if context_name == None:
                    if task["extra_code"] == task_name:
                        process = task["process"]
                        artist_task = artist
                        project = connect_shot_proj[item]
                    
                        if artist == userID:
                            artist_kr = self.currName_kr
                        else:
                            artist_kr = teamInfo[0][artist]["nameKr"]

                else:
                    if task["extra_code"] == task_name and task["context"] == task["process"]+"/"+context_name:
                        artist_task = artist
                        project = connect_shot_proj[item]
                        process = task["process"]

                        if artist == userID:
                            artist_kr = self.currName_kr
                        else:
                            artist_kr = teamInfo[0][artist]["nameKr"]                    

        start=""
        end=""
        shotManday = ""
        status = ""
        identifier = "noContext"
        progressValue_shot = 0.0
        progressValue_proj = 0.0

        if artist_task != None :
            bid_proj = 0.0
            act_proj = 0.0
            taskName = item[1]

            if "/" in taskName: # 군중,페이셜등의 context 식별자 설정
                identifier = "haveContext"

            for taskData in taskInfoJson[artist_task]:

                # facial, crowd 등의 컨텍스트 식별
                if identifier == "haveContext":
                    task = taskName.split('/')[0]
                    context = taskData["process"]+"/" + taskName.split('/')[1]
                else: 
                    task = taskName
                    context = taskData["process"]

                if project.lower() in list(self.projects.keys()):
                    projTitle = self.projects[project.lower()][1]

                # 샷 맨데이 설정
                if taskData["extra_code"] == task and taskData["context"] == context:

                    if taskData["start_date"] != None:
                        start = taskData["start_date"].split(" ")[0]
                    else: start = None                        

                    if taskData["end_date"] != None:
                        end = taskData["end_date"].split(" ")[0]
                    else: end = None

                    bid = taskData["bid_manday"]
                    act = taskData["active_manday"]
                    status = taskData["status"]

                    if bid != None and act != None:
                        if float(bid) != 0:
                            progressValue_shot = (float(act) / float(bid))*100
                            if progressValue_shot > 100:
                                progressValue_shot = 100
                        elif bid == 0:
                            progressValue_shot = 0                            

                    else:
                        progressValue_shot = 0

                    shotManday = str(act) + " / " + str(bid)


                # 프로젝트 맨데이 설정
                if taskData["project_code"] == self.projects[project.lower()][0]:
                    if taskData["bid_manday"] != None:
                        bid_proj += float(taskData["bid_manday"])

                    if taskData["active_manday"] != None:
                        act_proj += float(taskData["active_manday"])
                    else:
                        act_proj += 0

                    if bid_proj == 0:
                        progressValue_proj = 0
                    else:
                        progressValue_proj = (float(act_proj) / float(bid_proj))*100
                        if progressValue_proj > 100:
                            progressValue_proj = 100

                    projManday = str(int(act_proj)) + " / " + str(int(bid_proj))

        return start, end, shotManday, projManday, progressValue_shot, progressValue_proj, status, artist_task, artist_kr, projTitle, process





    # 리스트뷰내 아이템 더블클릭시 해당태스크 디테일정보를 띄우는 다이얼로그 창
    def openTaskInfoWin(self,listView, index):

        # 이미 실행되어있는 중복되는 창 삭제
        if self.taskInfomationDialog is not None:
            self.taskInfomationDialog.close()
            self.taskInfomationDialog.deleteLater()
            self.taskInfomationDialog = None

        taskInfoJson = self.file_manager.import_Json(currentPath+"/.task_info", userID+"_task_info.json")
        self.taskInfomationDialog = taskInfoDialog.TaskInfoDialog(self)
        self.taskInfomationDialog.setWindowTitle("Task Information")        

        start=""
        end=""
        shotManday = ""
        status = ""
        identifier = "noContext"
        self.progressValue_shot = 0.0
        self.progressValue_proj = 0.0

        if isinstance(listView, QListView):
            item = listView.model().data(index, Qt.DisplayRole)

            if "/" in item[1]:
                task_name = item[1].split('/')[0]
                context_name = item[1].split('/')[1]
            else:
                task_name = item[1]
                context_name = None

            project = connect_shot_proj[item]
            start, end, shotManday, projManday, progressValue_shot, progressValue_proj, status, artist, artistKr, projTitle, process = self.get_TaskInfo_wip(item, taskInfoJson, project)

            # 현재 태스크가 속한 프로젝트의 샷진행 현황 가져오기
            shotStatus, shotProgressValue, progressBar_set, wipValue, wip_shots = self.get_status_projectShot(taskInfoJson, project, artist ) 

            # 태스크 난이도 읽어오기
            grade = self.get_taskScore(project, process, task_name, context_name, artist)

            # 태스크 평가등급 리스트
            eval_list = ["Excellent", "Good", "Weak"]

            # 샷맨데이 프로그레스바 스타일셋업
            shot_act_manday = shotManday.split(" / ")[0]
            shot_bid_manday = shotManday.split(" / ")[1]

            if shot_bid_manday == "None":
                shot_bid_manday = 0

            else: shot_bid_manday = float(shot_bid_manday)

            shot_progressBar_style = self.setup_manday_progressBar(shot_act_manday, shot_bid_manday, progressValue_shot)

            # 프로젝트맨데이 프로그레스바 스타일셋업
            proj_act_manday = float(projManday.split(" / ")[0])
            proj_bid_manday = float(projManday.split(" / ")[1])

            if proj_bid_manday == "None":
                proj_bid_manday = 0
            else: proj_bid_manday = float(proj_bid_manday)

            proj_progressBar_style = self.setup_manday_progressBar(proj_act_manday, proj_bid_manday, progressValue_proj)


            # 다이얼로그 창, 텍스트 색상 설정
            self.taskInfomationDialog.setStyleSheet("""
                QDialog {
                    background-color: rgb(70,70,70);
                    color: black;
                }
                QLabel {
                    color : rgb(160,160,160);
                }
            """)



        if self.job == "팀원":
            self.taskInfomationDialog.ui.frame_grade.hide()
            self.taskInfomationDialog.ui.label_grade.hide()

        if int(shot_bid_manday) == 0:
            self.taskInfomationDialog.ui.progressBar_shotManday.setTextVisible(False)

        if int(proj_bid_manday) == 0:
            self.taskInfomationDialog.ui.progressBar_projManday.setTextVisible(False)

        self.taskInfomationDialog.ui.combo_evaluation.hide()
        self.taskInfomationDialog.ui.frame_value.hide()
        self.taskInfomationDialog.ui.combo_evaluation.addItems(eval_list)
        self.taskInfomationDialog.ui.label_taskName.setText(item[1])
        self.taskInfomationDialog.ui.label_projectName.setText(projTitle)
        self.taskInfomationDialog.ui.label_artistName.setText(artistKr)        
        self.taskInfomationDialog.ui.label_startDate.setText(start)        
        self.taskInfomationDialog.ui.label_endDate.setText(end)
        self.taskInfomationDialog.ui.label_grade.setText(grade)     
        self.taskInfomationDialog.ui.shot_act_bid.setText(shotManday)   
        self.taskInfomationDialog.ui.proj_act_bid.setText(projManday)   
        self.taskInfomationDialog.ui.progressBar_shotManday.setValue(progressValue_shot)
        self.taskInfomationDialog.ui.progressBar_projManday.setValue(progressValue_proj) 
        self.taskInfomationDialog.ui.progressBar_shotManday.setStyleSheet( shot_progressBar_style )        
        self.taskInfomationDialog.ui.progressBar_projManday.setStyleSheet( proj_progressBar_style )                
        #self.taskInfomationDialog.ui.progressBar_projManday.setMaximumHeight(5) # 프로젝트 맨데이 프로그래스바의 최대두께
        self.taskInfomationDialog.ui.progressBar_projManday.setTextVisible(False) 
        self.taskInfomationDialog.ui.progressBar_projShots.setTextVisible(False)         
        self.taskInfomationDialog.ui.progressBar_projShots.setStyleSheet(progressBar_set)         

        self.taskInfomationDialog.ui.proj_app_all.setText(shotStatus)   
        self.taskInfomationDialog.ui.progressBar_projShots.setValue(shotProgressValue) 

        self.taskInfomationDialog.ui.label_status.setText(status)   


        if status == "Approved":
            self.taskInfomationDialog.ui.frame_status.setStyleSheet(u"background-color: rgb(52, 101, 164);")
            self.taskInfomationDialog.ui.label_status.setStyleSheet("color: white;")   

        elif status == "Retake":
            self.taskInfomationDialog.ui.frame_status.setStyleSheet("background-color: rgb(201,33,30);")
            self.taskInfomationDialog.ui.label_status.setStyleSheet("color: white;")   

        elif status == "In-Progress":
            self.taskInfomationDialog.ui.frame_status.setStyleSheet("background-color: rgb(138, 226, 52);")
            self.taskInfomationDialog.ui.label_status.setStyleSheet("color: white;")   

        elif status == "OK":
            self.taskInfomationDialog.ui.frame_status.setStyleSheet("background-color: rgb(129,182,239);")
            self.taskInfomationDialog.ui.label_status.setStyleSheet("color: white;")   

        elif status == "Ready":
            self.taskInfomationDialog.ui.frame_status.setStyleSheet("background-color: rgb(213,213,119);")
            self.taskInfomationDialog.ui.label_status.setStyleSheet("color: white;")   

        elif status == "Omit":
            self.taskInfomationDialog.ui.frame_status.setStyleSheet("background-color: black;")
            self.taskInfomationDialog.ui.label_status.setStyleSheet("color: white;")   

        elif status == "Hold":
            self.taskInfomationDialog.ui.frame_status.setStyleSheet("background-color: gray;")
            self.taskInfomationDialog.ui.label_status.setStyleSheet("color: white;")   

        elif status == "Review":
            self.taskInfomationDialog.ui.frame_status.setStyleSheet("background-color: rgb(78, 154, 6);")
            self.taskInfomationDialog.ui.label_status.setStyleSheet("color: white;")   


        self.taskInfomationDialog.show()





    # 초기 스케쥴 리스트뷰 UI생성
    def setUp_listViewUI(self):

        numberListView = int(self.listViewNumLineEdit.text())

        today = QDate.currentDate()
        monday = today.addDays(-today.dayOfWeek()+1)
        friday = monday.addDays(4) 

        value_addDay = int(numberListView / 2)

        range_date = []
        for i in range(numberListView):
            currentValue_day = i - value_addDay
            weekDay = {1:'월', 2:'화', 3:'수', 4:'목', 5:'금', 6:'토', 7:'일'}

            if currentValue_day==0:
                listViewName = "listView_0"

            elif currentValue_day<0:
                listViewName = "listView_L_" + str(abs(currentValue_day))

            elif currentValue_day>0:
                listViewName = "listView_R_" + str(currentValue_day) 

            date = monday.addDays(i)
            label_dow = date.dayOfWeek()
            labelName = str(date.month()) + "/" + str(date.day()) + " ( " + weekDay[label_dow] + " )"
            label_date = QDate(date.year(), date.month(), date.day())

            #layout_parent = self.day_listViewLayout
            layout_parent = self.dayUi.splitter
            layout_child = listViewName + "_layout"

            # 리스트뷰 생성
            newListview = lv.ScheduleListView(labelName, listViewName, layout_parent, layout_child, 300, self, self.manager, self.dayScheduleFrame, self)

        
            # 새로 생성한 리스트뷰에 드래그 드롭이 발생한경우, self.get_dropEvent 메서드를 실행시킴
            newListview.listView_day.itemsDropped.connect(self.get_dropEvent)

            font = QFont()
            font.setPointSize(12)
            newListview.listView_day.setFont(font)

            if i==0: 
                range_date.append(label_date)

            elif i==numberListView-1: 
                range_date.append(label_date)


            # 리스트뷰내 아이템을 더블클릭하면 테스크 정보창이 생성
            newListview.listView_day.doubleClicked.connect(lambda index: self.openMedia(self.current_focused_list_view, index))#new_listView, index))

            # 초기 리스트뷰 생성에 컨텍스트 메뉴 연결
            newListview.listView_day.setContextMenuPolicy(Qt.CustomContextMenu)
            newListview.listView_day.customContextMenuRequested.connect(
                lambda pos, lv=newListview.listView_day: self.show_context_menu(lv, pos))

            # label_dow 가 토요일이나 일요일일경우 색깔변경
            if label_dow == 6 or label_dow == 7:
                newListview.listView_day.setStyleSheet("QListView { background-color: #232323; }")
            else:                
                newListview.listView_day.setStyleSheet("QListView { background-color: #2a2a2a; }")  

        self.update_range_Label(range_date[0], range_date[1])
        self.rangeEndDate = range_date[1] # ui 처음 생성시 전역변수 rangeEndDate에 enddate 저장



        #######################################################################################
        # 처음 ui를  켰을때 현재 등록시킨 스케쥴을 표시하게 하기위해 현재기준 앞뒤 한주씩의 기간 설정
        today = QDate.currentDate()
        monday = today.addDays(-today.dayOfWeek()+1)
        friday = monday.addDays(4) 
        lastWeek_monday = today.addDays(-7+(-today.dayOfWeek()+1))
        nextWeek_friday = today.addDays(11+(-today.dayOfWeek()+1))

        betweenDate = {}
        day_count = 0
        date = lastWeek_monday


        if jsonData != []:

            json_df = pd.DataFrame(jsonData) # jsonData 파일을 판다스 데이타프레임으로 변환
            artist_df = json_df.set_index('artist') # artist 열을 인덱스로 가지는 데이타프레임으로 변환
            user_df = artist_df.loc[artist_df.index == userID] # 아티스트 데이타프레임에서 현재 사용자의 데이타만 뽑은 데이타프레임으로 변환

            ########################################################################################
            # 지난주 월요일부터 다음주 금요일까지의 어싸인된 스케쥴을 확인후 해당 프로젝트만 선택되게 하기

            #오늘날짜 기준 지난주 월요일부터 다음주 금요일까지의 기간의 데이타프레임 제작
            dates = []
            while date <= nextWeek_friday:
                dateList = {}

                dateList["year"] = date.year()
                dateList["month"] = date.month()
                dateList["day"] = date.day()                        

                if dateList not in dates:
                    dates.append(dateList)

                date = date.addDays(1)

            betweenDate_df = pd.DataFrame(dates)
            columns_compare = ["year", "month", "day"] # 현재유저의 스케쥴어싸인 데이타프레임과 3주간의 날짜를 비교하여 일치하는것을 찾기위해 비교할 날짜 컬럼 리스트
            mask1 = betweenDate_df[columns_compare].apply(tuple, axis=1)
            mask2 = user_df[columns_compare].apply(tuple, axis=1)
            commonValues = set(mask1).intersection(set(mask2)) # 두마스크의 일치되는 것만 뽑아낸 set
            recentWork_df = user_df[mask2.isin(commonValues)] # 일치하는 날짜에 해당하는 스케쥴어싸인 데이타프레임
            userTasks = recentWork_df['tasks'] # task컬럼만 시리즈로 추출

            # jsonData 안의 현재 유저의 스케쥴 데이타들중 task부분만 추출하여 set연산자를 통해 현재 진행중인 프로젝트 이름만 추출
            projects = set()
            for task in userTasks: # 시리즈안의 프로젝트만 set 연산자를 통해 뽑아냄
                proj = set(task[0])
                projects = projects | proj

            # 추출된 프로젝트이름(코드)를 ui상 표시된 이름으로 변환하여 리스트에 저장
            projLong_list = []
            for projCode in projects:
                projName = self.projects[projCode][1]
                if projName not in projLong_list:
                    projLong_list.append(projName)

            # 뽑아낸 프로젝트 이름을 이용하여 UI상의 프로젝트 리스트뷰에서 해당 프로젝트를 선택
            proj_model = self.projListview.model()

            for row in range(proj_model.rowCount()):
                index = proj_model.index(row, 0)
                text_proj = proj_model.data(index)
                if text_proj in projLong_list:
                    self.projSelection_model.select(index, QItemSelectionModel.Select)


        # 현재 리스트뷰에서 오늘에 해당하는 날짜를 하이라이트 시키기
        numListview = self.listViewNumLineEdit.text()

        for i in range(int(numListview)):
            widget = self.dayUi.splitter.widget(i)
            layout = widget.layout()

            if layout:
                frame = layout.itemAt(2).widget()
                dateIndex = layout.itemAt(0).widget()
                date = dateIndex.text()
                dateSplit = date.split(" (")
                month = str(today.month())
                day = str(today.day())
                todayString = month + "/" + day

                if dateSplit[0] == todayString:
                    widget.setStyleSheet("background-color: #38613b;")
                    self.manager.set_active_container(frame, layout, False)

        self.ui_layout_manager.update_splitter_size() # 생성된 리스트뷰들의 갯수의 합보다 scroll area 의 크기를 더크게 셋팅하여 splitter 작동되도록 함





    def update_range_Label(self, startDate, endDate):
        
        startDay = startDate.day()
        startMonth = startDate.month()

        endDay = endDate.day()
        endMonth = endDate.month()

        self.startDay.setText(str(startDay))
        self.startMonth.setText(str(startMonth))

        self.comboDay.setCurrentText(str(endDay))
        self.comboMonth.setCurrentText(str(endMonth))

    


    def eventFilter(self, source, event):
        return self.event_manager.handle_event_filter(source, event)




    def makeListView(self, label, listView, parentLayout, childLayout, scaleUI):

        listViewLayout = QVBoxLayout()
        listViewLayout.setObjectName(childLayout)
        label_day = QLabel(self.dayScheduleFrame)
        label_day.setObjectName(label)
        label_day.setText(label)
        font_label = QFont()
        font_label.setPointSize(10)
        font_label.setBold(True)
        font_label.setWeight(75)
        label_day.setFont(font_label)
        listViewLayout.addWidget(label_day, 0, Qt.AlignHCenter)

        listView_day = ddlv.dropListView(self.dayScheduleFrame)
        listView_day.setEditTriggers(QAbstractItemView.NoEditTriggers) # 더블클릭시에 에딧모드로 들어가지 않도록 함
        listView_day.setObjectName(listView)
        listView_day.setMinimumSize(scaleUI, 0)
        listView_day.installEventFilter(self)

        font_listView = QFont()
        font_listView.setPointSize(9)
        listView_day.setFont(font_listView)
        listViewLayout.addWidget(listView_day)
        parentLayout.addLayout(listViewLayout)

        return listView_day, listViewLayout




    # 리스트뷰 내 아이템의 리스트 가져오기 / jsonData내에 있지만 리스트뷰에 보여지지 않는 아이템들도 가져오기
    def getListViewItem_inJson(self, listView, members):

        num_listview = self.dayUi.splitter.count()

        label_index = 0
        listView_list=[]
        for i in range(num_listview):

            widget = self.dayUi.splitter.widget(i)
            listview_layout = widget.layout()

            if listview_layout:
                frame = listview_layout.itemAt(2).widget()
                frameLayout = frame.layout()
                listview_widget = frameLayout.itemAt(0).widget()
                listView_list.append(listview_widget)

        date_labels = self.getLabelData()

        listView_date = ''
        for i in range(len(listView_list)):
            if listView.objectName() == (listView_list[i].objectName()):
                listView_date = date_labels[i]

        listView_year = listView_date.split('/')[0]
        listView_month = listView_date.split('/')[1]
        listView_day = listView_date.split('/')[2]

        taskList = []
        for data in jsonData: 
            task_dep = {}
            if str(data['year']) == listView_year and str(data['month']) == listView_month and str(data['day']) == listView_day:
                if data['tasks']:  
                    for proj in data['tasks'][0]:
                        for task in data['tasks'][0][proj]:
                            task_proj_part = {}
                            proj_part = []

                            proj_part.append(proj)
                            proj_part.append(data["department"])

                            task_proj_part[task] = proj_part

                            if task_proj_part not in taskList:
                                taskList.append(task_proj_part)

        tuple_tasks = self.convert_to_tuple(taskList, members)

        return (tuple_tasks)





    def getListViewItem(self, listView):

        model = listView.model()
        allIndex = []
        num = range(model.rowCount(QModelIndex()))

        for row in range(model.rowCount(QModelIndex())):
            index = model.index(row)
            allIndex.append(index)

        currItems = []
        for index in allIndex:
            item = listView.model().data(index, Qt.DisplayRole)

            if item not in currItems:
                currItems.append(item)

        return currItems




    # 리스트뷰내 선택되어진 아이템들의 리스트 가져오기
    def getListView_selectedItem(self, listView):
        """리스트뷰 선택된 아이템 조회 (UIController 사용)"""
        return self.ui_controller.get_listview_selected_items(listView)




    def searchEmptyProj(self, x):

        empty_proj = [key for key in x[0] if len(x[0][key]) == 0]

        if empty_proj == []:
            return x

        elif empty_proj != []:
            for proj in empty_proj:
                del x[0][proj]
                
            return x


    # 스케쥴 제이슨파일을 데이타프레임으로 바꾼뒤 클린업을 한후 데이타프레임 형식으로 반환
    def get_DataFrame_sechedule(self, userPath):
        """스케줄 DataFrame 조회 (원본 로직 유지)"""
        df = pd.read_json(userPath)

        if not df.empty:
            df['tasks'] = df['tasks'].apply(lambda x: self.searchEmptyProj(x))
            df_importedJson = df.loc[df['tasks'].apply(lambda x:x != [{}])]        

            return df_importedJson

        else: # 스케쥴이 존재하지 않는 멤버의 경우 빈 데이타프레임 그대로 반환
            return df



    # 스케쥴 데이타 가져오기
    def getJsonData(self):

        global teamTreeHierarchy
        global jsonData

        jsonFile_user = os.path.join(jsonFilePath, userID+"_Data.json")

        if(os.path.exists(jsonFile_user)):
            df_importedJson = self.get_DataFrame_sechedule(jsonFile_user)

            # 읽어온 스케쥴데이타의 데이타프레임을 딕셔너리로 바꾼뒤 jsonData에 추가
            jsonData.extend(df_importedJson.to_dict(orient='records'))

        # 현재 멤버트리뷰의 루트아이템을 읽어들인뒤 그 하위의 체크되어있는 모든 멤버들을 리스트로 만들고 영어이름으로 변환
        invisibleRoot = self.teamTreeModel.invisibleRootItem()
        rootItem = invisibleRoot.child(0)

        checked_items_list = []
        unchecked_items_list = []
        checked_items_list_Kr = []
        self.get_CheckedItems(rootItem, checked_items_list_Kr)
        #print (checked_items_list_Kr)

        checked_items_list_En = []
        teamInfo = self.edit_teamTree_item(self.teamTreeModel.invisibleRootItem())

        if teamInfo != [] and teamInfo != None:
            memberList = list(teamInfo[0].keys())

            saved_all_JsonData = []
            for member in memberList:

                # 하위멤버들의 기존 저장된 스케쥴을 jsonData 에 저장
                jsonFile_member = os.path.join(jsonFilePath, member+"_Data.json")

                if(os.path.exists(jsonFile_member)):
                    df_importedJson_mem = self.get_DataFrame_sechedule(jsonFile_member)

                    # 읽어온 스케쥴데이타의 데이타프레임을 딕셔너리로 바꾼뒤 jsonData에 추가
                    if not df_importedJson_mem.empty :
                        jsonData.extend(df_importedJson_mem.to_dict(orient='records'))




    # 팀트리 창에서 현재 선택되어있는 팀원들의 리스트를 가져오기
    def getCheckedMember(self):

        checkedMember = []
        invisibleRoot = self.teamTreeModel.invisibleRootItem()
        rootItem = invisibleRoot.child(0)
        teamInfo = self.team_manager.get_team_info(userID)

        checkedMembers = []
        self.get_CheckedItems(rootItem, checkedMembers)

        EN_members=[]
        if rootItem.checkState() == Qt.Checked: #현재유저가 체크가 되어있으면 영어리스트에 현재유저를 먼저추가함.
            EN_members.append(userID)

        if teamInfo != []:
            for checkedMember in checkedMembers:
                for member in teamInfo[0]:
                    if teamInfo[0][member]["nameKr"] == checkedMember:
                        EN_members.append(member)

        return EN_members




    def getLabelData(self):

        num_listview = self.dayUi.splitter.count()
        rangeHalfDate = QDate(self.rangeEndDate.addDays(-1*(num_listview)/2))
        numberListView = int(self.listViewNumLineEdit.text())

        currentDate = rangeHalfDate

        currentDay = currentDate.day()
        currentMonth = currentDate.month()
        currentYear = currentDate.year()
        value_addDay = int(num_listview / 2)

        label_data = []
        for i in range(num_listview):

            currentValue_day = i - value_addDay
            weekDay = {1:'월', 2:'화', 3:'수', 4:'목', 5:'금', 6:'토', 7:'일'}

            label_year = QDate(int(currentYear), int(currentMonth), int(currentDay)).addDays(currentValue_day).year()
            label_month = QDate(int(currentYear), int(currentMonth), int(currentDay)).addDays(currentValue_day).month()
            label_day = QDate(int(currentYear), int(currentMonth), int(currentDay)).addDays(currentValue_day).day()
            label_dow = QDate(int(currentYear), int(currentMonth), int(currentDay)).addDays(currentValue_day).dayOfWeek()
            label_date = QDate(label_year, label_month, label_day)

            listViewName = "listView_" + str(i)
            labelName = str(label_year) + "/" + str(label_month) + "/" + str(label_day)
            label_data.append(labelName)

        return label_data            




    # 드래그 드랍 아이템 시그널 받아오기
    def get_dropEvent(self, dropped_items, listView_name):
        self.event_manager.handle_drop_event(dropped_items, listView_name)





    # del키를 눌렀을때 해당 리스트뷰의 아이템들을 리스트뷰에서 삭제
    def keyPressEvent(self, event):
        self.event_manager.handle_key_press_event(event)



    # indexes( shotlist 리스트뷰에서 선택한 태스크들을 모든 스케쥴 리스트뷰에서 삭제함.(jsonData 에서도 모두 삭제))
    def delete_items(self, list_view, currListview, item):

        # 팀멤버 트리뷰에서 체크되어있는 멤버 가져오기
        checkedMembers = self.getCheckedMember()
        num_listview = self.dayUi.splitter.count()

        label_index = 0

        if list_view.model().rowCount():
            listView_task = []
            for i in range(num_listview):

                widget = self.dayUi.splitter.widget(i)
                listview_layout = widget.layout()

                if listview_layout:

                    frame = listview_layout.itemAt(2).widget()
                    frameLayout = frame.layout()
                    listview = frameLayout.itemAt(0).widget()

                    tasks = self.getListViewItem(listview) # 현재 ui상에 디스플레이되어있는 아이템들 가져오기

                    model_ = listview.model()
                    if item in tasks:
                        for row in range(model_.rowCount()):
                            listview_index = model_.index(row) # 이 인덱스는 스케쥴 리스트뷰에 있는 현재 보여지는 ui 에 있는 태스크들의 인덱스임.
                            listView_item = listview.model().data(listview_index, Qt.DisplayRole)
                                                                
                            if item == listView_item:
                                listview.model().removeRow(listview_index.row())




    # 현재 생성되어 있는 모든 리스트뷰의 리스트 반환
    def get_listview_list(self):

        num_listview = self.dayUi.splitter.count()

        listView_list=[]
        for i in range(num_listview):
            widget = self.dayUi.splitter.widget(i)
            listview_layout = widget.layout()

            if listview_layout:
                frame = listview_layout.itemAt(2).widget()
                frameLayout = frame.layout()
                listview_widget = frameLayout.itemAt(0).widget()
                listView_list.append(listview_widget)

        return listView_list





    # 스케쥴 리스트뷰에서 선택되어져 있는 태스크들을 모두삭제
    def delete_selected_items(self, list_view):

        listView_name = list_view.objectName() # 아이템 삭제가 발생한 리스트뷰의 이름을 가져오기
        num_listview = self.dayUi.splitter.count()

        label_index = 0

        listView_list=[]
        for i in range(num_listview):
            widget = self.dayUi.splitter.widget(i)
            listview_layout = widget.layout()

            if listview_layout:
                frame = listview_layout.itemAt(2).widget()
                frameLayout = frame.layout()
                listview_widget = frameLayout.itemAt(0).widget()
                listView_list.append(listview_widget)

        date_labels = self.getLabelData()

        del_date = ''
        for i in range(len(listView_list)):
            if listView_name == (listView_list[i].objectName()):
                del_date = date_labels[i]

        indexes = list_view.selectedIndexes()
        del_year = del_date.split('/')[0]
        del_month = del_date.split('/')[1]
        del_day = del_date.split('/')[2]

        for data in jsonData: # jsonData 내에 지우는 해당샷을 찾아서 삭제시킴

            if str(data['year']) == del_year and str(data['month']) == del_month and str(data['day']) == del_day:
                if indexes:
                    for index in sorted(indexes, reverse=True):
                        itemText = list_view.model().data(index, Qt.DisplayRole)

                        if data['tasks']:  
                            for proj in data['tasks'][0]:
                                taskList = data['tasks'][0][proj]
                                if itemText[1] in taskList:
                                    taskList.remove(itemText[1])


        if indexes:
            for index in sorted(indexes, reverse=True):
                itemText = list_view.model().data(index, Qt.DisplayRole)

                #리스트뷰에서 해당 아이템 삭제
                list_view.model().removeRow(index.row())



    # 특정 리스트뷰의 빈곳을 클릭하였을때 다른 모든 리스트뷰와 그안의 선택된 아이템들을 포함한 모든 아이템 선택해제
    def clear_selection(self):

        listviews = self.get_listview_list()
        
        for listview in listviews:
            listview.selectionModel().clearSelection()





    # 디벨롭 예정
    def find_layout(self, widget):

        num_listview = self.day_listViewLayout.count()

        listView_list = []
        for i in range(num_listview):
            listview_layout = self.day_listViewLayout.layout().itemAt(i)
            if listview_layout:
                listview_point = listview_layout.layout().itemAt(1)
                listview = listview_point.widget()
                listView_list.append(listview)
                



    # 현재 리스트뷰들의 태스크들로 제이슨파일 업데이트하기
    def updateJson(self,direction):

        global jsonData
        global connect_shot_proj

        # 팀멤버 트리뷰에서 체크되어있는 멤버 가져오기
        checkedMembers = self.getCheckedMember()

        teamInfo = [{}]
        for mem in list(self.memberCache_.keys()):

            memberInfo = self.team_manager.get_team_info(mem)
            if memberInfo != []:
                for teamMem in memberInfo[0]:
                    if teamMem not in teamInfo[0]:
                        teamInfo[0][teamMem] = memberInfo[0][teamMem]

        # 현재 스케쥴 리스트뷰의 갯수
        num_listview = self.dayUi.splitter.count()

        # 현재 스케쥴 리스트뷰의 태스크들을 listView_task 리스트에 저장
        listView_task = []
        for i in range(num_listview):
            widget = self.dayUi.splitter.widget(i)
            listview_layout = widget.layout()

            if listview_layout:
                frame = listview_layout.itemAt(2).widget()
                frameLayout = frame.layout()
                listview = frameLayout.itemAt(0).widget()

                tasks = self.getListViewItem(listview) # 현재 ui상에 디스플레이되어있는 아이템들 가져오기
                tasks_inJson = self.getListViewItem_inJson(listview, checkedMembers)

                if tasks_inJson:
                    for task_ in tasks_inJson:
                        if task_ not in tasks:
                            tasks.append(task_)

                listView_task.append(tasks)

        # 현재 스케쥴 리스트뷰의 날짜정보 가져오기
        listView_date = self.getCurrentListViewDate(0)

        # 현재 샷 리스트상에 로드된 샷들의 태스크정보 읽어오기
        taskInfoJson = self.file_manager.import_Json(currentPath+"/.task_info", userID+"_task_info.json")

        # 현재 체크된것과 상관없이 모든 팀원 리스트 읽어오기
        memberList = list(teamInfo[0].keys())

        if userID not in memberList:
            memberList.append(userID)

        member_projectTasks={}
        for member in taskInfoJson:
            connect_shot_proj_keyList = list(connect_shot_proj.keys())

            for i in range(len(listView_task)):
                projectTasks={}
                projs=[]
                taskList=[]
                task_data={}
                department = ""

                for task in listView_task[i]:

                    process=""

                    if task[4] == "M":
                        process = "matchmove"
                    elif task[4] == "A":
                        process = "animation"                        
                    elif task[4] == "R":
                        process = "creature"

                    projCode = ""
                    if task in connect_shot_proj_keyList:   # 프로젝트와 샷을 연결시킨 전역변수 connect_shot_proj 내에 샷이 있는지 확인후, 있으면 연결된 프로젝트 반환
                        proj = connect_shot_proj[task]
                        projCode = self.projects[proj][0]

                    if member in memberList:

                        for taskInfo in taskInfoJson[member]:
                            if projCode == taskInfo["project_code"]: # 서로다른 프로젝트의 같은 이름의 태스크를 구별하기 위해 프로젝트가 같은지 확인
                                identifier = 'noContext'
                                subContext_ = ''
                                orig_task = task

                                if task[1].find('/') != -1:
                                    task_context = task[1].split('/')
                                    taskName = task_context[0]
                                    subContext_ = task_context[1]
                                    identifier = 'haveContext'

                                # 식별자를 확인해서 컨텍스트를 가지고 있으면, 태스크 이름과 taskInfo의 컨텍스트와 비교하여 모두 같으면 jsonData에 저장될 projectTask리스트에 저장
                                if  identifier == 'haveContext':
                                    subContext = ''
                                    origContext = taskInfo["context"]
                                    if origContext.find('/') != -1:                                
                                        subContext = origContext.split('/')[1]

                                    if taskName == taskInfo["extra_code"] and subContext == subContext_ and taskInfo["process"] == process:
                                        
                                        if proj in projs:
                                            task_fullName = taskName+"/"+subContext_
                                            if task_fullName not in projectTasks[proj]:
                                                projectTasks[proj].append(task_fullName)

                                        elif proj not in projs:
                                            projs.append(proj)
                                            task_fullName = taskName+"/"+subContext_ 
                                            projectTasks[proj] = [task_fullName]

                                # 식별자를 확인해서 컨텍스트가 없으면, 태스크명만 확인해서 같으면 jsonData에 저장될 projectTask리스트에 저장
                                elif identifier == 'noContext':
                                    if task[1] == taskInfo["extra_code"] and ("/" not in taskInfo["context"]) and taskInfo["process"] == process:
                                        if proj in projs:
                                            if task[1] not in projectTasks[proj]:
                                                projectTasks[proj].append(task[1])

                                        elif proj not in projs:
                                            projs.append(proj)
                                            projectTasks[proj] = [task[1]]

                                task = orig_task # 컨텍스트 지정 여부를 확인하기 위해 태스크명과 컨텍스트 분리된 태스크명을 다시 원래태스크명으로 되돌리기 위한 코드


                if member == userID:
                    department = self.department                
                else:
                    department = teamInfo[0][member]["department"]

                task_data = {"artist":member, "year":listView_date[i]["year"], "month":listView_date[i]["month"], "day":listView_date[i]["day"], "tasks":[projectTasks], "department":department}#, "tasks":listView_task[i]}

                # jsonData내의 모든 스캐쥴데이터중 task_data와 동일한 날짜인 모든 데이터를 filterDate_inJson안에 리스트로 저장(아티스트변경에 따른 오류를 체크하기 위해)
                filterDate_inJson = [item for item in jsonData
                                        if item["year"] == listView_date[i]["year"]
                                        and item["month"] == listView_date[i]["month"]
                                        and item["day"] == listView_date[i]["day"]]

                # 생성된 task_data의 어싸인 업데이트가 있었는지 확인하고 있었다면, jsonData내의 스케쥴 데이타를 삭제하고, 현재 진행중이던 태스크의 스케쥴생성은 중단함.
                if self.validation_manager.check__Task_data(task_data, filterDate_inJson, jsonData):
                    continue

                if jsonData == []:
                    jsonData.append(task_data)
                else:
                    for data in jsonData:
                        #if (listView_date[i]["year"] == data["year"] and listView_date[i]["month"] == data["month"] and listView_date[i]["day"] == data["day"]):
                        if (task_data["year"] == data["year"] and task_data["month"] == data["month"] and task_data["day"] == data["day"]) and (data["artist"] == member):

                            # task_data와 같은 날짜의 jsonData에 아무런 값이 없을때 값이 없는 데이타를 지우고 task_data를 jsonData에 추가함(append메서드를 사용해야 ui에 표시됨- 이유확인필요)
                            if data["tasks"][0] == {} and task_data["tasks"][0] != {}:
                                jsonData.remove(data)
                                jsonData.append(task_data)

                            # task_data와 같은 날짜의 jsonData에 스케쥴 데이타가 존재하지만, task_data에 있는 프로젝트가 존재하지 않을때, 
                            # task_data에 현재 존재하는 jsonData의 데이타를 추가한뒤 이 jsonData를 지우고  task_data를 jsonData에 추가함
                            elif data["tasks"][0] != {} and task_data["tasks"][0] != {}:
                                project_task_dict = {}
                                for project in data["tasks"][0].keys():
                                    if project not in task_data["tasks"][0]:
                                        task_data["tasks"][0][project] = data["tasks"][0][project]

                                        if data in jsonData:                                                
                                            jsonData.remove(data)

                                        if task_data not in jsonData:
                                            jsonData.append(task_data)

                                    elif project in task_data["tasks"][0]:
                                        for task in data["tasks"][0][project]:
                                            if task not in task_data["tasks"][0][project]:
                                                task_data["tasks"][0][project].append(task)

                                        if data in jsonData:
                                            jsonData.remove(data)

                                        if task_data not in jsonData:
                                            jsonData.append(task_data)

                        elif  (task_data not in jsonData): 
                            if ((self.validation_manager.check_date_exists(jsonData, task_data) == False) or (data["artist"] != member)):
                                jsonData.append(task_data)

        self.refereshListViews( direction, jsonData)
        print ("============================================")

        return jsonData





    # 벨로즈의 어싸인된 아티스트가 변경되었을경우 json 파일을 그에 맞게 업데이트 시킴
    def updateAssignment(self, user_delAssign, user_newAssign, taskData):

        delAssign_jsonFile = os.path.join(jsonFilePath, user_delAssign+"_Data.json")

        # 아사인이 바뀐 이전 아티스트의 스케쥴 데이타 가져오기
        delAssign_json = []
        if (os.path.exists(delAssign_jsonFile)):
            with open(delAssign_jsonFile) as f:
                delAssign_json = json.load(f)

            # 다른 아티스트로 어사인이 바뀐 이전 아티스트의 태스크 목록에서 태스크 삭제
            for data in delAssign_json:
                if data["year"] == taskData["year"] and data["month"] == taskData["month"] and data["day"] == taskData["day"]:
                    projs = list(data['tasks'][0].keys())
                    projs_taskData = list(taskData['tasks'][0].keys())

                    if projs:
                        for proj in projs:
                            for task in data['tasks'][0][proj]:
                                if proj in projs_taskData:
                                    if task in taskData['tasks'][0][proj]:
                                        data['tasks'][0][proj].remove(task)  

                            if data['tasks'][0][proj] == []:
                                del data['tasks'][0][proj]

            # 하나의 태스크도 존재하지 않으면 스케쥴 데이터 자체를 삭제
            del_data = []
            for i in range(len(delAssign_json)):
                if delAssign_json[i]['tasks'] == [{}]:
                    del_data.append(delAssign_json[i])
                    #del delAssign_json[i]

            for data in del_data:
                if data in delAssign_json:
                    delAssign_json.remove(data)

            with open(delAssign_jsonFile, 'w') as file:
                json.dump(delAssign_json, file, indent=4)
                    



    # 한칸 앞, 또는 한칸 뒤로 이동을 위한 ui 의 요청에 따라 스케쥴 리스트뷰의 내용을 갱신
    def move_sideways(self, direction, scheduleData):

        old_listViewDate = self.getCurrentListViewDate(0)
        listViewDate = self.getCurrentListViewDate(direction)
        listViewNum = self.dayUi.splitter.count()

        num_currentListView = self.dayUi.splitter.count()

        movedDate = QDate(self.rangeEndDate.addDays(direction))
        self.setListviewDate(movedDate, num_currentListView)
        self.comboDay.setCurrentText(str(movedDate.day()))
        self.comboMonth.setCurrentText(str(movedDate.month()))        
        self.comboYear.setCurrentText(str(movedDate.year()))

        # 프로젝트 리스트에서 선택된 프로젝트 가져오기
        # 프로젝트 리스트에서 선택된 프로젝트들 가져오기
        selectedProj = self.get_selected_projects()

        # 현재 ui에 표시되어있는 스케쥴 리스트뷰 가져오기(옮기기전)
        listViews=[]
        frameDic = {} # 현재 선택된 프레임을 표시하여 딕셔너리 저장 : 하이라이트를 이동에 따라 힘께 이동시키기 위해
        layouts = []
        for i in range(listViewNum):
            widget = self.dayUi.splitter.widget(i)
            layout = widget.layout()
            layouts.append(layout)

            if layout:
                frame = layout.itemAt(2).widget()
                frameLayout = frame.layout()
                listView = frameLayout.itemAt(0).widget()
                listViews.append(listView)

                self.manager.Move_sideway_setContainer(frame, layout)

                # 현재 프래임의 선택 여부 확인후 딕셔너리 저장
                sel = 0
                if hasattr(frame, 'styleSheet') and frame.styleSheet() == "border: 1px solid blue;":
                    sel = 1
                frameDic[frame] = sel



        # 현재 멤버트리뷰의 루트아이템을 읽어들인뒤 그 하위의 체크되어있는 모든 멤버들을 리스트로 만들고 영어이름으로 변환
        invisibleRoot = self.teamTreeModel.invisibleRootItem()
        rootItem = invisibleRoot.child(0)

        checked_items_list_Kr = []
        self.get_CheckedItems(rootItem, checked_items_list_Kr)

        checked_items_list_En = []

        teamInfo = self.file_manager.import_Json(currentPath+"/.team_Info", userID+"_teamInfo.json")

        for mem in self.memberCache_:
            if self.memberCache_[mem] in checked_items_list_Kr:
                checked_items_list_En.append(mem)

        # jsonData 내의 데이타중 현재 체크되어있는 멤버의 데이타만 따로 저장
        refereshJsonData = []
        ######## < 삭제금지 >   
        # 컴프리헨션 가이드
        #for member in checked_items_list_En:
        #    for i in range(len(scheduleData)):
        #        if scheduleData[i]['artist'] == member:
        #            refereshJsonData.append(jsonData[i])
        refereshJsonData = [jsonData[i] for member in checked_items_list_En for i in range(len(scheduleData)) if scheduleData[i]['artist'] == member]


        # 데이터와 날짜를 이동시키기전 각 날짜에 대응하는 리스트뷰의 소팅정보 저장 -> 이동시킨 이후 이 소팅정보를 새로운 리스트뷰에 동일하게 적용하기 위해
        for i in range(len(old_listViewDate)):
            old_model = listViews[i].model()
            date = str(old_listViewDate[i]["year"]) + "-" + str(old_listViewDate[i]["month"]) + "-" + str(old_listViewDate[i]["day"])
            sortOrder_date[date] = old_model._sort_order
            sortColumn_date[date] = old_model._sort_column

        sortOrder = {}
        sortColumn = {}

        inData_listViews = []
        for i in range(len(listViewDate)):
            tasks=[]

            #### < 삭제금지 >
            # 컴프리헨션 가이드
            #for data in refereshJsonData:
            #    task_dep = {}
            #    if data['year'] == listViewDate[i]['year'] and data['month'] == listViewDate[i]['month'] and data['day'] == listViewDate[i]['day']:
            #        for projTask in data['tasks'][0]:
            #            if projTask in selectedProj:
            #                for task in data['tasks'][0][projTask]:
            #                    task_dep[task] = data["department"]
            #                    if task_dep not in tasks: # 리스트뷰안에 넣으려는 tasks리스트 안에 task가 이미 있는지 확인
            #                        tasks.append(task_dep)
            #tasks = [ {task:data["department"]  for task in data['tasks'][0][projTask]}
            #        for data in refereshJsonData if data['year'] == listViewDate[i]['year'] and data['month'] == listViewDate[i]['month'] and data['day'] == listViewDate[i]['day']
            #        for projTask in data['tasks'][0] if projTask in selectedProj]

            tasks = [ {task:[projTask.lower(), data["department"]]  for task in data['tasks'][0][projTask]}
                    for data in refereshJsonData if data['year'] == listViewDate[i]['year'] and data['month'] == listViewDate[i]['month'] and data['day'] == listViewDate[i]['day']
                    for projTask in data['tasks'][0] if projTask in selectedProj]




            tuple_tasks = self.convert_to_tuple(tasks, checked_items_list_En)
            inData_listViews.append(i)

            # 갱신된 새로운 아이템들이 등록된 모델을 리스트뷰에 등록하기전 이전 모델 삭제를 위해 이전모델을 old_model에 저장
            old_model = listViews[i].model()

            sortOrder[i] = old_model._sort_order
            sortColumn[i] = old_model._sort_column

            shotlist_model = ddlv.DragDropModel(tuple_tasks, sortColumn[i])
            listViews[i].setModel(shotlist_model)

            # 이전 모델 삭제
            del old_model

            self.set_itemBackgroundColor(shotlist_model, listViewDate[i],listViews[i])
            self.set_listView_view(listViews[i], shotlist_model)

        self.ui_layout_manager.reset_splitter_size() # 스케쥴 리스트뷰의 스플리터 사이즈 리셋

        return movedDate



    # 스케쥴 리스트뷰의 스플리터 사이즈 리셋



    # ui 의 요청에 따라 리스트뷰의 내용을 갱신하기
    def refereshListViews(self, direction, scheduleData):

        listViewDate = self.getCurrentListViewDate(direction)
        #listViewNum = self.day_listViewLayout.count()
        listViewNum = self.dayUi.splitter.count()

        # 프로젝트 리스트에서 선택된 프로젝트 가져오기
        selectedProj = self.get_selected_projects()

        listViews=[]
        for i in range(listViewNum):
            widget = self.dayUi.splitter.widget(i)
            layout = widget.layout()
            #layout = self.day_listViewLayout.itemAt(i)

            if layout:
                #listView = layout.itemAt(2).widget()
                frame = layout.itemAt(2).widget()
                frameLayout = frame.layout()
                listView = frameLayout.itemAt(0).widget()
                listViews.append(listView)

        # 현재 멤버트리뷰의 루트아이템을 읽어들인뒤 그 하위의 체크되어있는 모든 멤버들을 리스트로 만들고 영어이름으로 변환
        invisibleRoot = self.teamTreeModel.invisibleRootItem()
        rootItem = invisibleRoot.child(0)

        checked_items_list_Kr = []
        self.get_CheckedItems(rootItem, checked_items_list_Kr)

        checked_items_list_En = []
        teamInfo = self.file_manager.import_Json(currentPath+"/.team_Info", userID+"_teamInfo.json")

        for mem in self.memberCache_:
            if self.memberCache_[mem] in checked_items_list_Kr:
                checked_items_list_En.append(mem)

        invalidTask_list = []    # 차후에 삭제 예정 : 컨텍스트 오류가 있는 프로젝트 마감후
        fixedTask_list = []     # 차후에 삭제 예정 : 컨텍스트 오류가 있는 프로젝트 마감후

        refereshJsonData = []
        for member in checked_items_list_En:
            for i in range(len(scheduleData)):
                if scheduleData[i]['artist'] == member:
                    refereshJsonData.append(jsonData[i])

        ############################################################################################
        ############################################################################################
        # 컨텍스트 수정을 위한 임시 코드, 차후 삭제예정 : 컨텍스트 오류가 있는 프로젝트 마감후

                    if self.contextFix:

                        for task in self.contextFix:

                            invalidCTX = {}

                            proj=self.contextFix[task][0]

                            contextList = task.split('/')
                            invalid_task = contextList[0]+"/"+contextList[1]

                            if  proj in jsonData[i]['tasks'][0]: 
                                if self.contextFix[task][1] == member:
                                    if invalid_task in jsonData[i]['tasks'][0][proj]:

                                        editTask = jsonData[i]
                                        invalidTask_list.append(jsonData[i])

                                        editTask["tasks"][0][proj].remove(invalid_task)
                                        editTask["tasks"][0][proj].append(task)
                                        
                                        fixedTask_list.append(editTask)

                    ##### 에러가 발생했을때 해당 태스크에 대한 내용을 로그데이터에 기록하고, 프로세스는 계속 이어지도록
        ################################################################################################
        ################################################################################################


        # 기존 모델에서 정렬 정보 먼저 수집
        sortOrder = {}
        sortColumn = {}
        for i in range(len(listViewDate)):
            old_model = listViews[i].model()
            if old_model:
                sortOrder[i] = old_model._sort_order
                sortColumn[i] = old_model._sort_column
            else:
                sortOrder[i] = Qt.AscendingOrder
                sortColumn[i] = 0

        inData_listViews = []

        for i in range(len(listViewDate)):

            tasks=[]
            for data in refereshJsonData:
                if data['year'] == listViewDate[i]['year'] and data['month'] == listViewDate[i]['month'] and data['day'] == listViewDate[i]['day']:
                    for projTask in data['tasks'][0]:
                        if projTask in selectedProj:
                            #proj_part.append(projTask)
                            for task in data['tasks'][0][projTask]:
                                task_dep = {}
                                proj_part = []
                                proj_part.append(projTask)
                                proj_part.append(data["department"])
                                task_dep[task] = proj_part  

                                if task_dep not in tasks: # 리스트뷰안에 넣으려는 tasks리스트 안에 task가 이미 있는지 확인``
                                    tasks.append(task_dep)

            tuple_tasks = self.convert_to_tuple(tasks, checked_items_list_En)
            inData_listViews.append(i)

            # 갱신된 새로운 아이템들이 등록된 모델을 리스트뷰에 등록하기전 이전 모델 삭제를 위해 이전모델을 old_model에 저장
            old_model = listViews[i].model()

            sortOrder[i] = old_model._sort_order
            sortColumn[i] = old_model._sort_column

            shotlist_model = ddlv.DragDropModel(tuple_tasks, sortColumn[i])
            listViews[i].setModel(shotlist_model)

            # 이전 모델 삭제
            del old_model

            self.set_itemBackgroundColor(shotlist_model, listViewDate[i], listViews[i])


        # 스케쥴 리스트뷰의 새로고침 과정에서 각 리스트뷰의 소팅정보를 sortOrder_date, sortColumn_date 에 저장하였다가, 새로고침이 끝난후 해당 날짜를 찾아
        # 저장된 소팅정보대로 이전의 정렬과 같게 만들어줌
        for i in range(len(listViewDate)):
            shotList_model = listViews[i].model()
            date = str(listViewDate[i]["year"]) + "-" + str(listViewDate[i]["month"]) + "-" + str(listViewDate[i]["day"])
            dateDic = {"year":listViewDate[i]["year"], "month":listViewDate[i]["month"], "day":listViewDate[i]["day"]}

            if date in sortOrder_date:
                self.sort_manager.sortScheduleListview(listViews[i], shotList_model, sortOrder_date[date], sortColumn_date[date], dateDic)

            # refereshListView 메서드내에서 실행되어 스케쥴데이터가 업데이트될때마다 뷰필터 체크박스를 확인하여 내용을 갱신시킴
            self.set_listView_view(listViews[i], shotList_model)



    def reSelect_proj(self):

        selectedIndexes = self.projListview.selectionModel().selectedIndexes()
        self.projListview.clearSelection()

        for index in selectedIndexes:
            #print (index.data())
            self.projSelection_model.select(index, QItemSelectionModel.Select)            




    # 현재 선택되어 있는 멤버들의 이번주 완료해야할 프로젝트 
    def get_projs_thisWeek(self, taskInfoJson, checked_mem):
        
        today = datetime.now()
        today_weekday = today.weekday()
        mon = today - timedelta(days=today_weekday)
        fri = mon + timedelta(days=4)
        mem_proj = {}

        # mon을 월요일 0시로 설정
        mon=mon.replace(hour=0, minute=0, second=0, microsecond=0)

        # fri를 금요일 24시(토요일0시)로 설정
        fri=fri+timedelta(days=1)
        fri=fri.replace(hour=0, minute=0, second=0, microsecond=0)


        if taskInfoJson:
            projs_currentMem = []
            for mem in taskInfoJson:

                if mem in checked_mem:

                    df_taskInfo = pd.DataFrame(taskInfoJson[mem])
                    df_taskInfo["end_date"] = pd.to_datetime(df_taskInfo["end_date"])
                    df_thisWeek = df_taskInfo[(df_taskInfo["end_date"] >= mon) & (df_taskInfo["end_date"] <= fri)]  
                    thisWeek_proj_codes = df_thisWeek["project_code"].unique().tolist()
                    projs = [proj for proj in self.projects for proj_code in thisWeek_proj_codes if self.projects[proj][0] == proj_code] 
                    mem_proj[mem] = projs

                    for proj in projs:
                        if proj not in projs_currentMem:
                            projs_currentMem.append(proj)

            return projs_currentMem




    def getProjs_Mem_selAll(self, taskInfoJson, memEN=None, deselectAll=None, unchecked=None, teamMember=None) :


        # 팀 트리뷰에서 체크된 멤버 리스트 읽어오기
        invisibleRoot = self.teamTreeModel.invisibleRootItem()
        rootItem = invisibleRoot.child(0)

        checked_items_list = []
        self.get_CheckedItems(rootItem, checked_items_list)

        checkedList_EN = [member for member in list(self.memberCache_.keys()) for memKR in checked_items_list if self.memberCache_[member] == memKR]

        today = QDate.currentDate()
        monday = today.addDays(-today.dayOfWeek()+1)
        friday = monday.addDays(4) 
        lastWeek_monday = today.addDays(-7+(-today.dayOfWeek()+1))
        nextWeek_friday = today.addDays(11+(-today.dayOfWeek()+1))

        betweenDate = {}
        day_count = 0
        date = lastWeek_monday

        #selected_mem = []
        projs_currentMem = []
        projs_currentMem = self.get_projs_thisWeek(taskInfoJson, checkedList_EN)
        projLong_list = []

        for memEN in checkedList_EN:
            if memEN in taskInfoJson:
                df_taskInfo = pd.DataFrame(taskInfoJson[memEN])

                # end_date를 datetime으로 변환
                df_taskInfo["end_date"] = pd.to_datetime(df_taskInfo["end_date"])
                today = datetime.now()
                today_weekday = today.weekday()
                mon = today - timedelta(days=today_weekday)
                fri = mon + timedelta(days=4)

                df_thisWeek = df_taskInfo[(df_taskInfo["end_date"] >= mon) & (df_taskInfo["end_date"] <= fri)]

                thisWeek_proj_codes = df_thisWeek["project_code"].unique().tolist()

                projs=[]
                if memEN in checkedList_EN:
                    projs = [proj for proj in self.projects for proj_code in thisWeek_proj_codes if self.projects[proj][0] == proj_code] 

                for proj in projs:
                    if proj not in projs_currentMem:
                        projs_currentMem.append(proj)

                # 이번주를 포함한 지난주 다음주 3주간의 스케쥴이 어싸인되어있는 프로젝트를 찾아 task_3w 에 저장
                tasks_3w = []
                if jsonData != []:
                    for data in jsonData:
                        data_date = QDate(data["year"], data["month"], data["day"])
                        if lastWeek_monday <= data_date <= nextWeek_friday and data['artist'] == memEN:
                            tasks_3w.append(data)

                for task_3w in tasks_3w:
                    if task_3w["tasks"]:
                        for proj in list(task_3w["tasks"][0].keys()):
                            if proj not in projs_currentMem:
                                projs_currentMem.append(proj)


        # 각 프로젝트의 전체 이름 가져오기
        if projs_currentMem:
            for proj in projs_currentMem:                    
                if proj in self.projects:
                    projLong_list.append(self.projects[proj][1])



        # 뽑아낸 프로젝트 이름을 이용하여 UI상의 프로젝트 리스트뷰에서 해당 프로젝트를 선택
        proj_model = self.projListview.model()

        # 프로젝트 리스트뷰 새로고침을 위해 모두 선택해제
        self.projSelection_model.clearSelection()

        for row in range(proj_model.rowCount()):
            index = proj_model.index(row, 0)
            text_proj = proj_model.data(index)

            if text_proj != '관리개발 (testShot)':# and text_proj in projLong_list:
                if text_proj not in projLong_list:
                    self.projSelection_model.select(index, QItemSelectionModel.Deselect)
                else:
                    print (text_proj)
                    self.projSelection_model.select(index, QItemSelectionModel.Select)

        # 샷리스트뷰 셋업이 끝난뒤 projLong_list 초기화
        projLong_list = []

        # 최종적으로 선택된 프로젝트들의 리스트를 self.currentProjs 에 저장
        self.currentProjs = projs_currentMem



    def getProjs_Mem(self, taskInfoJson, memEN, deselectAll=None, unchecked=None, teamMember=None) :

        # 팀 트리뷰에서 체크된 멤버 리스트 읽어오기
        invisibleRoot = self.teamTreeModel.invisibleRootItem()
        rootItem = invisibleRoot.child(0)

        checked_items_list = []
        self.get_CheckedItems(rootItem, checked_items_list)

        checkedList_EN = [member for member in list(self.memberCache_.keys()) for memKR in checked_items_list if self.memberCache_[member] == memKR]

        today = QDate.currentDate()
        monday = today.addDays(-today.dayOfWeek()+1)
        friday = monday.addDays(4) 
        lastWeek_monday = today.addDays(-7+(-today.dayOfWeek()+1))
        nextWeek_friday = today.addDays(11+(-today.dayOfWeek()+1))

        betweenDate = {}
        day_count = 0
        date = lastWeek_monday

        projs_currentMem = []
        projLong_list = []
        current_all_Projects = []

        # 현재 멤버가 체크박스 on인경우 어싸인되어 있는 프로젝트들중 이번주 완료해야할 태스크가 있는 프로젝트들을 가져옴
        projs_currentMem = self.get_projs_thisWeek(taskInfoJson, checkedList_EN)


        if memEN in taskInfoJson:

            df_taskInfo = pd.DataFrame(taskInfoJson[memEN])

            # end_date를 datetime으로 변환
            df_taskInfo["end_date"] = pd.to_datetime(df_taskInfo["end_date"])
            today = datetime.now()

            today_weekday = today.weekday()
            mon = today - timedelta(days=today_weekday)
            fri = mon + timedelta(days=4)

            df_thisWeek = df_taskInfo[(df_taskInfo["end_date"] >= mon) & (df_taskInfo["end_date"] <= fri)]
            thisWeek_proj_codes = df_thisWeek["project_code"].unique().tolist()

            projs=[]
            if memEN in checkedList_EN:
                projs = [proj for proj in self.projects for proj_code in thisWeek_proj_codes if self.projects[proj][0] == proj_code] 

            for proj in projs:
                if proj not in projs_currentMem:
                    projs_currentMem.append(proj)


            #projLong_list = []
            # 이번주를 포함한 지난주 다음주 3주간의 스케쥴이 어싸인되어있는 프로젝트를 찾아 task_3w 에 저장
            tasks_3w = []
            if jsonData != []:
                for data in jsonData:
                    data_date = QDate(data["year"], data["month"], data["day"])
                    if lastWeek_monday <= data_date <= nextWeek_friday and data['artist'] == memEN:
                        tasks_3w.append(data)

            for task_3w in tasks_3w:
                if task_3w["tasks"]:
                    for proj in list(task_3w["tasks"][0].keys()):
                        if proj not in projs_currentMem:
                            projs_currentMem.append(proj)


            current_all_Projects = list(dict.fromkeys(self.currentProjs + projs_currentMem ))

            for proj in current_all_Projects:                    
                if proj in self.projects:
                    projLong_list.append(self.projects[proj][1])



            """
            for proj in projs_currentMem:                    
                if proj in self.projects:
                    projLong_list.append(self.projects[proj][1])
            """



            # 뽑아낸 프로젝트 이름을 이용하여 UI상의 프로젝트 리스트뷰에서 해당 프로젝트를 선택
            proj_model = self.projListview.model()

            for row in range(proj_model.rowCount()):
                index = proj_model.index(row, 0)
                text_proj = proj_model.data(index)

                if text_proj != '관리개발 (testShot)':# and text_proj in projLong_list:
                    if text_proj not in projLong_list:
                        self.projSelection_model.select(index, QItemSelectionModel.Deselect)
                    else:
                        self.projSelection_model.select(index, QItemSelectionModel.Select)

            # 샷리스트뷰 셋업이 끝난뒤 projLong_list 초기화
            projLong_list = []
            self.currentProjs = current_all_Projects

        elif memEN not in taskInfoJson:

            #current_all_Projects = self.currentProjs
            #print (self.currentProjs)

            projLong_list = []
            currentProj_long_list = []

            # 현재  특정멤저의 체크박스를 해제한상태(taskInfoJson  에 memEN이 없는상태) 이므로 그외 남은  멤버의 프로젝트  선택된 상태
            for member in checkedList_EN:
                projLong_list = self.get_projects_3Weeks(member, lastWeek_monday, nextWeek_friday, projs_currentMem)

            if projs_currentMem:
                self.currentProjs = [proj for proj in self.currentProjs if proj in projs_currentMem]
            else:
                self.currentProjs = []


            proj_model = self.projListview.model()
            exceptCurrentProj = []

            if projLong_list == None:
                self.projListview.clearSelection()


            else:
                for row in range(proj_model.rowCount()):
                    index = proj_model.index(row, 0)
                    text_proj = proj_model.data(index)
                    if text_proj != '관리개발 (testShot)' and text_proj in projLong_list:
                        if not self.projSelection_model.isSelected(index):
                            self.projSelection_model.select(index, QItemSelectionModel.Select)

                    elif text_proj != '관리개발 (testShot)' and text_proj not in projLong_list:# and text_proj in currentProj_long_list:
                        self.projSelection_model.select(index, QItemSelectionModel.Deselect)

        # 샷리스트뷰 셋업이 끝난뒤 projLong_list 초기화
        projLong_list = []





    def get_projects_3Weeks(self, memEN, lastWeek_monday, nextWeek_friday, projs_currentMem):

        projLong_list = []
        # 이번주를 포함한 지난주 다음주 3주간의 스케쥴이 어싸인되어있는 프로젝트를 찾아 task_3w 에 저장
        tasks_3w = []
        if jsonData != []:
            for data in jsonData:
                data_date = QDate(data["year"], data["month"], data["day"])
                if lastWeek_monday <= data_date <= nextWeek_friday and data['artist'] == memEN:
                    tasks_3w.append(data)

        for task_3w in tasks_3w:
            if task_3w["tasks"]:
                for proj in list(task_3w["tasks"][0].keys()):
                    if proj not in projs_currentMem:
                        projs_currentMem.append(proj)

        for proj in projs_currentMem:                    
            if proj in self.projects:
                projLong_list.append(self.projects[proj][1])


        if projLong_list:
            return projLong_list





















    # 트리뷰의 멤버를 선택하면 프로젝트 리스트에 어싸인된 프로젝트들 자동으로 선택되게 하기
    def sel_assigned_proj(self, set="multi", index=None, items=None, teamMember=None, deselectAll=None):

        taskInfoJson = self.file_manager.import_Json(currentPath+"/.task_info", userID+"_task_info.json")

        checkedItems = []
        if index != None and set=="single" and items > 1:
            memKR = self.teamTreeModel.data(index, Qt.DisplayRole)
            memEN = [mem for mem in self.memberCache_ if self.memberCache_[mem] == memKR]
            if memEN:
                self.getProjs_Mem(taskInfoJson, memEN[0])

        elif index != None and set == "single" and items == 0 or items == 1:
            memKR = self.teamTreeModel.data(index, Qt.DisplayRole)
            memEN = [mem for mem in self.memberCache_ if self.memberCache_[mem] == memKR]
            if memEN:
                self.getProjs_Mem(taskInfoJson, memEN[0])            

        elif index == None and set == "multi":
            self.getProjs_Mem_selAll(taskInfoJson)                    


    # 리스트뷰 뷰필터 stateChanged 시그널 연결시켜 체크박스를 키거나 끌때 리스트뷰에 바로 적용되도록 연결 
    def listView_filter(self, state):
        self.refereshListViews(0, jsonData)
        self.reloadShotList()

    # 샷리스트뷰와 스케쥴 리스트뷰 모두에 이번주 완료예정인 태스크만 보이도록 함(토글)
    def deadLine_filter(self):

        if self.deadLine_flag == 1:
            self.deadLine_flag = 0
            self.ui.deadlineButton.setStyleSheet("QPushButton { background-color: rgb(53,53,53); }")

        elif self.deadLine_flag == 0:
            self.deadLine_flag = 1
            self.ui.deadlineButton.setStyleSheet("QPushButton { background-color: rgb(90,0,0); }")

        self.refereshListViews(0, jsonData)
        self.reloadShotList()



    # refereshListView 메서드내에서 실행되어 스케쥴데이터가 업데이트될때마다 뷰필터 체크박스를 확인하여 내용을 갱신시킴
    def set_listView_view(self, listView, model):

        if model != None:
            for row in range(model.rowCount()):
                index = model.index(row)
                taskData = model.data(index, Qt.DisplayRole)

                # 이번주 마감인 태스크들의 온/오프
                taskColor = model.data(index, Qt.BackgroundRole) # 배경색 가져오기

                # taskColor가 None인 경우 안전 처리
                if taskColor is not None:
                    if (taskColor.red() != 94 and taskColor.green() != 29 and  taskColor.blue() != 35) and self.deadLine_flag==1:
                        listView.setRowHidden(row, True)
                    elif (taskColor.red() != 94 and taskColor.green() != 29 and  taskColor.blue() != 35) and self.deadLine_flag==0:
                        listView.setRowHidden(row, False)
                else:
                    # taskColor가 None인 경우 기본 동작 (데드라인 플래그에 따라 처리)
                    if self.deadLine_flag==0:
                        listView.setRowHidden(row, False)                


                # 스테이터스 체크박스의 상태에 따른  온/오프
                app_check = self.dayUi.checkBox_approved_v.isChecked()#self.checkBox_app.isChecked()
                inprogress_check = self.dayUi.checkBox_inprogress_v.isChecked()
                ready_check = self.dayUi.checkBox_ready_v.isChecked()
                review_check = self.dayUi.checkBox_review_v.isChecked()
                hold_check = self.dayUi.checkBox_hold_v.isChecked()
                ok_check = self.dayUi.checkBox_ok_v.isChecked()
                wait_check = self.dayUi.checkBox_wait_v.isChecked()
                omit_check = self.dayUi.checkBox_omit_v.isChecked()
                retake_check = self.dayUi.checkBox_retake_v.isChecked()


                if app_check == False and taskData[3] == "Approved":
                    listView.setRowHidden(row, True)

                if inprogress_check == False and taskData[3] == "In-Progress":
                    listView.setRowHidden(row, True)

                if ready_check == False and taskData[3] == "Ready":
                    listView.setRowHidden(row, True)

                if review_check == False and taskData[3] == "Review":
                    listView.setRowHidden(row, True)

                if hold_check == False and taskData[3] == "Hold":
                    listView.setRowHidden(row, True)

                if ok_check == False and taskData[3] == "OK":
                    listView.setRowHidden(row, True)

                if wait_check == False and taskData[3] == "Waiting":
                    listView.setRowHidden(row, True)

                if omit_check == False and taskData[3] == "Omit":
                    listView.setRowHidden(row, True)

                if retake_check == False and taskData[3] == "Retake":
                    listView.setRowHidden(row, True)




    # 아이템의 상태에 따라 백그라운드 색깔 설정
    def set_itemBackgroundColor(self, listview_Model, listView_date, listview=None):

        if listview is not None:
            frame = listview.parent()
            parent_widget = frame.parent()
            today = QDate.currentDate()

            if listView_date["year"] == today.year() and listView_date["month"] == today.month() and listView_date["day"] == today.day() :
                parent_widget.setStyleSheet("background-color: #38613b;")
            else:
                parent_widget.setStyleSheet("")

        if listview_Model.rowCount() != 0:
            taskInfoJson = self.file_manager.import_Json(currentPath+"/.task_info", userID+"_task_info.json")

            for row in range(listview_Model.rowCount(QModelIndex())):
                index = listview_Model.index(row)
                item = listview_Model.data(index, Qt.DisplayRole)

                project = connect_shot_proj[item]
                endDate_py = connect_shot_schedule[item][1]

                if endDate_py: # 스케쥴 지정이 되어있는경우
                    date_obj = datetime.strptime(endDate_py, "%Y-%m-%d %H:%M:%S")
                    endDate_obj = QDate(date_obj.year, date_obj.month, date_obj.day)
                    
                    listviewDate = QDate((listView_date["year"]), (listView_date["month"]), (listView_date["day"]))

                    today = QDate.currentDate()
                    monday = today.addDays(-today.dayOfWeek()+1)
                    friday = monday.addDays(4)

                    if listviewDate  == endDate_obj: # 엔드데이트와 현재리스트뷰 날짜가 일치하는경우 
                        color = QColor(94,29,35)                    
                        listview_Model.set_background_color(row, color)                    

                    elif (monday<=listviewDate<=friday) and (listviewDate  != endDate_obj) and (monday<=endDate_obj<=friday): # 엔드데이트가 속해있는 주간내에 있는경우 옅은색 표시
                        color = QColor(94,29,35, 50)                    
                        listview_Model.set_background_color(row, color)                    

                    elif listviewDate > endDate_obj: # 스케쥴이 넘어간 위치인경우 검은색 표시
                        color = QColor(0,0,0)                    
                        listview_Model.set_background_color(row, color)                    

                    else: # 그외 색없음
                        color = QColor(0,0,0,0)                     
                        listview_Model.set_background_color(row, color)                    

                else: # 스케쥴이 없는경우 색없음
                    color = QColor(0,0,0,0)                    
                    listview_Model.set_background_color(row, color)                    





    def convert_to_resultData_forMirror(self, tasks, members):

        allProj = []
        allProj = list(self.projects.keys())
        taskInfo = {}
        proj_member_shots={}
        task_manday_list = []
        task_manday_status_list = []
        mirror_proj_task_dic = {}

        bidEdit_json = self.file_manager.import_Json(currentPath+"/.temp_BDmanday", "edited_mandayData.json")

        for proj in allProj:
            member_shot = {}
            progressRate = 0

            for member in members:

                taskInfoJson = self.file_manager.import_Json_Thread(currentPath+"/.task_info", userID+"_task_info.json")

                if member in list(taskInfoJson.keys()):
                    taskInfo[member] = taskInfoJson[member]

                elif member not in list(taskInfoJson.keys()):
                    taskInfo[member] = self.getTaskInfo(member)
                    taskInfoJson[member] = taskInfo[member]
                    self.file_manager.export_Json(currentPath+"/.task_info", userID+"_task_info.json", taskInfoJson)


                process = ""
                for i in range(len(taskInfo[member])):

                    projCode = taskInfo[member][i]["project_code"]
                    if projCode == self.projects[proj.lower()][0]:
                        task = taskInfo[member][i]["extra_code"]
                        context = taskInfo[member][i]["context"]
                        if  context.find('/') != -1:
                            contextList = context.split('/')
                            task = task + "/" + contextList[1]

                        dep_process = taskInfo[member][i]["process"]
                        if dep_process == "matchmove": dep = "MMV"
                        elif dep_process == "animation": dep = "ANI"
                        elif dep_process == "creature": dep = "RIG"

                        if tasks:
                            for project in list(tasks.keys()):
                                if project == proj:
                                    taskList = list(tasks[proj].keys())
                                    if (task in taskList) and (tasks[proj][task] == dep):

                                        process = taskInfo[member][i]["process"]
                                        bd_manday = (taskInfo[member][i]["bid_manday"])
                                        act_manday = taskInfo[member][i]["active_manday"]
                                        status = taskInfo[member][i]["status"]

                                        if bd_manday == None:
                                            bd_manday = "None"

                                        if act_manday == None:
                                            act_manday = "None"

                                        progressRate = str(act_manday) + "/" + str(bd_manday)

                                        # 파트 지정
                                        if process == "matchmove":
                                            department = "M"

                                        elif process == "animation":
                                            department = "A"

                                        elif process == "creature":
                                            department = "R"

                                        task_manday_status = [proj.upper(), task, progressRate, status, department]

                                        if task_manday_status not in task_manday_status_list:
                                            task_manday_status_list.append(task_manday_status)

                                        if proj in mirror_proj_task_dic:
                                            if task_manday_status not in mirror_proj_task_dic[proj]:
                                                mirror_proj_task_dic[proj].append(task_manday_status)

                                        if proj not in mirror_proj_task_dic:
                                            mirror_proj_task_dic[proj] = []
                                            mirror_proj_task_dic[proj].append(task_manday_status)


        if task_manday_status_list:
            # 맨데이가 수정된 경우 미러데이터에 이를 업데이트하기
            for proj in mirror_proj_task_dic:
                update_tasks = []
                for task in mirror_proj_task_dic[proj]:
                    for edit_task in bidEdit_json:
                        if edit_task[0] == task[1] and edit_task[2] == task[4]:
                            act = task[2].split("/")[0]
                            bid = edit_task[1]
                            manday = act+"/"+bid
                            task[2] = manday
                            update_tasks.append(task)

            return task_manday_status_list, mirror_proj_task_dic





    def convert_to_tuple(self, tasks, members):

        mirrorData = self.file_manager.import_Json(currentPath+"/.scheduleData_mirror", userID+"_scheduleMirror.json")
        selectedProj = self.get_selected_projects()

        taskInfo = {}
        proj_member_shots={}

        task_manday_list = []
        task_manday_status_list = []

        task_manday_list__ = []
        task_manday_status_list__ = []

        userTaskInfo_path = currentPath+"/.task_info/" + userID +"_task_info.json"
        taskInfoJson = self.file_manager.import_Json(currentPath+"/.task_info", userID+"_task_info.json")


        for sel_proj in selectedProj:
            member_shot = {}
            progressRate = 0
            taskTuple=()
            grade="G"

            taskInfo = self.get_cached_task_info(members, taskInfoJson)

            for member in members:
                process = ""
                for i in range(len(taskInfo[member])):
                    task = taskInfo[member][i]["extra_code"]
                    context = taskInfo[member][i]["context"]
                    if  context.find('/') != -1:
                        contextList = context.split('/')

                        numContext=contextList.count("/")

                        full_context = ""
                        for j in range(len(contextList)):
                            if j>0:
                                full_context += "/"+contextList[j]

                        task=task+full_context

                    dep_process = taskInfo[member][i]["process"]
                    if dep_process == "matchmove": dep = "MMV"
                    elif dep_process == "animation": dep = "ANI"
                    elif dep_process == "creature": dep = "RIG"

                    projCode = taskInfo[member][i]["project_code"]

                    if projCode == self.projects[sel_proj.lower()][0]:
                        if tasks:
                            for task_dep in tasks:
                                taskList = list(task_dep.keys())

                                if (task in taskList) and (task_dep[task][1] == dep) and (task_dep[task][0] == sel_proj.lower()):
                                    process = taskInfo[member][i]["process"]
                                    bd_manday = (taskInfo[member][i]["bid_manday"])
                                    act_manday = taskInfo[member][i]["active_manday"]
                                    status = taskInfo[member][i]["status"]

                                    if bd_manday == None:
                                        bd_manday = "None"

                                    if act_manday == None:
                                        act_manday = "None"

                                    progressRate = str(act_manday) + "/" + str(bd_manday)

                                    # 파트 지정
                                    if process == "matchmove":
                                        department = "M"

                                    elif process == "animation":
                                        department = "A"

                                    elif process == "creature":
                                        department = "R"

                                    task_manday_status = (sel_proj.upper(), task, progressRate, status, department)
                                    task_manday_status_list.append(task_manday_status)

        if task_manday_status_list:
            task_manday_status_list = self.convert_mandayEditList(task_manday_status_list)
            return task_manday_status_list


                    # 지난 프로젝트를 선택해서 리스트를 표시하는 기능을 사용하게 되면 쓰게될 부분 - 삭제하지 말것
    """
                    else:
                        if task in mirrorDataList[0] and mirrorDataList[0][task] == dep_process:
                            keys_list = list(mirrorDataList[0].keys())
                            index = keys_list.index(task)
                            
                            scheduleMirror_List = mirrorData[index]
                            
                            m_grade = scheduleMirror_List[0]
                            m_task = scheduleMirror_List[1]
                            m_progressRate = scheduleMirror_List[2]
                            m_status = scheduleMirror_List[3]
                            m_department = scheduleMirror_List[4]

                            scheduleMirror_tuple = (m_grade, m_task, m_progressRate, m_status, m_department)                            

                            if scheduleMirror_tuple not in task_manday_status_list:
                                task_manday_status_list.append(scheduleMirror_tuple)
    """



    def getCurrentListViewDate(self, shiftDays):
        numberListView = self.dayUi.splitter.count()

        listView_labels = []
        for i in range(numberListView):
            addNum = shiftDays + (-1*((numberListView-1) - i))
            listViewDate = self.rangeEndDate.addDays(addNum)
            date = {"year":listViewDate.year() , "month":listViewDate.month(), "day":listViewDate.day()}
            listView_labels.append(date)

        return listView_labels


    # 한번도 스케쥴 데이터가 생성된적이 없는경우 초기 제이슨파일 생성
    def makeInit_data(self, path, userID):

        taskInfo = {}
        taskInfo[userID] = self.getTaskInfo(userID)

        # 현재사용자의 한글이름을 가져오기 위해 유저정보에 가져오기
        userInfoJsonPath = path + "/.user_Info"
        userInfo = self.file_manager.import_Json(userInfoJsonPath, userID+"_userInfo.json")

        taskJsonPath = path + "/.task_info"
        self.file_manager.export_Json(taskJsonPath, userID+"_task_info.json", taskInfo)

        # 초기 스케쥴 데이타 생성
        existingData = jsonData
        currentYear = ""
        currentMonth = ""
        currentDay = ""

        if existingData != False:
            self.refereshListViews(0, existingData)

        elif (os.path.exists(jsonFilePath)==False):

            currentMonth = self.comboMonth.currentText()
            currentDay = self.comboDay.currentText()
            currentYear = self.comboYear.currentText()    




    def changeComboDate(self):

        numListview = self.listViewNumLineEdit.text()
        newEndMonth = self.comboMonth.currentText()
        newEndDay = self.comboDay.currentText()
        newEndYear = self.comboYear.currentText()

        newEndDate = QDate(int(newEndYear), int(newEndMonth), int(newEndDay))
        diffrenceDays = self.rangeEndDate.daysTo(newEndDate)

        self.updateJson(diffrenceDays)
        self.setListviewDate(newEndDate, numListview)
        self.rangeEndDate = newEndDate

        layouts = []
        for i in range(int(numListview)):
            widget = self.dayUi.splitter.widget(i)
            layout = widget.layout()

            layouts.append(layout)

            if layout:
                frame = layout.itemAt(2).widget()
                frameLayout = frame.layout()
                listView = frameLayout.itemAt(0).widget()

                self.manager.Move_sideway_setContainer(frame, layout)





    def getListviewDate(self):

        currentMonth = self.comboMonth.currentText()
        currentDay = self.comboDay.currentText()
        currentYear = self.comboYear.currentText()        

        weekDay = {1:'Mon', 2:'Tue', 3:'Wed', 4:'Thu', 5:'Fri', 6:'Sat', 7:'Sun'}
        currentDate = QDate(int(currentYear), int(currentMonth), int(currentDay))

        dateList = []
        for i in range(5):
            month = (QDate(int(currentYear), int(currentMonth), int(currentDay)).addDays(i-2)).month()
            day = (QDate(int(currentYear), int(currentMonth), int(currentDay)).addDays(i-2)).day()            
            dow = (QDate(int(currentYear), int(currentMonth), int(currentDay)).addDays(i-2)).dayOfWeek()
            date = str(month) + "."+str(day) + "  ("+weekDay[dow]+")"
            dateList.append(date)

        self.setListviewDate(currentDate)

        return dateList




    def setListviewDate(self, setDate, numListview):

        dateList=[]
        for i in range(int(numListview)):
            addNum = int(numListview)-(i+1)
            dateList.append(QDate(setDate).addDays(-1*addNum))

        weekDay = {1:'월', 2:'화', 3:'수', 4:'목', 5:'금', 6:'토', 7:'일'} 

        # set date label
        for i in range(int(numListview)):

            widget = self.dayUi.splitter.widget(i)
            layout = widget.layout()

            if layout:
                item_0 = layout.itemAt(0)
                item_1 = layout.itemAt(2)
                label_listView = item_0.widget()

                frame = layout.itemAt(2).widget()
                frameLayout = frame.layout()
                listView = frameLayout.itemAt(0).widget()

                dow = dateList[i].dayOfWeek()
                label = str(dateList[i].month()) + "/" + str(dateList[i].day()) +  " ( "+weekDay[dow]+" )"
                label_listView.setText(label)


                if dow == 6 or dow == 7:
                    listView.setStyleSheet("QListView { background-color: #232323; }")
                else:                
                    listView.setStyleSheet("QListView { background-color: #2a2a2a; }")

        startDate = QDate(setDate).addDays(-1*int(numListview)+1)
        self.startDay.setText(str(QDate(startDate).day()))
        self.startMonth.setText(str(QDate(startDate).month()))
        self.rangeEndDate = setDate # 엔드데이트를 전역변수에 저장
        


    def forwardDay(self):
        forwardDate = self.move_sideways(1, jsonData)
        self.rangeEndDate = forwardDate # 전역변수 업데이트



    def backwardDay(self):
        # 리스트뷰를 한칸 옮기기전 현재데이타 저장.
        backwardDate = self.move_sideways(-1, jsonData)
        self.rangeEndDate = backwardDate # 전역변수 업데이트



    def setDate(self):

        currentDate = QDate.currentDate()
        year = currentDate.year()
        month = currentDate.month()
        day = currentDate.day()
        dow = currentDate.dayOfWeek()

        lastDate = QDate(year, month, 1).addMonths(1).addDays(-1)
        lastDay = lastDate.day()
        
        Days = []
        Months = ["1","2","3","4","5","6","7","8","9","10","11","12"]
        Years = [str(year-2), str(year-1), str(year), str(year+1), str(year+2)]
        for i in range(lastDay):
            Days.append(str(i+1))

        return year, month, day, dow, Years, Months, Days



    # 트리뷰 루트아이템 이하 모든 아이템을 리스트에 저장하는 재귀함수   
    def getAllItem_treeView(self, item, allItems):

        allItems.append(item.text())
        for row in range(item.rowCount()):
            child_item = item.child(row)
            self.getAllItem_treeView(child_item, allItems)

        return allItems



    # 트리뷰 루트아이템 이하 모든 아이템들중 체크박스 체크되어있는 아이템만 리스트에 저장하는 재귀함수
    def get_CheckedItems(self, item, checkedItems): 
        """팀 관리자의 get_checked_items 메소드를 호출합니다."""
        self.team_manager.get_checked_items(item, checkedItems)



    # 트리뷰를 새로고침하기전 저장한 체크박스 on되어있는 리스트변수를 참조하여 새로고침후 다시 원래대로 체크박스 on
    def set_CheckItems(self, item, checkedItems):
        """팀 관리자의 set_check_items 메소드를 호출합니다."""
        self.team_manager.set_check_items(item, checkedItems)



    # 멤버트리뷰의 각 멤버의 체크박스를 온오프에 작동하는 메서드
    def clickedAction_memberCheck(self, index): 

        itemChecked = self.teamTreeModel.data(index, Qt.CheckStateRole)
        invisibleRoot = self.teamTreeModel.invisibleRootItem()
        rootItem = invisibleRoot.child(0)

        checked_items_list = []
        self.get_CheckedItems(rootItem, checked_items_list)                             # 트리뷰를 새로고침 하기전 체크되어있는 상태의 아이템리스트 가져오기
        self.teamInfo = self.edit_teamTree_item(self.teamTreeModel.invisibleRootItem()) # 새로 추가되거나 삭제된 멤버에 대한 체크박스상태를 반영하기 위해서 새로고침
        self.set_CheckItems(rootItem, checked_items_list)                               # 트리뷰를 새로고침한후 원래대로 체크박스 체크해주기          

        # 프로젝트 자동선택이 꺼져있고, 트리뷰 컨택스트 메뉴실행이 아닌경우
        if self.ui.sel_allProj_checkBox.checkState() == Qt.Unchecked and self.select_ctxMenu == 0:
            self.reloadShotList()

        # 프로젝트 자동선택이 켜져있고있고, 트리뷰 컨택스트 메뉴실행이 아닌상태이고, 싱글 멤버 선택인 경우
        elif self.ui.sel_allProj_checkBox.checkState() == Qt.Checked and self.select_ctxMenu == 0: # and len(checked_items_list) == 1 or len(checked_items_list) == 0:

            self.reloadShotList()
            items = len(checked_items_list)
            self.sel_assigned_proj("single", index, items)
            self.refereshListViews(0, jsonData)            


    # 제이슨으로 저장된 리더의 팀멤버가 있다면 정보를 읽어옴

    def edit_teamTree_item(self, parent_root):

        global teamInfo_

        # 현재유저의 팀원정보 가져오기
        teamInfoData = self.team_manager.get_team_info(userID)

        if userID not in self.memberCache_:
            self.memberCache_[userID] = self.currName_kr

        if teamInfoData != []:
            existMember_inJson = list(teamInfoData[0].keys())
        else:
            existMember_inJson = []            

        for mem in existMember_inJson:
            if mem not in self.memberCache_:
                self.memberCache_[mem] = teamInfoData[0][mem]["nameKr"]

        del_mem = []
        for mem in self.memberCache_:
            if mem not in existMember_inJson:
                del_mem.append(mem)

        for mem in del_mem:
            if mem != userID and mem in self.memberCache_:
                del self.memberCache_[mem]

        invisibleRoot = self.teamTreeModel.invisibleRootItem()
        rootItem = invisibleRoot.child(0)

        checked_items_list = []
        unchecked_items_list = []

        new_hierarchy = self.team_manager.get_tree_all_items(rootItem, checked_items_list, unchecked_items_list, existMember_inJson, teamTreeHierarchy, 0)

        # 중복되는 멤버의 경우 teamInfo파일에는 json형태로 존재하지만, treeView에는 표시되지 않고 있던 상태에서,
        # 중복되는 멤버중 하나를 삭제해서 원래있던 멤버가 다시 treeView에 표시되게 하기 위해, 트리뷰에서 현재 안보이는 상태의 멤버의 이름을,
        # memberCache 에서 읽어와 unchecked_items_list 에 저장
        for memberName in list(self.memberCache_.values()):
            if memberName not in unchecked_items_list:
                unchecked_items_list.append(memberName)

        all_treeMemberList = list(set(checked_items_list + unchecked_items_list))

        memberList = []
        self.print_all_items(memberList, None)

        if teamInfoData != []:
            child_Members = []
            child_teamInfo = []
            allTeamInfos = [{}]

            for member in teamInfoData[0]:
                self.checkUserInfo(member) # member의 정보가 user_info 폴더에 json파일이 있는지 확인후 없으면 json 파일 생성
                memberNameKr = teamInfoData[0][member]['nameKr']

                # 중복되는 멤버가 있다면 doubledMem 에 저장
                doubledMem = []
                for i in new_hierarchy:
                    if i>2 and new_hierarchy:
                        if memberNameKr in new_hierarchy[i]:
                            doubledMem.append(memberNameKr)

                memberItem = QStandardItem(memberNameKr)
                memberItem.setCheckable(True)

                if member not in list(self.memberCache_.keys()):
                    self.memberCache_[member] = memberNameKr

                if memberNameKr in checked_items_list:
                    memberItem.setCheckState(Qt.Checked)

                if memberNameKr not in memberList:
                    self.teamLeader.appendRow(memberItem)

                if memberNameKr not in all_treeMemberList:
                    all_treeMemberList.append(memberNameKr)

                #### 팀원의 삭제로 변화가 생긴 정보를 all_treeMemberList에 반영하기 위해 cacheMember_ 변수를 확인하여 바뀐정보를 반영함
                del_members = []
                for mem in all_treeMemberList:
                    if mem not in list(self.memberCache_.values()):
                        del_members.append(mem)

                for del_mem in del_members:
                    if del_mem != self.currName_kr:
                        all_treeMemberList.remove(del_mem)

                # 최상위 리더의 바로 아래 리더들의 하위 멤버들중, 중복되는 멤버가 있을경우 최상위 리더의 중복멤버는 트리뷰에서 삭제
                # (중복되면 체크박스가 정상적으로 작동하지 않음)
                rows_to_delete = []
                if memberNameKr in doubledMem:
                    model = self.teamTree.model()
                    teamLeader_index = model.indexFromItem(self.teamLeader)
                    child_count = model.rowCount(teamLeader_index)

                    for i in range(child_count):
                        child_index = model.index(i,0,teamLeader_index)
                        child_item =  model.itemFromIndex(child_index)

                        if memberNameKr == child_item.text():
                            rows_to_delete.append(i)
                            
                for row in reversed(rows_to_delete):
                    self.teamLeader.removeRow(row)

                # 각 리더의 하위 팀원들을 팀트리뷰에 추가함
                members, teamInfos = self.get_teamInfo_member(member, memberItem, all_treeMemberList, checked_items_list, existMember_inJson, allTeamInfos, 2)
                child_Members = members + child_Members
                child_teamInfo = allTeamInfos + child_teamInfo


            if child_Members != None: # 팀원중에 리더가 있어서 차일드로 팀원들을 가지고 있는경우
                all_treeMemberList = list(set(all_treeMemberList + child_Members))

                if child_teamInfo != []:
                    for info in child_teamInfo[0]:

                        if info not in teamInfoData[0]:
                            teamInfoData[0][info] = child_teamInfo[0][info]

                #< 삭제금지 >
                #print ("## 아래 세가지 데이터가 일치해야함")
                #print (all_treeMemberList) #현재 트리뷰 ui상의 멤버리스트
                #print (existMember_inJson) # 현재 유저의 외부 teamInfo.json 상의 팀원리스트
                #print (self.memberCache_)       # 현재상태의 전체 팀원의 리스트를 담고있는 전역변수
                #print ("##")

            self.teamTree.expandAll()
            teamInfo_ = teamInfoData

            return teamInfoData


    # 트리뷰 모델내에 있는 아이템을 모두 출력하는 테스트 메서드
    def print_all_items(self, memberList, parent_item=None):
                
        if parent_item == None:
            parent_item  = self.teamTreeModel.invisibleRootItem()

        for row in range(parent_item.rowCount()):
            child_item = parent_item.child(row,0)
            if child_item:
                member = child_item.text()
                memberList.append(member)
                self.print_all_items(memberList, child_item)



    # 재귀함수(하위 팀원중 리더인 팀원의 팀원들까지 읽어오기 위한 재귀함수)
    def get_teamInfo_member(self, teamMember, teamMemberItem, all_treeMemList, checkedItems, existMem, allInfos, hi):

        hi = hi+1
        added_memberList = []
        member_teamInfo = self.team_manager.get_team_info(teamMember)
        if member_teamInfo != []:
            for info_member in list(member_teamInfo[0].keys()):
                allInfos[0][info_member] = member_teamInfo[0][info_member]

        if member_teamInfo != []:
            for member in member_teamInfo[0]:
                if member != userID:
                    self.checkUserInfo(member) # 트리뷰상에 존재하는 팀원의 정보가 user_info 폴더에 json파일이 있는지 확인후 없으면 json 파일 생성
                    memberNameKr = member_teamInfo[0][member]['nameKr']

                    if member not in list(self.memberCache_.keys()):
                        self.memberCache_[member] = memberNameKr
                    
                        if member not in existMem:
                            existMem.append(member)

                    memberItem = QStandardItem()
                    memberItem.setData(memberNameKr, Qt.DisplayRole)
                    memberItem.setCheckable(True)

                    if memberNameKr in checkedItems:
                        memberItem.setCheckState(Qt.Checked)


                    # 팀원의 팀원을 트리뷰에 표시할때 상위 팀원들에 중복되는 팀원이 있는지 확인하기위해, teamTreeHierarchy 변수의 상위 하이라키를 검색해서 확인하는 코드
                    valueNum = 0
                    if hi !=0:
                        for i in range(hi):
                            if hi < hi-1:
                                num = teamTreeHierarchy[i].count(memberNameKr)
                                valueNum = valueNum + num
                                
                    if valueNum < 2: # 하위멤버가 다른팀의 멤버들과 중복되어 2개이상인지 확인
                        teamMemberItem.appendRow(memberItem)

                    all_treeMemList.append(memberNameKr)
                    added_memberList.append(memberNameKr)

                    self.get_teamInfo_member(member, memberItem,all_treeMemList, checkedItems, existMem, allInfos, hi)
        
        return added_memberList, member_teamInfo






    # 트리뷰의 하위 멤버를 모두 읽어오기 위한 재귀함수


    # 리스트를 값으로 가지고 있는 딕셔너리 안에 특정 값이 존재하는지 확인하는 메서드
    # check_value_in_lists moved to utils/validation.py



    def showTeamTreeView(self):
        global teamTreeHierarchy
        teamInfo = self.edit_teamTree_item(self.teamTreeModel.invisibleRootItem())
        currentMemList = list(teamInfo[0].keys())

        if userID not in currentMemList:
            currentMemList.append(userID) 

        #### jsonData 내에 포함되어있는 삭제된 팀원의 데이타를 찾아서 삭제시키기 ###############################
        delDatas = []
        for data in jsonData: 
            if data['artist'] not in currentMemList:  
                delDatas.append(data)

        for delData in delDatas:
            if delData in jsonData:
                jsonData.remove(delData)
        #####################################################################################################

        ## 멤버가 추가되거나 삭제되는 경우에 멤버트리뷰를 갱신하고, 전역변수 teamTreeHierarchy 에도 멤버 변동사항을 적용함.
        teamInfoData = self.team_manager.get_team_info(userID)

        invisibleRoot = self.teamTreeModel.invisibleRootItem()
        rootItem = invisibleRoot.child(0)

        userMembers = []
        childMembers = []
        allMembers = []
        teamHierarchy = {}
        # 변동사항에 맞게 하이라키를 재구성함.
        self.checkExists_atTree(rootItem, userMembers, childMembers, allMembers,teamHierarchy, currentMemList, 0)

        # 팀원이 더해지는 경우
        for userMember in teamInfoData[0]: # json파일내에 존재하는 현재 유저의 팀원들
            nameKr = self.memberCache_[userMember]
            if nameKr not in userMembers: # userMembers는 현재 트리뷰ui상에 존재하는 팀원들
                if nameKr not in allMembers: # 중복되는 멤버들이 있는지 확인

                    memberItem = QStandardItem(nameKr)
                    memberItem.setCheckable(True)                    
                    self.teamLeader.appendRow(memberItem)
                    teamHierarchy[2].append(nameKr)

        # 팀원을 삭제하는 경우
        for userMember in allMembers:
            if userMember not in list(self.memberCache_.values()):
                parent_index = self.teamTreeModel.index(0,0,QModelIndex())
                for row in range(self.teamTreeModel.rowCount(parent_index)):
                    child_index = self.teamTreeModel.index(row,0, parent_index)
                    item = self.teamTreeModel.itemFromIndex(child_index)
                    
                    if item and item.text() == userMember:
                        self.teamTreeModel.removeRow(row, parent_index)


        # 재구성된 하이라키를 전역변수 teamTreeHierarchy에 덮어씀.
        teamTreeHierarchy = teamHierarchy # 멤버가 추가된경우 서로 다른 하이라키에 중복되는 멤버가 있을수 있으나, 아래 getJsonData 내의 edit_teamTree_item 메서드에서 정리됨

        self.getJsonData() # 저장된 json 스케쥴 데이타를 가져와서 jsonData 파일에 저장함
        self.refereshListViews(0, jsonData) # 추가 혹은 삭제된 멤버의 스케쥴데이타가 업데이트된 내용이 적용된 jsonData변수를 바탕으로 ui업데이트
        self.reloadShotList() # 멤버트리뷰가 새로고침된후 그에 맞게 샷리스트뷰 리로드




    def checkExists_atTree(self, item, members, childs, allMembers, teamHierarchy, currentMemList, hi):
        hi = hi+1
        name = item.text()
        allMembers.append(name)

        if hi==2:
            members.append(name)

        if hi>2:
            childs.append(name)

        if hi in teamHierarchy:
            teamHierarchy[hi].append(name)
        else:
            teamHierarchy[hi] = [name]

        for row in range(item.rowCount()):
            child_item = item.child(row)
            self.checkExists_atTree(child_item,  members, childs, allMembers, teamHierarchy, currentMemList,hi)    



    def checkUserInfo(self, user):
        """사용자 정보를 확인하고 반환합니다."""
        # 기존 구조 유지를 위해 TeamManager의 check_user_info 호출 후 기존 로직 사용
        self.team_manager.check_user_info(user)
        
        name_kr = ""
        role = ""
        job = ""
        team = ""
        department = ""

        userInfoPath = currentPath+"/.user_Info"
        userInfo_json = os.path.join(userInfoPath, user+"_userInfo.json")

        userInfoData = self.file_manager.import_Json(userInfoPath, user+"_userInfo.json")

        if userInfoData:
            name_kr = userInfoData[0]["name_kr"]
            job = userInfoData[0]["job"]
            role = userInfoData[0]["role"]
            team = userInfoData[0]["team"]
            department = userInfoData[0]["department"]
        #if user information does not exist, write a json file to userInfoPath
        elif(os.path.exists(userInfo_json)==False):
            name_kr, role, team, job, department = getUserInfo(user)
            userInfo = [{"name_kr": name_kr, "role":role, "team":team, "job":job, "department":department}]

            try:
                with open(userInfo_json, 'w') as file:
                    json.dump(userInfo, file, indent=4)
                #return True
                return name_kr, role, team, job, department

            except Exception as e:
                print (f"Error saving to file: {e} ")
                sys.exit(1)

        return name_kr, role, team, job, department



    # 팀원 편집창 실행 메서드
    def showMemberEditDialog(self):
        dialogUI = memberDialog.EditMemberDialog()
        dialogUI.setWindowTitle("Edit Member")    
        dialogUI.exec_()
        self.showTeamTreeView()
       
    def showWorkdataWin(self):
        pass
        
    #메인창을 스케쥴창으로 변환        
    def showScheduleFrame(self):
        self.dayScheduleFrame.setVisible(True)
        self.dashboardFrame.setVisible(False)
        self.leftSideInfoFrm.setVisible(True)
        self.dayScheduleFrame.raise_()
        

    #메인창을 대쉬보드창으로 변환
    def showDashboardFrame(self):
        self.dayScheduleFrame.setVisible(False)
        self.dashboardFrame.setVisible(True)
        self.leftSideInfoFrm.setVisible(False)
        self.dashboardFrame.raise_()


    # 제이슨 파일로 저장하고 사용자에게 다이얼로그 띄움
    def saveFile(self):
        existJsonData = self.updateJson(0)

        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Save")
        msg_box.setText("현재 스케쥴을 저장하시겠습니까?")
        msg_box.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        msg_box.setDefaultButton(QMessageBox.Ok)
        result = msg_box.exec()

        if result == QMessageBox.Ok:
            if(self.scheduleExportJson()):
                QMessageBox.information(self,"save", "현재 스케쥴이 저장되었습니다")
            else:
                QMessageBox.critical(None, "Error", "스케쥴 저장에 실패하였습니다.")  




    # 제이슨 파일로 저장
    def scheduleExportJson(self):

        # 커서
        QApplication.setOverrideCursor(Qt.WaitCursor)

        existJsonData = self.updateJson(0)

        userID_list = []
        for data in existJsonData:
            if data['artist'] not in userID_list:
                if data['tasks'] != [{}]:
                    userID_list.append(data['artist'])


        for user in userID_list:
            user_jsonDatas=[]
            for data in existJsonData:
                projs = list(data['tasks'][0].keys())
                if data['tasks'] != [{}] and data['artist']==user:
                    for proj in projs:
                        if data['tasks'][0][proj] == []:
                            del data['tasks'][0][proj]

                        # 저장할 제이슨 데이타가 아직 비어있는 초기에 우선 첫번째 데이터 추가.
                        if user_jsonDatas == []:
                            user_jsonDatas.append(data)

                        elif data not in user_jsonDatas: # 위 구문에서 첫번째 데이터가 추가된 상태에서 두번째 데이터가 저장될때에는 저장된 프로젝트나 태스크의 중복여부를 검사하여 분류해서 저장.
                           for userJsonData in user_jsonDatas:
                                if userJsonData['year'] == data['year'] and userJsonData['month'] == data['month'] and userJsonData['day'] == data['day']:
                                    if proj in list(userJsonData['tasks'][0].keys()):
                                        if (proj in list(userJsonData['tasks'][0].keys())) and (userJsonData['tasks'][0][proj] != []):
                                            for task in data['tasks'][0][proj]:
                                                if task not in userJsonData['tasks'][0][proj]:
                                                    userJsonData['tasks'][0][proj].append(task)

                                        elif  proj not in list(userJsonData['tasks'][0].keys()):
                                            userJsonData['tasks'][0][proj] = data['tasks'][0][proj]

                                else:
                                    if data not in user_jsonDatas:
                                        user_jsonDatas.append(data)

            saveData = []
            for userData in user_jsonDatas:
                if userData['tasks'] == [{}]:
                    del userData
                else:
                    saveData.append(userData)

            jsonFilePath = currentPath+"/.scheduleData"
            jsonFile = os.path.join(jsonFilePath, user+"_Data.json")

            try:
                with open(jsonFile, 'w') as file:
                    json.dump(saveData, file, indent=4)

            except Exception as e:
                print (f"Error saving to file: {e} ")
                return False

            # 스케쥴 파일 백업
            backup_dir = "schedule"
            backup_name = user +"_Data"

        # 스케쥴 데이터와 연결되는 맨데이, 스테이터스, 평가 등의 데이타를 백업
        self.makeMirrorSchedule(jsonData)

        # 분기별로 분류된 스케쥴 데이타를 각 멤버별 이름 파일로 저장
        self.export_QuarterlyWorkload()

        # 커서 
        QApplication.restoreOverrideCursor()
        return True



    # 백업파일 생성


    # 백업파일의 오래된 파일 삭제






    # 이벤트루프와 쓰레드가 동시에 사용하게 되면 충돌이 생기는것으로 보여 다른이름으로 만든 import_Json과 동일한 메서드







    # 이벤트루프와 쓰레드가 동시에 사용하게 되면 충돌이 생기는것으로 보여 다른이름으로 만든 export_Json과 동일한 메서드




    def findProj(self, taskName):

        showCode = ""      
        for data in jsonData:
            for projDatas in data["tasks"]:
                for show, tasks in projDatas.items():
                    if taskName[1] in tasks:
                        showCode = show

        if(showCode != ""):
            return showCode

        else:
            return "notExist"            



    # 스케쥴에 어싸인된 태스크의 색 변환
    def changeColor_assignedTask(self):
        checkedMember = []
        invisibleRoot = self.teamTreeModel.invisibleRootItem()
        rootItem = invisibleRoot.child(0)

        teamInfo = [{}]
        for mem in list(self.memberCache_.keys()):
            memberInfo = self.team_manager.get_team_info(mem)
            if memberInfo != []:
                for teamMem in memberInfo[0]:
                    if teamMem not in teamInfo[0]:
                        teamInfo[0][teamMem] = memberInfo[0][teamMem]

        checkedMember = []
        self.get_CheckedItems(rootItem, checkedMember)

        # 멤버트리뷰의 체크된 멤버 리스트의 영문아이디 읽어오기
        memberEN = []
        if rootItem.checkState() == Qt.Checked:
            memberEN.append(userID)
            if userID not in self.memberEN:
                self.memberEN.append(userID)

        for member in teamInfo[0]:
            if teamInfo[0][member]["nameKr"] in checkedMember:
                memberEN.append(member)                    
                if member not in self.memberEN:
                    self.memberEN.append(member)

        taskListDic = self.get_ShotlistDic(memberEN) # 멤버 트리뷰에서 팀원의 선택 및 해제에 맞게 샷 리스트를 업데이트
        containDic = self.containsTask(taskListDic)  # 현재 배치된 스케쥴에 해당 태스크가 포함되어있는지 확인하는 메서드

        # 선택되어져 있던 아이템들이 리프레시된 이후에 선택된 상태를 유지시키기 위해 선택되어졌던 아이템 리스트 만들기
        selected_indexes = self.shotListview.selectedIndexes()
        selectedItems = [index.data() for index in selected_indexes]

        delegate = ddlv.ProgressDelegate(containDic)
        self.shotListview.setItemDelegate(delegate)

        shotlist_model = self.shotListview.model()

        # 리프레시 하기전 셀렉트 되어있던 상태대로 다시 셀렉팅
        if (shotlist_model != None):
            for row in range(shotlist_model.rowCount()):
                index = shotlist_model.index(row)
                item = shotlist_model.data(index, Qt.DisplayRole)
                if item in selectedItems:
                    self.shotListview.selectionModel().select(index, QItemSelectionModel.Select)



    # 샷 리스트뷰의 리스트 갱신
    def reloadShotList(self):

        QApplication.setOverrideCursor(Qt.WaitCursor)

        global connect_shot_proj

        checkedMember = []
        invisibleRoot = self.teamTreeModel.invisibleRootItem()
        rootItem = invisibleRoot.child(0)

        teamInfo = [{}]
        for mem in list(self.memberCache_.keys()):
            memberInfo = self.team_manager.get_team_info(mem)
            if memberInfo != []:
                for teamMem in memberInfo[0]:
                    if teamMem not in teamInfo[0]:
                        teamInfo[0][teamMem] = memberInfo[0][teamMem]

        checkedMember = []
        self.get_CheckedItems(rootItem, checkedMember)

        # 멤버트리뷰의 체크된 멤버 리스트의 영문아이디 읽어오기
        memberEN = []
        if rootItem.checkState() == Qt.Checked:
            memberEN.append(userID)
            if userID not in self.memberEN:
                self.memberEN.append(userID)

        for member in teamInfo[0]:
            if teamInfo[0][member]["nameKr"] in checkedMember:
                memberEN.append(member)                    
                if member not in self.memberEN:
                    self.memberEN.append(member)


        if self.select_ctxMenu == 1 or self.reload_tatic == 1:  # 컨텍스트 메뉴 select All 이거나 refresh 버튼이 눌려진 경우
            taskListDic = self.get_ShotlistDic_MT(memberEN) # 멤버 트리뷰에서 팀원의 선택 및 해제에 맞게 샷 리스트를 업데이트  

        elif self.select_ctxMenu != 1:
            taskListDic = self.get_ShotlistDic(memberEN) # 멤버 트리뷰에서 팀원의 선택 및 해제에 맞게 샷 리스트를 업데이트  


        containDic = self.containsTask(taskListDic)  # 현재 배치된 스케쥴에 해당 태스크가 포함되어있는지 확인하는 메서드

        #< 삭제금지>
        ###   selectAll 컨텍스트 메뉴가 실행되면,  팀원이 많은경우 멀티 쓰레드상 데이타를 읽어오것이 느려져서 아래 if문 아래 프로세스가
        ###   먼저 진행되어 샷리스트뷰가 꺼졌다 켜지는 현상이 발생함. 이 현상을 없애기 위해 아래 if 문을 추가하였음
        ###   selectAll 로 reloadShotListView가 실행된경우 아래 if 문에 의해 if문 아래 내용은 건너뛰게 되며
        ###   대신 멀티 쓰레드가 끝난후 프로젝트 리스트뷰의 프로젝트 선택 프로세스가 진행될때 호출되는 reloadShotListView는 
        ###   컨텍스트 메뉴(selectAll) 호출이 아니므로 아래 if문 이후의 프로세스가 진행되어 샷 리스트뷰가 갱신됨
        if self.select_ctxMenu != 1 or self.reload_tatic == 1:  # 컨텍스트 메뉴 select All 이 아니거나 refresh 버튼이 눌려진 경우

            # 선택되어져 있던 아이템들이 리프레시된 이후에 선택된 상태를 유지시키기 위해 선택되어졌던 아이템 리스트 만들기
            selected_indexes = self.shotListview.selectedIndexes()
            selectedItems = [index.data() for index in selected_indexes]

            # 샷 리스트뷰의 모델을 새로 만들어 샷리스트뷰에 새로 등록
            old_model = self.shotListview.model() # 기존 모델

            taskList = list(taskListDic.keys())

            shotlist_model = ddlv.DragDropModel(taskList, containDic, self.sort_column) 
            self.shotListview.setModel(shotlist_model)

            del old_model # 샷리스트뷰의 기존모델은 삭제

            delegate = ddlv.ProgressDelegate(containDic)
            self.shotListview.setItemDelegate(delegate)

            self.shotListview.setEditTriggers(QAbstractItemView.NoEditTriggers) # 리스트뷰내 아이템을 더블클릭하였을때 편집모드가 작동되지 않도록 함

            # 리프레시 되기전의 정렬상태와 같이 일치시키기
            self.sort_manager.sortShotlistview()

            # 리프레시 하기전 셀렉트 되어있던 상태대로 다시 셀렉팅
            if (shotlist_model != None):
                for row in range(shotlist_model.rowCount()):
                    index = shotlist_model.index(row)
                    item = shotlist_model.data(index, Qt.DisplayRole)
                    if item in selectedItems:
                        self.shotListview.selectionModel().select(index, QItemSelectionModel.Select)

            # 스테이터스의 뷰 셋업을 샷리스트뷰에도 적용시키기
            self.set_listView_view(self.shotListview, shotlist_model)

        QApplication.restoreOverrideCursor()




    # 태스크리스트 아이템들의 백그라운드 색깔 지정하기
    def setItemColor_taskList(self):

        shotList_model = self.shotListview.model()
        if shotList_model.rowCount() != 0:
            taskInfoJson = self.file_manager.import_Json(currentPath+"/.task_info", userID+"_task_info.json")

            today = QDate.currentDate()
            monday = today.addDays(-today.dayOfWeek()+1)
            friday = monday.addDays(4)

            for row in range(shotList_model.rowCount(QModelIndex())):
                index = shotList_model.index(row)
                item = shotList_model.data(index, Qt.DisplayRole)

                project = connect_shot_proj[item]
                endDate_py = connect_shot_schedule[item][1]

                if endDate_py: # 스케쥴 지정이 되어있는경우
                    date_obj = datetime.strptime(endDate_py, "%Y-%m-%d %H:%M:%S")
                    Qdate_obj = QDate(date_obj.year, date_obj.month, date_obj.day)
                    
                    if monday <= Qdate_obj <= friday:#QcurrentDate  == Qdate_obj: # 엔드데이트와 현재리스트뷰 날짜가 일치하는경우 
                        color = QColor(94,29,35)                    
                        shotList_model.set_background_color(row, color)                    
                    else:
                        color = QColor(0,0,0,0)                     
                        shotList_model.set_background_color(row, color)                    

                else: # 스케쥴이 없는경우
                    color = QColor(0,0,0,0)                    
                    shotList_model.set_background_color(row, color)  





    # 현재 배치된 스케쥴에 해당 태스크가 포함되어있는지 확인하는 메서드
    def containsTask(self, taskListDic):

        containStatus = {}
        checkedMem, uncheckedMem = self.get_teamTree_member() # 현재 선택된 멤버 리스트 가져오기        

        for task in list(taskListDic.keys()):
            proj = taskListDic[task]
            for mem in checkedMem:
                memEN = ""
                for member in self.memberCache_:
                    if self.memberCache_[member] == mem:
                        memEN = member

                for data in jsonData:
                    if data["artist"] == memEN:
                        if proj in data["tasks"][0]:
                            if task[1] in data["tasks"][0][proj]:  
                                containStatus[task] = 1

        return containStatus

            


    # 멤버 트리뷰의 선택한 인덱스의 태스크 정보를 읽어오기 
    def getTaskInfo_member(self, index):

        teamInfo = self.team_manager.get_team_info(userID)

        itemChecked = self.teamTreeModel.data(index, Qt.CheckStateRole)
        if itemChecked == 2:
            memberKr = self.teamTreeModel.data(index)

            memberEN = ''
            taskInfo = {}
            
            for member in teamInfo[0]:
                if self.teamInfo[0][member]["nameKr"] == memberKr:
                    taskInfo[member] = self.getTaskInfo(member)
                    memberEN = member

            return taskInfo[membrEN]




    # 멤버 트리뷰에서 팀원의 선택 및 해제에 맞게 샷 리스트를 업데이트
    def getShotlist_member(self, members):

        selectedProj = self.get_selected_projects()
        taskInfo = {}
        proj_member_shots={}
        for sel_proj in selectedProj:

            member_shot = {}

            for member in members:

                #< 삭제금지 >
                # 1. 현재유저의 태스크정보를 읽어와 json파일에 저장후 memberCache 변수에 저장
                # 2. 멤버트리뷰에 유저가 체크를 추가할때마다 먼저 memberCache 변수 확인후 없다면 해당 체크된 멤버의 태스크 정보를 가져온후 jason 업데이트후 memberCache 저장
                # 3. 메서드 호출시 먼저 memberCache를 확인하고 태스크정보를 가지고 오려는 유저가 memberCache안에 있다면 json파일만 열어서 정보 가져가기
                if member in list(self.taskInfoJson.keys()):
                    taskInfo[member] = self.taskInfoJson[member]                    

                elif member not in list(self.taskInfoJson.keys()): 
                    taskInfo[member] = self.getTaskInfo(member)
                    self.taskInfoJson[member] = taskInfo[member]
                    self.file_manager.export_Json(currentPath+"/.task_info", userID+"_task_info.json", self.taskInfoJson)


                progressRate = 0
                shot_manday_list = []
                shot_manday_status_list = []
                shotList=[]
                grade="G"
                process=""
                for i in range(len(taskInfo[member])):
                    projCode = taskInfo[member][i]["project_code"]

                    if projCode == self.projects[sel_proj.lower()][0]:

                        shot = taskInfo[member][i]["extra_code"]
                        bd_manday = (taskInfo[member][i]["bid_manday"])
                        act_manday = taskInfo[member][i]["active_manday"]
                        status = taskInfo[member][i]["status"]
                        process = taskInfo[member][i]["process"]

                        if act_manday == None:
                            act_manday = "None"

                        if bd_manday == None:
                            bd_manday = "None"                            

                        progressRate = str(act_manday) + "/" + str(bd_manday)


                        context = taskInfo[member][i]["context"]
                        if  context.find('/') != -1:
                            contextList = context.split('/')
                            shot = shot + "/" + contextList[1]

                        shot_manday = (shot, progressRate)
                        shot_manday_list.append(shot_manday)

                        # 파트 지정
                        part=""
                        if process == "matchmove":
                            part = "M"

                        elif process == "animation":
                            part = "A"

                        elif process == "creature":
                            part = "R"

                        shot_manday_status = (grade, shot, progressRate, status, part)
                        shot_manday_status_list.append(shot_manday_status)
                        member_shot[member] = shot_manday_status_list                        

            proj_member_shots[sel_proj] = member_shot


        projs = list(proj_member_shots.keys())
        all_shots = []
        members_proj = []
        for proj in projs:
            members_proj =  list(proj_member_shots[proj].keys())
            for member in members_proj:
                shotList_member = proj_member_shots[proj][member]
                for shot in shotList_member:
                    all_shots.append(shot)


        ###### 각 프로젝트의 어싸인된 태스크들과 해당 프로젝트를 짝지어 전역변수에 저장
        for proj in proj_member_shots:

            members_proj =  list(proj_member_shots[proj].keys())

            for member in members_proj:
                if member in list(proj_member_shots[proj].keys()):
                    for task in proj_member_shots[proj][member]:
                        connect_shot_proj[task] = proj

        all_shots = self.convert_mandayEditList(all_shots)

        # taskInfo에는 현재 선택되어진 멤버의 태스크정보들만 들어있으므로 이 정보로 기존 정보(체크가 해제된 멤버의 태스크정보까지 포함된)를 덮어씀
        self.file_manager.export_Json(currentPath+"/.task_info", userID+"_task_info.json", taskInfo)

        return all_shots  










# 멤버 트리뷰에서 팀원의 선택 및 해제에 맞게 샷 리스트를 업데이트
    def get_ShotlistDic(self, members):

        global connect_shot_schedule

        tempBid_path = currentPath + "/.temp_BDmanday"
        editedTasks_json = self.file_manager.import_Json(tempBid_path, "edited_mandayData.json")

        selectedProj = self.get_selected_projects()

        taskInfoJson = self.file_manager.import_Json(currentPath+"/.task_info", userID+"_task_info.json")

        # 프로젝트 자동선택이 on 일경우 선택된 팀원의 task정보 읽어오기
        taskInfo = {}
        if self.ui.sel_allProj_checkBox.checkState() == Qt.Checked:
            for member in members:
                if self.reload_tatic == 0:
                    if member in list(self.taskInfoJson.keys()):
                        taskInfo[member] = self.taskInfoJson[member]                    


                    elif member not in list(self.taskInfoJson.keys()): 
                        taskInfo[member] = self.getTaskInfo(member)
                        self.taskInfoJson[member] = taskInfo[member]
                        self.file_manager.export_Json(currentPath+"/.task_info", userID+"_task_info.json", self.taskInfoJson)


                elif self.reload_tatic == 1:
                        taskInfo[member] = self.getTaskInfo(member)
                        self.taskInfoJson[member] = taskInfo[member]
                        self.file_manager.export_Json(currentPath+"/.task_info", userID+"_task_info.json", self.taskInfoJson)


        proj_member_shots={}
        for sel_proj in selectedProj:

            member_shot = {}
            for member in members:

                #< 삭제금지 >
                # 1. 현재유저의 태스크정보를 읽어와 json파일에 저장후 memberCache 변수에 저장
                # 2. 멤버트리뷰에 유저가 체크를 추가할때마다 먼저 memberCache 변수 확인후 없다면 해당 체크된 멤버의 태스크 정보를 가져온후 jason 업데이트후 memberCache 저장
                # 3. 메서드 호출시 먼저 memberCache를 확인하고 태스크정보를 가지고 오려는 유저가 memberCache안에 있다면 json파일만 열어서 정보 가져가기

                if self.reload_tatic == 0:
                    if member in list(self.taskInfoJson.keys()):
                        taskInfo[member] = self.taskInfoJson[member]                    

                    elif member not in list(self.taskInfoJson.keys()): 
                        taskInfo[member] = self.getTaskInfo(member)
                        self.taskInfoJson[member] = taskInfo[member]
                        self.file_manager.export_Json(currentPath+"/.task_info", userID+"_task_info.json", self.taskInfoJson)

                elif self.reload_tatic == 1:
                        taskInfo[member] = self.getTaskInfo(member)
                        self.taskInfoJson[member] = taskInfo[member]
                        self.file_manager.export_Json(currentPath+"/.task_info", userID+"_task_info.json", self.taskInfoJson)


                progressRate = 0
                shot_manday_list = []
                shot_manday_status_list = []

                #task_schedule = []
                task_schedule_dic = {}

                shotList=[]
                #grade="G"
                process=""
                for i in range(len(taskInfo[member])):
                    projCode = taskInfo[member][i]["project_code"]

                    if projCode == self.projects[sel_proj.lower()][0]:

                        shot = taskInfo[member][i]["extra_code"]
                        bd_manday = (taskInfo[member][i]["bid_manday"])
                        act_manday = taskInfo[member][i]["active_manday"]
                        status = taskInfo[member][i]["status"]
                        process = taskInfo[member][i]["process"]
                        start_date = taskInfo[member][i]["start_date"]
                        end_date = taskInfo[member][i]["end_date"]

                        if act_manday == None:
                            act_manday = "None"

                        if bd_manday == None:
                            bd_manday = "None"                            

                        progressRate = str(act_manday) + "/" + str(bd_manday)
                        context = taskInfo[member][i]["context"]
                        numContext=context.count("/")

                        if numContext>0:
                            contextList = context.split('/')
                            full_context = ""
                            for i in range(len(contextList)):
                                if i>0:
                                    full_context += "/"+contextList[i]
                            shot=shot+full_context

                        ######## 차후에 삭제예정 : 컨텍스트 오류있는 프로젝트 완료후 #####################################
                        if numContext > 1:

                            contextFix_data = []
                            contextFix_data.append(sel_proj.lower())
                            contextFix_data.append(member)
                            self.contextFix[shot] = contextFix_data

                        ###############################################################################################

                        shot_manday = (shot, progressRate)
                        shot_manday_list.append(shot_manday)

                        # 파트 지정
                        part=""
                        if process == "matchmove":
                            part = "M"

                        elif process == "animation":
                            part = "A"

                        elif process == "creature":
                            part = "R"

                        for editTask in editedTasks_json:

                            editTask_proj = ""
                            for proj_low in self.projects:
                                if proj_low[1] == editTask[3]:
                                    editTask_proj = proj_low[1]                                    

                            if editTask[0] == shot and editTask[2] == part and editTask_proj==sel_proj:
                                act = progressRate.split("/")[0]
                                progressRate = act+"/"+editTask[1]

                        shot_manday_status = (sel_proj.upper(), shot, progressRate, status, part)
                        shot_manday_status_list.append(shot_manday_status)

                        task_schedule = [start_date, end_date]
                        task_schedule_dic[shot_manday_status] = task_schedule

                        member_shot[member] = shot_manday_status_list


                # connect_shot_schedule 와 새로 생성된 task_schedule_dic 을 합치기
                connect_shot_schedule = {**connect_shot_schedule, **task_schedule_dic}

                # connect_shot_schedule 의  태스크중 task_schedule_dic 와 태스크 이름은 같지만 맨데이나 스테이터스가 다른 태스크의 경우 
                # connect_shot_schedule 의 원래 태스크를 삭제하고 task_schedule_dic의 태스크로 대체시킴
                del_tasks = []
                add_tasks = []
                for connTask in connect_shot_schedule:
                    for editTask in task_schedule_dic:
                        if connTask[0] == editTask[0] and connTask[1] == editTask[1] and connTask[4] == editTask[4] and connTask[2] != editTask[2]:
                            del_tasks.append(connTask)
                            add_tasks.append(editTask)

                for del_task in del_tasks:
                    del connect_shot_schedule[del_task]

                for add_task in add_tasks:
                    connect_shot_schedule[add_task] = task_schedule_dic[add_task]

                # connect_shot_schedule의 태스크들중 비디 맨데이가 수정된 태스크는 수정된 맨데이를 적용
                taskList_connectSchedule = list(connect_shot_schedule.keys())
                editedTasks = [editTask for editTask in editedTasks_json for task in taskList_connectSchedule  if task[1]==editTask[0] and task[4]==editTask[2] and self.projects[task[0].lower()][1] ==  editTask[3]]

                for taskName in editedTasks:
                    scheduleTask = [scheduleTaskName for scheduleTaskName in connect_shot_schedule if scheduleTaskName[1]== taskName[0] and scheduleTaskName[4]== taskName[2] and self.projects[scheduleTaskName[0].lower()][1] ==  taskName[3]]

                    if len(scheduleTask) == 1:
                        list_scheduleTask = list(scheduleTask[0])
                        bidManday = taskName[1]
                        mandayStatus = list_scheduleTask[2].split('/')[0] + "/" + bidManday
                        list_scheduleTask[2] = mandayStatus
                        tuple_scheduleTask = tuple(list_scheduleTask)
                        schedule = connect_shot_schedule[scheduleTask[0]]
                        del connect_shot_schedule[scheduleTask[0]]
                        connect_shot_schedule[tuple_scheduleTask] = schedule

            proj_member_shots[sel_proj] = member_shot
        
        projs = list(proj_member_shots.keys())

        all_shots = []
        all_shotDic = {}
        members_proj = []
        for proj in projs:
            members_proj =  list(proj_member_shots[proj].keys())
            for member in members_proj:
                shotList_member = proj_member_shots[proj][member]
                for shot in shotList_member:
                    all_shots.append(shot)
                    all_shotDic[shot]=proj

        ###### 각 프로젝트의 어싸인된 태스크들과 해당 프로젝트를 짝지어 전역변수에 저장
        for proj in proj_member_shots:
            members_proj =  list(proj_member_shots[proj].keys())
            for member in members_proj:
                if member in list(proj_member_shots[proj].keys()):
                    for task in proj_member_shots[proj][member]:
                        connect_shot_proj[task] = proj

        ## 모든 태스크를 확인하여 bd맨데이 수정된 태스크는 이를 반영하여 다시 저장
        all_shots = self.convert_mandayEditList(all_shots) 

        # all_shotDic 내의 모든 태스크를 맨데이가 수정된 all_shots의 태스크들과 비교하여, 
        # 맨데이 수정된 태스크는 원래 태스크를 지우고 all_shots의 태스크로 대체함.
        shotDicKeys = list(all_shotDic.keys())
        for shotDic_task in shotDicKeys:
            for shot in all_shots:
                if shot[1]==shotDic_task[1] and shot[0]==shotDic_task[0] and shot[4]==shotDic_task[4]:
                    if shot[2] != shotDic_task[2]:
                        proj = all_shotDic[shotDic_task]
                        del all_shotDic[shotDic_task]
                        all_shotDic[shot] = proj


        # taskInfo에는 현재 선택되어진 멤버의 태스크정보들만 들어있으므로 이 정보로 기존 정보(체크가 해제된 멤버의 태스크정보까지 포함된)를 덮어씀
        self.file_manager.export_Json(currentPath+"/.task_info", userID+"_task_info.json", taskInfo)

        return all_shotDic




    # 팀원 업무데이타 로딩의 멀티쓰레드의 반환값을 가져오는 시그널의 슬롯
    @Slot(list)
    def loadData_result(self, results):

        QApplication.setOverrideCursor(Qt.WaitCursor)
        global connect_shot_schedule

        i=0
        for result in results:

            # connect_shot_schedule 와 새로 생성된 task_schedule_dic 을 합치기
            connect_shot_schedule = {**connect_shot_schedule, **result[2]}
            member = list(result[0].keys())[0]
            self.taskInfo[member] = result[1][member]

            for proj in result[0][member]:
                if proj not in self.proj_member_shots:
                    self.proj_member_shots[proj] = {}
                    self.proj_member_shots[proj][member] = result[0][member][proj]

                if proj in self.proj_member_shots:
                    self.proj_member_shots[proj][member] = result[0][member][proj]                

            projs = list(self.proj_member_shots.keys())

            all_shots = []
            members_proj = [] 
            for proj in projs:
                members_proj =  list(self.proj_member_shots[proj].keys())
                for member in members_proj:
                    shotList_member = self.proj_member_shots[proj][member]
                    for shot in shotList_member:
                        all_shots.append(shot)
                        self.all_shotDic[shot]=proj 


            ###### 각 프로젝트의 어싸인된 태스크들과 해당 프로젝트를 짝지어 전역변수에 저장
            for proj in self.proj_member_shots:
                members_proj =  list(self.proj_member_shots[proj].keys())
                for member in members_proj:
                    if member in list(self.proj_member_shots[proj].keys()):
                        for task in self.proj_member_shots[proj][member]:
                            connect_shot_proj[task] = proj


            ## 모든 태스크를 확인하여 bd맨데이 수정된 태스크는 이를 반영하여 다시 저장
            all_shots = self.convert_mandayEditList(all_shots) 


            # all_shotDic 내의 모든 태스크를 맨데이가 수정된 all_shots의 태스크들과 비교하여, 
            # 맨데이 수정된 태스크는 원래 태스크를 지우고 all_shots의 태스크로 대체함.
            shotDicKeys = list(self.all_shotDic.keys())
            for shotDic_task in shotDicKeys:
                for shot in all_shots:
                    if shot[1]==shotDic_task[1] and shot[0]==shotDic_task[0] and shot[4]==shotDic_task[4]:
                        if shot[2] != shotDic_task[2]:
                            proj = self.all_shotDic[shotDic_task]
                            del self.all_shotDic[shotDic_task]
                            self.all_shotDic[shot] = proj


        if self.MT_reloadShotList == 1 and self.MT_selAll_thread == 1:
            print ("(selAll)thread_win!!")

            self.MT_reloadShotList = 0
            self.MT_selAll_thread = 0

        self.file_manager.export_Json(currentPath+"/.task_info", userID+"_task_info.json", self.taskInfo)
        self.sel_assigned_proj()
        self.refereshListViews(0, jsonData)

        QApplication.restoreOverrideCursor()



    # 팀원 업무데이타 로딩을 위한 멀티쓰레딩 실행 메서드
    def process_MT_loadData(self): 

        QApplication.setOverrideCursor(Qt.WaitCursor)

        tempBid_path = currentPath + "/.temp_BDmanday"
        editedTasks_json = self.file_manager.import_Json(tempBid_path, "edited_mandayData.json")
        selectedProj = self.get_selected_projects()

        taskInfo = {}
        # 활성작업이 최대치 미만이고 대기작업이 남아있으면 추가
        while len(self.active_workers) < self.max_concurrent and self.pending_workers:

            self.proj_member_shots = {}
            conn_shot_sche = connect_shot_schedule

            i, member = self.pending_workers.pop(0)
            worker = loadData.LoadingWorkData(i, member, taskInfo, selectedProj, editedTasks_json, self.signals, conn_shot_sche, self.reload_tatic, self)
            self.active_workers.append(worker)
            self.threadpool.start(worker)

        #################################################################################################################################
        for worker in self.active_workers[:]:
            if worker.is_finished:
                self.active_workers.remove(worker)
                self.completed_tasks += 1
                self.results.append(worker.result)


        if self.completed_tasks == self.total_tasks:
            self.timer.stop()
            self.signals.finished.emit(self.results)                
        #################################################################################################################################

        QApplication.restoreOverrideCursor()





    # (멀티쓰레딩)멤버 트리뷰에서 팀원의 선택 및 해제에 맞게 샷 리스트를 업데이트
    def get_ShotlistDic_MT(self, members):

        # signals.finished 중복연결 방지
        try:
            self.signals.finished.disconnect()
        except Exception:
            pass

        # QTimer, active_workers 등 완전 초기화
        if hasattr(self, 'timer'):
            self.timer.stop()
            del self.timer

        global connect_shot_schedule
        self.all_shotDic = {}
        self.taskInfo = {}
        self.thread_finished = False

        tempBid_path = currentPath + "/.temp_BDmanday"
        editedTasks_json = self.file_manager.import_Json(tempBid_path, "edited_mandayData.json")

        selectedProj = self.get_selected_projects()

        taskInfo = {}
        proj_member_shots={}

        self.active_workers = []
        self.pending_workers = [(i, members[i]) for i in range(len(members))]
        self.completed_tasks = 0
        self.max_concurrent = 32 #동시실행 제한
        self.total_tasks = len(members)
        self.results = []
        self.signals.finished.connect(self.loadData_result)  

        self.timer = QTimer()
        self.timer.timeout.connect(self.process_MT_loadData)
        self.timer.setInterval(100)
        self.timer.start()

        return self.all_shotDic





    # 맨데이가 수정된 샷의 경우 수정된 값으로 비디 맨데이를 수정한후 샷리스트를 반환
    def convert_mandayEditList(self, shotList):

        global connect_shot_proj
        bidEdit_json = self.file_manager.import_Json(currentPath+"/.temp_BDmanday", "edited_mandayData.json")

        edited_taskList = {}
        del_tasks = []
        for task in shotList:

            list_task = list(task) # 튜플형태의 task를 리스트로 변경
            proj = task[0].lower()

            # shotList 안에서 맨데이가 수정된 태스크를 찾기 위해서 제이슨파일 확인
            for editBD_task in bidEdit_json:
                # 맨데이는 수정되었으므로, 태스크명과 파트명만 같은 태스크를 찾아 맨데이 업데이트하기
                if list_task[1] == editBD_task[0] and list_task[4] == editBD_task[2] and self.projects[proj][1] == editBD_task[3]:
                    editedBD_manday = editBD_task[1]
                    new_mandayStatus = list_task[2].split("/")[0] + "/" + editedBD_manday
                    list_task[2] = new_mandayStatus

                    # shotlist 로드될때 만들어진 connect_shot_proj 변수 내 맨데이를 업데이트 하기위하여 해당 태스크(지우고 업데이트 할) 체크
                    for conn_task in connect_shot_proj:
                        if list_task[1] == conn_task[1] and list_task[4] == conn_task[4]:
                            edited_taskList[tuple(list_task)] = connect_shot_proj[conn_task]
                            del_tasks.append(task)

        # ( 튜플이기 때문에 업데이트 할수 없고 삭제하고 새로 추가해야함 )
        # 맨데이 업데이트되어진 태스크를 찾아 업데이트 되기전 데이타 지우고 업데이트된 태스크 추가
        for delTask in del_tasks:
            if delTask in shotList:
                shotList.remove(delTask)

        for addTask in list(edited_taskList.keys()):
            shotList.append(addTask)

        # ( 튜플이기 때문에 업데이트 할수 없고 삭제하고 새로 추가해야함 )
        # connect_shot_proj 내에서 맨데이 업데이트를 위해, 이전 맨데이를 가진 태스크를 삭제위해 del_task_proj 에 모으기
        del_task_proj = []
        for delTask in edited_taskList:
            for connect_del_task in connect_shot_proj:
                # 태스크명, 파트, 프로젝트 가 같은지 확인
                if delTask[1] == connect_del_task[1] and delTask[4] == connect_del_task[4] and delTask[0] == connect_del_task[0]:
                    del_task_proj.append(connect_del_task)

        # 업데이트 되기전 이전 맨데이를 가진 태스크를 connect_shot_proj에서 삭제
        for delItem in del_task_proj:
            if delItem in connect_shot_proj:
                del connect_shot_proj[delItem]

        # 삭제후 connect_shot_proj 에 수정된 비디멘디로 업데이트된 태스크 추가
        for editBD_Task in edited_taskList:
            connect_shot_proj[editBD_Task] = edited_taskList[editBD_Task]

        return shotList




    def getTaskInfo(self, name):
        """TACTIC API를 통한 태스크 정보 조회 (TacticAPIClient 사용)"""
        return tactic_client.get_task_info(name)




def getProjectList():
    """TACTIC API를 통한 프로젝트 목록 조회 (TacticAPIClient 사용)"""
    return tactic_client.get_project_list()



def getUserInfo(user):
    """TACTIC API를 통한 사용자 정보 조회 (TacticAPIClient 사용)"""
    return tactic_client.get_user_info(user)



if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)

    app.setStyle(QtWidgets.QStyleFactory.create("Fusion"))
    app.setPalette(darkTheme.set_palette())
    app.setStyleSheet(darkTheme.set_styleSheet())

    dx_manager = DxManager()
    dx_manager.setWindowTitle("dxManager")
    dx_manager.show()
    app.exec_()

